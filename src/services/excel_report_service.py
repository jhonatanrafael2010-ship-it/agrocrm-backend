"""
excel_report_service.py — v5 (PowerBI-inspired redesign)
"""

from io import BytesIO
from datetime import date as _date, datetime, timedelta
import datetime as _dt
from collections import Counter, defaultdict

from flask import jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.formatting.rule import DataBarRule, CellIsRule
from models import db, Visit

META_VISITAS_CLIENTE = 5

# ── Palette ──────────────────────────────────────────────────────────
NAV_DARKEST    = "0F172A"   # banner background (near-black navy)
NAV_DARK       = "1E293B"   # sub-bar / footer bar
SURFACE_WHITE  = "FFFFFF"
SURFACE_F8     = "F8FAFC"   # zebra row tint
BORDER_LINE    = "E2E8F0"   # hairline separators

ACCENT_EMERALD = "10B981"   # KPI 1 / primary section accent
ACCENT_BLUE    = "3B82F6"   # KPI 2
ACCENT_PURPLE  = "8B5CF6"   # KPI 3
ACCENT_AMBER   = "F59E0B"   # KPI 4 / Atraso accent
ACCENT_DEFAULT = ACCENT_EMERALD

TEXT_DARK  = "0F172A"
TEXT_MID   = "475569"
TEXT_LIGHT = "94A3B8"

STATUS_DONE        = "16A34A"
STATUS_CANCELED    = "DC2626"
STATUS_DONE_BG     = "DCFCE7"
STATUS_PENDING_BG  = "FEF3C7"
STATUS_CANCELED_BG = "FEE2E2"

BRAND_MID = "166534"   # kept for legacy DataBarRule colour arg


def _styles():
    return {
        # Banner
        "banner_fill": PatternFill("solid", fgColor=NAV_DARKEST),
        "banner_font": Font(name="Calibri", color=SURFACE_WHITE, bold=True, size=20),
        "nav_dark_fill": PatternFill("solid", fgColor=NAV_DARK),

        # KPI card accents
        "kpi_accent_emerald": PatternFill("solid", fgColor=ACCENT_EMERALD),
        "kpi_accent_blue":    PatternFill("solid", fgColor=ACCENT_BLUE),
        "kpi_accent_purple":  PatternFill("solid", fgColor=ACCENT_PURPLE),
        "kpi_accent_amber":   PatternFill("solid", fgColor=ACCENT_AMBER),
        "kpi_body_fill":      PatternFill("solid", fgColor=SURFACE_WHITE),
        "kpi_value_font":     Font(name="Calibri", color=TEXT_DARK, bold=True, size=28),
        "kpi_label_font":     Font(name="Calibri", color=TEXT_MID, bold=False, size=9),

        # Table header
        "header_fill": PatternFill("solid", fgColor=NAV_DARKEST),
        "header_font": Font(name="Calibri", color=SURFACE_WHITE, bold=True, size=10),

        # Section title
        "section_fill": PatternFill("solid", fgColor="F0FDF4"),
        "section_font": Font(name="Calibri", color=TEXT_DARK, bold=True, size=11),

        # Table rows
        "row_font":    Font(name="Calibri", color=TEXT_DARK, size=10),
        "zebra_fill":  PatternFill("solid", fgColor=SURFACE_F8),
        "white_fill":  PatternFill("solid", fgColor=SURFACE_WHITE),
        "border_data": Border(bottom=Side(style="hair", color=BORDER_LINE)),

        # Typography
        "bold":         Font(name="Calibri", bold=True, color=TEXT_DARK),
        "muted_italic": Font(name="Calibri", color=TEXT_MID, size=9, italic=True),
        "footer":       Font(name="Calibri", color=TEXT_LIGHT, size=8, italic=True),

        # Alignments
        "center":      Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left":        Alignment(horizontal="left",   vertical="top",    wrap_text=True),
        "left_center": Alignment(horizontal="left",   vertical="center", wrap_text=True),

        # Legacy key (used in _render_visits/_render_atraso style references)
        "border_section": Border(bottom=Side(style="medium", color=ACCENT_EMERALD)),
    }


def _br_date(d):
    if not d:
        return ""
    if isinstance(d, str):
        try:
            return _date.fromisoformat(d[:10]).strftime("%d/%m/%Y")
        except Exception:
            return d[:10]
    try:
        return d.strftime("%d/%m/%Y")
    except Exception:
        return str(d)


def _translate_status(status):
    mapping = {
        "done": "Concluída", "concluida": "Concluída", "concluído": "Concluída",
        "planned": "Planejada", "planejada": "Planejada", "pendente": "Pendente",
        "canceled": "Cancelada", "cancelado": "Cancelada",
    }
    return mapping.get((status or "").lower(), status or "—")


def _has_valid_photo(visit):
    photos = getattr(visit, "photos", []) or []
    return any(getattr(p, "url", None) for p in photos)


def _truncate(text, limit=200):
    if not text:
        return ""
    text = str(text).strip()
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + "…"


def _build_consultants_map():
    consultants_map = {}
    try:
        from models import Consultant
        consultants_map = {c.id: c.name for c in Consultant.query.all()}
        if consultants_map:
            return consultants_map
    except Exception as e:
        print(f"[excel_report] Consultant.query falhou: {e}")
    try:
        from routes import CONSULTANTS
        consultants_map = {c["id"]: c["name"] for c in CONSULTANTS}
    except Exception:
        pass
    return consultants_map


def _resolve_visit_culture(visit):
    c = (visit.culture or "").strip()
    if c:
        return c
    planting = getattr(visit, "planting", None)
    if planting:
        c = (getattr(planting, "culture", None) or "").strip()
        if c:
            return c
    return ""


def _build_active_clients_for_seasons(seasons_list, region_filter=None):
    from models import Planting, Plot, Property, Client

    if not seasons_list:
        return set()

    plantings_q = (
        Planting.query
        .join(Plot, Planting.plot_id == Plot.id)
        .join(Property, Plot.property_id == Property.id)
        .join(Client, Property.client_id == Client.id)
    )
    if region_filter:
        plantings_q = plantings_q.filter(Client.region == region_filter)

    plantings = plantings_q.all()

    active_ids = set()
    for p in plantings:
        if not p.planting_date or not p.culture:
            continue
        culture_norm = (p.culture or "").strip().lower()
        for s in seasons_list:
            if culture_norm != s["culture"].strip().lower():
                continue
            try:
                s_start = _date.fromisoformat(s["start"])
                s_end = _date.fromisoformat(s["end"])
            except Exception:
                continue
            if s_start <= p.planting_date <= s_end:
                try:
                    cid = p.plot.property.client_id if p.plot and p.plot.property else None
                    if cid:
                        active_ids.add(cid)
                except Exception:
                    pass
                break

    return active_ids


def _dedupe_visits(visits_list):
    grouped = {}
    for v in visits_list:
        if not v.client_id or not v.date:
            continue
        culture = _resolve_visit_culture(v)
        key = (v.client_id, v.date, culture)
        if key not in grouped:
            grouped[key] = {
                "client_id": v.client_id,
                "date": v.date,
                "culture": culture,
                "consultant_id": v.consultant_id,
                "has_photo": _has_valid_photo(v),
                "launches": [v],
            }
        else:
            grouped[key]["launches"].append(v)
            if _has_valid_photo(v):
                grouped[key]["has_photo"] = True
            if not grouped[key]["consultant_id"]:
                grouped[key]["consultant_id"] = v.consultant_id

    return list(grouped.values())


def _week_label(d):
    iso_year, iso_week, _ = d.isocalendar()
    monday = d - timedelta(days=d.weekday())
    sunday = monday + timedelta(days=6)
    return f"{monday.strftime('%d/%m')}–{sunday.strftime('%d/%m')}", (iso_year, iso_week)


def generate_monthly_xlsx(request):
    try:
        from models import (
            Visit, Client, Property, Plot,
            get_season_by_key, AVAILABLE_SEASONS,
        )

        month = request.args.get("month")
        start = request.args.get("start")
        end = request.args.get("end")

        filter_region = (request.args.get("region") or "").strip()
        filter_season_key = (request.args.get("season") or "").strip()
        filter_consultant_id = (request.args.get("consultant") or "").strip()
        try:
            filter_consultant_id = int(filter_consultant_id) if filter_consultant_id else None
        except ValueError:
            filter_consultant_id = None

        if month:
            y, m = [int(x) for x in month.split("-")]
            start_date = _date(y, m, 1)
            next_month = _date(y + 1, 1, 1) if m == 12 else _date(y, m + 1, 1)
            end_date = next_month - _dt.timedelta(days=1)
        else:
            if not start or not end:
                return jsonify(message="Informe ?month=YYYY-MM ou ?start=YYYY-MM-DD&end=YYYY-MM-DD"), 400
            start_date = _date.fromisoformat(start)
            end_date = _date.fromisoformat(end)

        season = get_season_by_key(filter_season_key) if filter_season_key else None
        if season:
            season_start = _date.fromisoformat(season["start"])
            season_end = _date.fromisoformat(season["end"])
            start_date = max(start_date, season_start)
            end_date = min(end_date, season_end)
            seasons_for_carteira = [season]
            season_culture = season["culture"]
        else:
            seasons_for_carteira = AVAILABLE_SEASONS
            season_culture = None

        carteira_ids = _build_active_clients_for_seasons(
            seasons_for_carteira,
            region_filter=filter_region or None,
        )

        visits_clients_q = (
            db.session.query(Visit.client_id)
            .filter(Visit.date >= start_date)
            .filter(Visit.date <= end_date)
            .filter(Visit.client_id.isnot(None))
        )
        if filter_consultant_id:
            visits_clients_q = visits_clients_q.filter(Visit.consultant_id == filter_consultant_id)

        visits_clients_ids = {row[0] for row in visits_clients_q.distinct().all()}
        effective_client_ids = carteira_ids | visits_clients_ids
        total_clients = len(carteira_ids)

        if effective_client_ids:
            visits_query = (
                Visit.query
                .filter(Visit.date >= start_date)
                .filter(Visit.date <= end_date)
                .filter(Visit.client_id.in_(effective_client_ids))
            )
            if filter_consultant_id:
                visits_query = visits_query.filter(Visit.consultant_id == filter_consultant_id)
            visits_raw = visits_query.order_by(Visit.date.asc().nullslast()).all()
        else:
            visits_raw = []

        if season_culture:
            sc_norm = season_culture.strip().lower()
            visits_raw = [
                v for v in visits_raw
                if _resolve_visit_culture(v).strip().lower() == sc_norm
            ]

        unique_visits = _dedupe_visits(visits_raw)

        client_ids_in_visits = sorted({v["client_id"] for v in unique_visits if v["client_id"]})
        prop_ids = sorted({l.property_id for v in unique_visits for l in v["launches"] if l.property_id})
        plot_ids = sorted({l.plot_id for v in unique_visits for l in v["launches"] if l.plot_id})

        all_client_ids = sorted(carteira_ids | set(client_ids_in_visits))
        total_clients = len(all_client_ids)
        clients_map = (
            {c.id: c.name for c in Client.query.filter(Client.id.in_(all_client_ids)).all()}
            if all_client_ids else {}
        )
        props_map = (
            {p.id: p.name for p in Property.query.filter(Property.id.in_(prop_ids)).all()}
            if prop_ids else {}
        )
        plots_map = (
            {pl.id: pl.name for pl in Plot.query.filter(Plot.id.in_(plot_ids)).all()}
            if plot_ids else {}
        )
        consultants_map = _build_consultants_map()

        total_visits_unique = len(unique_visits)
        visits_with_photo = sum(1 for v in unique_visits if v["has_photo"])
        total_launches_with_photo = sum(
            1 for v in unique_visits for l in v["launches"] if _has_valid_photo(l)
        )
        unique_clients_attended = len({v["client_id"] for v in unique_visits if v["has_photo"]})
        coverage = (unique_clients_attended / total_clients) if total_clients else 0

        photo_visits_by_client = Counter()
        for v in unique_visits:
            if v["has_photo"]:
                photo_visits_by_client[v["client_id"]] += 1

        clients_in_target = sum(
            1 for cnt in photo_visits_by_client.values() if cnt >= META_VISITAS_CLIENTE
        )

        meta_total = total_clients * META_VISITAS_CLIENTE
        target_pct = (visits_with_photo / meta_total) if meta_total else 0

        filter_labels = []
        if filter_region:
            filter_labels.append(f"Região: {filter_region}")
        if season:
            filter_labels.append(f"Safra: {season['label']} ({season['culture']})")
        filters_applied = " • ".join(filter_labels) if filter_labels else "Carteira completa (todas as safras conhecidas)"

        wb = Workbook()
        wb.remove(wb.active)
        ws_dash = wb.create_sheet("Dashboard")
        ws_visits = wb.create_sheet("Visitas")
        ws_atraso = wb.create_sheet("Atraso")
        ws_prods = wb.create_sheet("Produtos")

        period_label = f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"

        _render_dashboard(
            ws_dash, unique_visits, period_label, filters_applied,
            total_visits_unique, visits_with_photo, total_launches_with_photo,
            total_clients,
            coverage, clients_in_target, target_pct, meta_total,
            unique_clients_attended,
            photo_visits_by_client, clients_map, consultants_map,
            set(all_client_ids)
        )
        _render_visits(
            ws_visits, visits_raw, period_label,
            total_launches_with_photo, unique_clients_attended,
            clients_map, props_map, plots_map, consultants_map,
            filters_applied
        )
        _render_atraso(
            ws_atraso, total_clients, photo_visits_by_client, clients_map,
            period_label, filters_applied, set(all_client_ids)
        )
        _render_products(
            ws_prods, visits_raw, period_label,
            clients_map, props_map, consultants_map,
            filters_applied
        )

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"relatorio_visitas_{start_date.isoformat()}_a_{end_date.isoformat()}.xlsx"
        return send_file(
            bio,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        import traceback
        print("⚠️ erro generate_monthly_xlsx:", e)
        traceback.print_exc()
        return jsonify(message=f"Erro ao gerar relatório: {e}"), 500


# ── Visual helpers ────────────────────────────────────────────────────

def _kpi_card(ws, col1, col2, row, title, value, fmt, s, accent_fill=None):
    """PowerBI-style KPI tile: colored top accent → label → big number → pad."""
    if accent_fill is None:
        accent_fill = s["kpi_accent_emerald"]
    c1 = ord(col1) - 64
    c2 = ord(col2) - 64

    # Row N — accent color bar
    ws.merge_cells(f"{col1}{row}:{col2}{row}")
    for cc in range(c1, c2 + 1):
        ws.cell(row, cc).fill = accent_fill

    # Row N+1 — label (small, muted)
    ws.merge_cells(f"{col1}{row+1}:{col2}{row+1}")
    lbl = ws[f"{col1}{row+1}"]
    lbl.value = title
    lbl.fill = s["kpi_body_fill"]
    lbl.font = s["kpi_label_font"]
    lbl.alignment = Alignment(horizontal="center", vertical="bottom")

    # Row N+2 — big number
    ws.merge_cells(f"{col1}{row+2}:{col2}{row+2}")
    val = ws[f"{col1}{row+2}"]
    val.value = value
    val.fill = s["kpi_body_fill"]
    val.font = s["kpi_value_font"]
    val.alignment = Alignment(horizontal="center", vertical="center")
    if fmt:
        val.number_format = fmt

    # Row N+3 — bottom padding
    ws.merge_cells(f"{col1}{row+3}:{col2}{row+3}")
    for cc in range(c1, c2 + 1):
        ws.cell(row + 3, cc).fill = s["kpi_body_fill"]

    # Subtle right border to separate cards
    right_b = Border(right=Side(style="thin", color=BORDER_LINE))
    for rr in range(row, row + 4):
        ws.cell(rr, c2).border = right_b


def _section_title_premium(ws, cell_ref, text, merge_range, s,
                            accent_color=None, bg_color="F0FDF4"):
    """Section header with thick left accent border — PowerBI style."""
    if accent_color is None:
        accent_color = ACCENT_EMERALD
    ws.merge_cells(merge_range)
    cell = ws[cell_ref]
    cell.value = text
    cell.font = s["section_font"]
    cell.fill = PatternFill("solid", fgColor=bg_color)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    cell.border = Border(
        left=Side(style="thick", color=accent_color),
        bottom=Side(style="thin", color=BORDER_LINE),
    )


def _style_header_row(ws, row, col_start, col_end, s):
    for col in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = s["header_fill"]
        cell.font = s["header_font"]
        cell.alignment = s["center"]


def _render_dashboard(
    ws, unique_visits, period_label, filters_applied,
    total_visits_unique, visits_with_photo, total_launches_with_photo,
    total_clients,
    coverage, clients_in_target, target_pct, meta_total,
    unique_clients_attended,
    photo_visits_by_client, clients_map, consultants_map,
    effective_client_ids
):
    s = _styles()
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 100

    for col in range(1, 13):
        ws.column_dimensions[get_column_letter(col)].width = 17

    # ── Banner ────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 52
    ws.merge_cells("A2:L2")
    ws["A2"] = "PAINEL GERENCIAL — NutriCRM"
    ws["A2"].fill = s["banner_fill"]
    ws["A2"].font = s["banner_font"]
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[3].height = 3
    ws.merge_cells("A3:L3")
    ws["A3"].fill = s["nav_dark_fill"]

    ws.row_dimensions[4].height = 20
    ws.merge_cells("A4:L4")
    ws["A4"] = (
        f"Período: {period_label}    •    "
        f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    )
    ws["A4"].font = Font(name="Calibri Light", color=TEXT_MID, size=10, italic=True)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[5].height = 18
    ws.merge_cells("A5:L5")
    ws["A5"] = f"Filtros: {filters_applied}"
    ws["A5"].font = Font(name="Calibri", color=ACCENT_EMERALD, size=10, bold=True)
    ws["A5"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[6].height = 16
    ws.merge_cells("A6:L6")
    ws["A6"] = (
        f"Regra: visita concluída = visita com foto. "
        f"Lançamentos no mesmo (cliente, dia, cultura) contam como 1 visita.    "
        f"Meta = {META_VISITAS_CLIENTE} visitas/cliente."
    )
    ws["A6"].font = Font(name="Calibri", color=TEXT_LIGHT, size=9, italic=True)
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")

    # ── KPI cards (rows 8–11) ─────────────────────────────────────────
    ws.row_dimensions[7].height = 10   # gutter
    ws.row_dimensions[8].height = 6    # accent bar
    ws.row_dimensions[9].height = 18   # label
    ws.row_dimensions[10].height = 40  # value
    ws.row_dimensions[11].height = 10  # bottom pad

    _kpi_card(ws, "A", "C",  8, "VISITAS REALIZADAS",
              visits_with_photo, "#,##0", s, s["kpi_accent_emerald"])
    _kpi_card(ws, "D", "F",  8, "LANÇAMENTOS",
              total_launches_with_photo, "#,##0", s, s["kpi_accent_blue"])
    _kpi_card(ws, "G", "I",  8, f"NA META ({META_VISITAS_CLIENTE}+)",
              clients_in_target, "#,##0", s, s["kpi_accent_purple"])
    _kpi_card(ws, "J", "L",  8, "% DA META",
              target_pct, "0.0%", s, s["kpi_accent_amber"])

    # ── Summary stats bar (row 13) ────────────────────────────────────
    ws.row_dimensions[12].height = 8
    ws.row_dimensions[13].height = 20
    ws.merge_cells("A13:L13")
    ws["A13"] = (
        f"Carteira ativa: {total_clients} clientes  •  "
        f"Atendidos: {unique_clients_attended}  •  "
        f"Cobertura: {coverage:.1%}  •  "
        f"Meta total: {meta_total} visitas ({total_clients} × {META_VISITAS_CLIENTE})"
    )
    ws["A13"].font = Font(name="Calibri", color=TEXT_MID, size=10)
    ws["A13"].fill = PatternFill("solid", fgColor=SURFACE_F8)
    ws["A13"].alignment = Alignment(horizontal="center", vertical="center")

    # ── Three tables side by side (rows 15+) ──────────────────────────
    ws.row_dimensions[14].height = 14
    ws.row_dimensions[15].height = 24
    ws.row_dimensions[16].height = 22

    _section_title_premium(ws, "A15", "VISITAS POR CONSULTOR",
                            "A15:C15", s,
                            accent_color=ACCENT_EMERALD, bg_color="F0FDF4")
    ws["A16"] = "Consultor"
    ws["B16"] = "Visitas"
    ws["C16"] = "% do total"
    _style_header_row(ws, 16, 1, 3, s)

    cons_counts = Counter(
        v["consultant_id"] for v in unique_visits
        if v["consultant_id"] and v["has_photo"]
    )
    r = 17
    for cid, cnt in cons_counts.most_common():
        ws[f"A{r}"] = consultants_map.get(cid, f"Consultor #{cid}")
        ws[f"B{r}"] = cnt
        ws[f"C{r}"] = (cnt / visits_with_photo) if visits_with_photo else 0
        ws[f"C{r}"].number_format = "0.0%"
        ws.row_dimensions[r].height = 22
        for col in range(1, 4):
            ws.cell(r, col).border = s["border_data"]
            ws.cell(r, col).font = s["row_font"]
            ws.cell(r, col).alignment = s["left_center"] if col == 1 else s["center"]
            if (r - 17) % 2 == 1:
                ws.cell(r, col).fill = s["zebra_fill"]
        r += 1

    _section_title_premium(ws, "E15", "VISITAS POR CULTURA",
                            "E15:G15", s,
                            accent_color=ACCENT_BLUE, bg_color="EFF6FF")
    ws["E16"] = "Cultura"
    ws["F16"] = "Visitas"
    ws["G16"] = "% do total"
    _style_header_row(ws, 16, 5, 7, s)

    cult_counts = Counter()
    for v in unique_visits:
        if not v["has_photo"]:
            continue
        cult = (v["culture"] or "—").strip() or "—"
        cult_counts[cult] += 1

    r = 17
    for culture, cnt in cult_counts.most_common():
        ws[f"E{r}"] = culture
        ws[f"F{r}"] = cnt
        ws[f"G{r}"] = (cnt / visits_with_photo) if visits_with_photo else 0
        ws[f"G{r}"].number_format = "0.0%"
        ws.row_dimensions[r].height = 22
        for col in range(5, 8):
            ws.cell(r, col).border = s["border_data"]
            ws.cell(r, col).font = s["row_font"]
            ws.cell(r, col).alignment = s["left_center"] if col == 5 else s["center"]
            if (r - 17) % 2 == 1:
                ws.cell(r, col).fill = s["zebra_fill"]
        r += 1

    _section_title_premium(ws, "I15", "TOP 5 CLIENTES (CONCLUÍDAS)",
                            "I15:K15", s,
                            accent_color=ACCENT_PURPLE, bg_color="FAF5FF")
    ws["I16"] = "Cliente"
    ws["J16"] = "Visitas"
    ws["K16"] = "% da meta"
    _style_header_row(ws, 16, 9, 11, s)

    top5 = photo_visits_by_client.most_common(5)
    r = 17
    for cid, cnt in top5:
        ws[f"I{r}"] = clients_map.get(cid, f"Cliente {cid}")
        ws[f"J{r}"] = cnt
        ws[f"K{r}"] = min(cnt / META_VISITAS_CLIENTE, 1.0)
        ws[f"K{r}"].number_format = "0%"
        ws.row_dimensions[r].height = 22
        for col in range(9, 12):
            ws.cell(r, col).border = s["border_data"]
            ws.cell(r, col).font = s["row_font"]
            ws.cell(r, col).alignment = s["left_center"] if col == 9 else s["center"]
            if (r - 17) % 2 == 1:
                ws.cell(r, col).fill = s["zebra_fill"]
        r += 1

    # ── Weekly chart ─────────────────────────────────────────────────
    week_buckets = defaultdict(lambda: {"label": "", "key": None, "count": 0})
    for v in unique_visits:
        if not v["date"] or not v["has_photo"]:
            continue
        label, key = _week_label(v["date"])
        bucket = week_buckets[key]
        bucket["label"] = label
        bucket["key"] = key
        bucket["count"] += 1

    weeks_sorted = sorted(week_buckets.values(), key=lambda x: x["key"])

    section_row_days = 27
    ws.row_dimensions[section_row_days].height = 24
    ws.row_dimensions[section_row_days + 1].height = 22
    _section_title_premium(
        ws, f"A{section_row_days}",
        "EVOLUÇÃO SEMANAL DE VISITAS",
        f"A{section_row_days}:B{section_row_days}",
        s, accent_color=ACCENT_BLUE, bg_color="EFF6FF"
    )
    ws[f"A{section_row_days+1}"] = "Semana"
    ws[f"B{section_row_days+1}"] = "Visitas"
    _style_header_row(ws, section_row_days + 1, 1, 2, s)

    r = section_row_days + 2
    for w in weeks_sorted:
        ws[f"A{r}"] = w["label"]
        ws[f"B{r}"] = w["count"]
        ws.row_dimensions[r].height = 22
        for col in (1, 2):
            ws.cell(r, col).border = s["border_data"]
            ws.cell(r, col).font = s["row_font"]
            ws.cell(r, col).alignment = s["center"]
            if (r - (section_row_days + 2)) % 2 == 1:
                ws.cell(r, col).fill = s["zebra_fill"]
        r += 1
    end_weeks_row = r - 1

    if end_weeks_row > section_row_days + 1:
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Visitas por semana"
        chart.y_axis.title = "Quantidade"
        chart.x_axis.title = None
        chart.legend = None
        chart.height = 11
        chart.width = 22

        data = Reference(ws, min_col=2, min_row=section_row_days + 1,
                         max_col=2, max_row=end_weeks_row)
        cats = Reference(ws, min_col=1, min_row=section_row_days + 2,
                         max_col=1, max_row=end_weeks_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        if chart.series:
            gp = GraphicalProperties(solidFill=ACCENT_EMERALD)
            gp.line.solidFill = "059669"
            chart.series[0].graphicalProperties = gp

        ws.add_chart(chart, "D" + str(section_row_days))

    # ── Client progress table ─────────────────────────────────────────
    progresso_row = max(end_weeks_row, section_row_days + 22) + 3

    ws.row_dimensions[progresso_row].height = 24
    ws.row_dimensions[progresso_row + 1].height = 22
    _section_title_premium(
        ws, f"A{progresso_row}",
        "PROGRESSO DA META POR CLIENTE",
        f"A{progresso_row}:L{progresso_row}",
        s, accent_color=ACCENT_AMBER, bg_color="FFFBEB"
    )

    ws[f"A{progresso_row+1}"] = "Cliente"
    ws[f"B{progresso_row+1}"] = "Concluídas"
    ws[f"C{progresso_row+1}"] = "% da meta (5)"
    ws[f"D{progresso_row+1}"] = "Progresso"
    _style_header_row(ws, progresso_row + 1, 1, 4, s)

    r = progresso_row + 2
    all_meta_ids = set(effective_client_ids) | set(photo_visits_by_client.keys())
    for cid in sorted(all_meta_ids, key=lambda k: -photo_visits_by_client.get(k, 0)):
        cnt = photo_visits_by_client.get(cid, 0)
        ws[f"A{r}"] = clients_map.get(cid, f"Cliente {cid}")
        ws[f"B{r}"] = cnt
        pct = min(cnt / META_VISITAS_CLIENTE, 1.0)
        ws[f"C{r}"] = pct
        ws[f"D{r}"] = pct
        ws[f"C{r}"].number_format = "0%"
        ws[f"D{r}"].number_format = "0%"
        ws.row_dimensions[r].height = 22
        for col in range(1, 5):
            ws.cell(r, col).border = s["border_data"]
            ws.cell(r, col).font = s["row_font"]
            ws.cell(r, col).alignment = s["left_center"] if col == 1 else s["center"]
            if (r - (progresso_row + 2)) % 2 == 1:
                ws.cell(r, col).fill = s["zebra_fill"]
        r += 1
    end_meta_row = r - 1

    if end_meta_row >= progresso_row + 2:
        rule = DataBarRule(
            start_type="num", start_value=0,
            end_type="num", end_value=1,
            color=ACCENT_EMERALD, showValue=False,
        )
        ws.conditional_formatting.add(f"D{progresso_row+2}:D{end_meta_row}", rule)

        ws.conditional_formatting.add(
            f"B{progresso_row+2}:B{end_meta_row}",
            CellIsRule(operator="lessThanOrEqual", formula=["1"],
                       fill=PatternFill("solid", fgColor=STATUS_CANCELED_BG),
                       font=Font(color=STATUS_CANCELED, bold=True))
        )
        ws.conditional_formatting.add(
            f"B{progresso_row+2}:B{end_meta_row}",
            CellIsRule(operator="between", formula=["2", "4"],
                       fill=PatternFill("solid", fgColor=STATUS_PENDING_BG),
                       font=Font(color="92400E", bold=True))
        )
        ws.conditional_formatting.add(
            f"B{progresso_row+2}:B{end_meta_row}",
            CellIsRule(operator="greaterThanOrEqual", formula=["5"],
                       fill=PatternFill("solid", fgColor=STATUS_DONE_BG),
                       font=Font(color=STATUS_DONE, bold=True))
        )

    # ── Footer ────────────────────────────────────────────────────────
    footer_row = end_meta_row + 3
    ws.row_dimensions[footer_row].height = 3
    ws.merge_cells(f"A{footer_row}:L{footer_row}")
    ws[f"A{footer_row}"].fill = s["nav_dark_fill"]

    ws.merge_cells(f"A{footer_row+1}:L{footer_row+1}")
    ws[f"A{footer_row+1}"] = "NutriCRM  •  Documento gerencial  •  Confidencial"
    ws[f"A{footer_row+1}"].font = s["footer"]
    ws[f"A{footer_row+1}"].alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 4
    ws.column_dimensions["I"].width = 22
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 14
    ws.column_dimensions["L"].width = 14


def _render_visits(ws, visits_raw, period_label, total_lancamentos, unique_clients,
                   clients_map, props_map, plots_map, consultants_map,
                   filters_applied):
    s = _styles()
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 48
    ws.merge_cells("A1:L1")
    ws["A1"] = "RELATÓRIO DE VISITAS TÉCNICAS"
    ws["A1"].fill = s["banner_fill"]
    ws["A1"].font = s["banner_font"]
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 3
    ws.merge_cells("A2:L2")
    ws["A2"].fill = s["nav_dark_fill"]

    ws.row_dimensions[3].height = 22
    ws["A3"] = "Período:"
    ws["A3"].font = s["bold"]
    ws["B3"] = period_label
    ws["B3"].font = s["row_font"]
    ws["D3"] = "Lançamentos:"
    ws["D3"].font = s["bold"]
    ws["E3"] = total_lancamentos
    ws["E3"].font = s["row_font"]
    ws["G3"] = "Clientes atendidos:"
    ws["G3"].font = s["bold"]
    ws["H3"] = unique_clients
    ws["H3"].font = s["row_font"]

    ws.row_dimensions[4].height = 20
    ws.merge_cells("A4:L4")
    ws["A4"] = f"Filtros: {filters_applied}"
    ws["A4"].font = Font(name="Calibri", color=ACCENT_EMERALD, size=10, bold=True)
    ws["A4"].alignment = Alignment(horizontal="left", vertical="center")

    headers = [
        "Data", "Cliente", "Propriedade", "Talhão", "Consultor",
        "Cultura", "Variedade", "Fenologia", "Status", "Foto",
        "Dias plantio", "Observações",
    ]
    header_row = 6
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i).value = h
    _style_header_row(ws, header_row, 1, len(headers), s)
    ws.row_dimensions[header_row].height = 26
    ws.freeze_panes = "A7"

    row_idx = header_row
    for v in visits_raw:
        row_idx += 1
        ws.row_dimensions[row_idx].height = 18

        client_name = clients_map.get(v.client_id, "—") if v.client_id else "—"
        prop_name = props_map.get(v.property_id, "—") if v.property_id else "—"
        plot_name = plots_map.get(v.plot_id, "—") if v.plot_id else "—"
        cons_name = consultants_map.get(v.consultant_id, f"#{v.consultant_id}" if v.consultant_id else "—")

        culture = _resolve_visit_culture(v) or "—"
        variety = v.variety or (v.planting.variety if getattr(v, "planting", None) else "—")

        dias_plantio = "—"
        planting = getattr(v, "planting", None)
        planting_date = getattr(planting, "planting_date", None) if planting else None
        if planting_date and v.date:
            try:
                dias_plantio = (v.date - planting_date).days
            except Exception:
                dias_plantio = "—"

        has_photo = "Sim" if _has_valid_photo(v) else "Não"
        status_pt = _translate_status(v.status)
        obs = _truncate(v.recommendation, 200)

        row_values = [
            _br_date(v.date), client_name, prop_name, plot_name, cons_name,
            culture, variety, v.fenologia_real or "—", status_pt, has_photo,
            dias_plantio, obs,
        ]
        for i, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=i)
            cell.value = val
            cell.border = s["border_data"]
            cell.font = s["row_font"]
            cell.alignment = s["center"] if i in (1, 9, 10, 11) else s["left"]
            if (row_idx - header_row) % 2 == 0:
                cell.fill = s["zebra_fill"]

        status_cell = ws.cell(row=row_idx, column=9)
        if status_pt == "Concluída":
            status_cell.fill = PatternFill("solid", fgColor=STATUS_DONE_BG)
            status_cell.font = Font(name="Calibri", color=STATUS_DONE, bold=True, size=10)
        elif status_pt == "Cancelada":
            status_cell.fill = PatternFill("solid", fgColor=STATUS_CANCELED_BG)
            status_cell.font = Font(name="Calibri", color=STATUS_CANCELED, bold=True, size=10)
        elif status_pt in ("Planejada", "Pendente"):
            status_cell.fill = PatternFill("solid", fgColor=STATUS_PENDING_BG)
            status_cell.font = Font(name="Calibri", color="92400E", bold=True, size=10)

    if row_idx > header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{row_idx}"

    widths = [12, 28, 22, 18, 18, 12, 18, 14, 14, 8, 12, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _render_atraso(ws, total_clients, photo_visits_by_client, clients_map,
                   period_label, filters_applied, effective_client_ids):
    s = _styles()
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 48
    ws.merge_cells("A1:E1")
    ws["A1"] = "CLIENTES EM ATRASO — AÇÃO PRIORITÁRIA"
    ws["A1"].fill = s["banner_fill"]
    ws["A1"].font = s["banner_font"]
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 3
    ws.merge_cells("A2:E2")
    ws["A2"].fill = s["nav_dark_fill"]

    ws.row_dimensions[3].height = 18
    ws.merge_cells("A3:E3")
    ws["A3"] = (
        f"Período: {period_label}    •    "
        f"Meta: {META_VISITAS_CLIENTE} visitas/cliente"
    )
    ws["A3"].font = s["muted_italic"]
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[4].height = 20
    ws.merge_cells("A4:E4")
    ws["A4"] = f"Filtros: {filters_applied}"
    ws["A4"].font = Font(name="Calibri", color=ACCENT_AMBER, size=10, bold=True)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

    headers = ["Cliente", "Concluídas", "Faltam", "% da meta", "Prioridade"]
    header_row = 6
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i).value = h
    _style_header_row(ws, header_row, 1, 5, s)
    ws.row_dimensions[header_row].height = 26
    ws.freeze_panes = "A7"

    all_clients = []
    all_meta_ids = set(effective_client_ids) | set(photo_visits_by_client.keys())
    for cid in all_meta_ids:
        name = clients_map.get(cid, f"Cliente {cid}")
        cnt = photo_visits_by_client.get(cid, 0)
        if cnt < META_VISITAS_CLIENTE:
            all_clients.append((cid, name, cnt))

    all_clients.sort(key=lambda x: x[2])

    r = header_row + 1
    for cid, name, cnt in all_clients:
        falta = META_VISITAS_CLIENTE - cnt
        pct = cnt / META_VISITAS_CLIENTE
        prioridade = "ALTA" if cnt <= 1 else ("MÉDIA" if cnt <= 3 else "BAIXA")

        ws.cell(r, 1).value = name
        ws.cell(r, 2).value = cnt
        ws.cell(r, 3).value = falta
        ws.cell(r, 4).value = pct
        ws.cell(r, 4).number_format = "0%"
        ws.cell(r, 5).value = prioridade
        ws.row_dimensions[r].height = 20

        for col in range(1, 6):
            ws.cell(r, col).border = s["border_data"]
            ws.cell(r, col).font = s["row_font"]
            ws.cell(r, col).alignment = s["center"] if col != 1 else s["left_center"]
            if (r - (header_row + 1)) % 2 == 1:
                ws.cell(r, col).fill = s["zebra_fill"]

        prio_cell = ws.cell(r, 5)
        if prioridade == "ALTA":
            prio_cell.fill = PatternFill("solid", fgColor=STATUS_CANCELED_BG)
            prio_cell.font = Font(name="Calibri", color=STATUS_CANCELED, bold=True, size=10)
        elif prioridade == "MÉDIA":
            prio_cell.fill = PatternFill("solid", fgColor=STATUS_PENDING_BG)
            prio_cell.font = Font(name="Calibri", color="92400E", bold=True, size=10)
        else:
            prio_cell.fill = PatternFill("solid", fgColor=STATUS_DONE_BG)
            prio_cell.font = Font(name="Calibri", color=STATUS_DONE, bold=True, size=10)

        r += 1

    if r > header_row + 1:
        ws.auto_filter.ref = f"A{header_row}:E{r-1}"

    widths = [34, 14, 12, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _render_products(ws, visits_raw, period_label,
                     clients_map, props_map, consultants_map,
                     filters_applied):
    s = _styles()
    ws.sheet_view.showGridLines = False

    ws.row_dimensions[1].height = 48
    ws.merge_cells("A1:I1")
    ws["A1"] = "PRODUTOS APLICADOS"
    ws["A1"].fill = s["banner_fill"]
    ws["A1"].font = s["banner_font"]
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 3
    ws.merge_cells("A2:I2")
    ws["A2"].fill = s["nav_dark_fill"]

    ws.row_dimensions[3].height = 18
    ws.merge_cells("A3:I3")
    ws["A3"] = f"Período: {period_label}"
    ws["A3"].font = s["muted_italic"]
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[4].height = 20
    ws.merge_cells("A4:I4")
    ws["A4"] = f"Filtros: {filters_applied}"
    ws["A4"].font = Font(name="Calibri", color=ACCENT_BLUE, size=10, bold=True)
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "Data", "Cliente", "Propriedade", "Cultura", "Fenologia",
        "Consultor", "Produto", "Dose / Unidade", "Data aplicação",
    ]
    header_row = 6
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i).value = h
    _style_header_row(ws, header_row, 1, len(headers), s)
    ws.row_dimensions[header_row].height = 26
    ws.freeze_panes = "A7"

    r = header_row + 1
    has_data = False
    for v in visits_raw:
        produtos = getattr(v, "products", None) or []
        for p in produtos:
            has_data = True
            client_name = clients_map.get(v.client_id, "—") if v.client_id else "—"
            prop_name = props_map.get(v.property_id, "—") if v.property_id else "—"
            cons_name = consultants_map.get(v.consultant_id, "—") if v.consultant_id else "—"
            culture = _resolve_visit_culture(v) or "—"
            fenologia = v.fenologia_real or "—"

            dose_unidade = ""
            try:
                dose = getattr(p, "dose", "") or ""
                unit = getattr(p, "unit", "") or ""
                dose_unidade = f"{dose} {unit}".strip()
            except Exception:
                pass

            row_values = [
                _br_date(v.date), client_name, prop_name,
                culture, fenologia, cons_name,
                getattr(p, "product_name", "—"), dose_unidade,
                _br_date(getattr(p, "application_date", None) or v.date),
            ]
            ws.row_dimensions[r].height = 18
            for i, val in enumerate(row_values, start=1):
                cell = ws.cell(row=r, column=i)
                cell.value = val
                cell.border = s["border_data"]
                cell.font = s["row_font"]
                cell.alignment = s["center"] if i in (1, 4, 5, 8, 9) else s["left_center"]
                if (r - header_row - 1) % 2 == 0:
                    cell.fill = s["zebra_fill"]
            r += 1

    if not has_data:
        ws.merge_cells(start_row=header_row + 2, start_column=1,
                       end_row=header_row + 2, end_column=9)
        empty_cell = ws.cell(row=header_row + 2, column=1)
        empty_cell.value = "Nenhum produto registrado neste período."
        empty_cell.font = Font(name="Calibri", color=TEXT_MID, size=10, italic=True)
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[header_row + 2].height = 36

    if r > header_row + 1:
        ws.auto_filter.ref = f"A{header_row}:I{r-1}"

    widths = [12, 26, 22, 12, 14, 18, 22, 18, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
