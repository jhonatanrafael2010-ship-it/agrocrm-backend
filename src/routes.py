# =========================
# Python (stdlib)
# =========================
from datetime import datetime, date, timedelta
import calendar
import os
import re
import uuid
import tempfile
import unicodedata
import datetime
from io import BytesIO
from urllib.request import Request, urlopen
from datetime import date as _date, datetime as _dt, timedelta as _timedelta

# =========================
# Third-party
# =========================
import jwt
import requests
from PIL import Image as PILImage
from PIL import ImageFile, ImageOps
from sqlalchemy import text
from werkzeug.utils import secure_filename
from flask_cors import cross_origin

# PIL config
ImageFile.LOAD_TRUNCATED_IMAGES = True

# =========================
# Flask
# =========================
from flask import (
    Blueprint,
    current_app,
    jsonify,
    render_template_string,
    request,
    send_file,
)

# =========================
# XML / HTML utils
# =========================
from xml.sax.saxutils import escape as xml_escape
from html import escape as html_escape

# =========================
# ReportLab (PDF)
# =========================
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT, TA_JUSTIFY
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.pdfgen.canvas import Canvas
from reportlab.platypus import (
    Flowable,
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from reportlab.graphics import renderPDF
from reportlab.graphics.barcode import qr
from reportlab.graphics.shapes import Drawing, String, Rect

# =========================
# OpenPyXL (Excel)
# =========================
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import DataBarRule

# =========================
# App / Models / Utils
# =========================
from models import (
    db,
    User,
    Client,
    Property,
    Plot,
    Visit,
    Planting,
    Opportunity,
    Photo,
    PhenologyStage,
    Variety,
    Culture,
    WhatsAppContactBinding,
    WhatsAppInboundMessage,
    ChatbotConversationState,
    TelegramContactBinding,
    Consultant,
    CONSULTANTS,
    VisitProduct,
)
from utils.r2_client import get_r2_client

import json
from difflib import SequenceMatcher

from services.chatbot_service import (
    ChatbotService,
    parse_chatbot_message,
    send_telegram_message,
    send_telegram_document,
)
import io
import subprocess
from openai import OpenAI
from zoneinfo import ZoneInfo
from pathlib import Path
import random






# =====================================================
# 🔒 LIMITES DE SEGURANÇA (EVITA ESTOURO DE MEMÓRIA)
# =====================================================
MAX_VISITS = 12        # máximo de visitas no PDF cumulativo
MAX_PHOTOS_V = 6       # máximo de fotos por visita






bp = Blueprint('api', __name__, url_prefix='/api')

UPLOAD_DIR = os.environ.get("UPLOAD_DIR", "/opt/render/project/src/uploads")

def normalize_phone_number(phone: str) -> str:
    if not phone:
        return ""
    phone = re.sub(r"\D", "", phone)
    if not phone.startswith("55"):
        phone = f"55{phone}"
    return phone


def _normalize_text(text: str) -> str:
    """
    Remove acentos, deixa minúsculo e limpa espaços extras.
    """
    if not text:
        return ""
    text = text.strip().lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(ch for ch in text if unicodedata.category(ch) != "Mn")
    text = re.sub(r"\s+", " ", text)
    return text

def parse_optional_float(value):
    if value in (None, "", "null"):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None

def normalize_lookup_text(value: str) -> str:
    if not value:
        return ""
    value = unicodedata.normalize("NFD", value.strip().lower())
    return "".join(ch for ch in value if unicodedata.category(ch) != "Mn")

def _safe_date(year: int, month: int, day: int) -> date | None:
    """
    Cria data de forma segura. Se inválida, retorna None.
    """
    try:
        return date(year, month, day)
    except ValueError:
        return None


def format_date_br(dt: date | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%d/%m/%Y")


def _last_day_of_month(year: int, month: int) -> int:
    return calendar.monthrange(year, month)[1]


def parse_human_date(text: str, base_date: date | None = None) -> date | None:
    """
    Interpreta datas em linguagem humana para uso no bot.

    Aceita exemplos:
    - hoje
    - amanha / amanhã
    - ontem
    - anteontem
    - 2 dias atras / 2 dias atrás
    - ha 3 dias / há 3 dias
    - semana passada
    - semana retrasada
    - 15
    - 24/02
    - 24/02/2026
    - 2026-02-24

    Retorna:
    - datetime.date quando conseguir interpretar
    - None quando não entender
    """
    if not text:
        return None

    today = base_date or get_local_today()
    normalized = _normalize_text(text)

    # Remove pontuação lateral comum
    normalized = normalized.strip(" .,!?:;")

    # -----------------------------
    # Casos exatos simples
    # -----------------------------
    simple_map = {
        "hoje": 0,
        "amanha": 1,
        "ontem": -1,
        "anteontem": -2,
        "semana passada": -7,
        "semana retrasada": -14,
    }

    if normalized in simple_map:
        return today + timedelta(days=simple_map[normalized])

    # -----------------------------
    # X dias atrás / ha X dias
    # -----------------------------
    match = re.fullmatch(r"(\d+)\s+dia[s]?\s+atras", normalized)
    if match:
        days = int(match.group(1))
        return today - timedelta(days=days)

    match = re.fullmatch(r"ha\s+(\d+)\s+dia[s]?", normalized)
    if match:
        days = int(match.group(1))
        return today - timedelta(days=days)

    # -----------------------------
    # X semanas atrás
    # -----------------------------
    match = re.fullmatch(r"(\d+)\s+semana[s]?\s+atras", normalized)
    if match:
        weeks = int(match.group(1))
        return today - timedelta(days=weeks * 7)

    # -----------------------------
    # "dia 15" ou apenas "15"
    # Interpreta como dia do mês atual.
    # Se quiser, dá para mover para mês anterior quando passar muito.
    # -----------------------------
    match = re.fullmatch(r"(dia\s+)?(\d{1,2})", normalized)
    if match:
        day = int(match.group(2))

        current_month_last_day = _last_day_of_month(today.year, today.month)
        if day <= current_month_last_day:
            candidate = date(today.year, today.month, day)

            if candidate < today - timedelta(days=15):
                next_month = today.month + 1
                next_year = today.year
                if next_month > 12:
                    next_month = 1
                    next_year += 1

                next_month_last_day = _last_day_of_month(next_year, next_month)
                if day <= next_month_last_day:
                    return date(next_year, next_month, day)

            return candidate

        return None

    # -----------------------------
    # dd/mm ou dd-mm
    # Ex: 24/02
    # Assume ano atual
    # -----------------------------
    match = re.fullmatch(r"(\d{1,2})[\/\-](\d{1,2})", normalized)
    if match:
        day = int(match.group(1))
        month = int(match.group(2))
        return _safe_date(today.year, month, day)

    # -----------------------------
    # dd/mm/yyyy ou dd-mm-yyyy
    # -----------------------------
    match = re.fullmatch(r"(\d{1,2})[\/\-](\d{1,2})[\/\-](\d{4})", normalized)
    if match:
        day = int(match.group(1))
        month = int(match.group(2))
        year = int(match.group(3))
        return _safe_date(year, month, day)

    # -----------------------------
    # yyyy-mm-dd
    # -----------------------------
    match = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", normalized)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        day = int(match.group(3))
        return _safe_date(year, month, day)

    return None



def normalize_intent_text(text: str) -> str:
    return normalize_lookup_text(text or "").strip()





def parse_yes_no(value: str):
    if not value:
        return None

    normalized = normalize_lookup_text(value)

    if normalized in ("sim", "s", "yes", "y"):
        return True

    if normalized in ("nao", "não", "n", "no"):
        return False

    return None   


def parse_pending_reply(text: str):
    if not text:
        return None

    value = text.strip().upper()
    normalized = normalize_lookup_text(text).strip()

    if value == "NOVA":
        return {"mode": "create_new", "index": None}

    if value.isdigit():
        return {"mode": "update_existing", "index": int(value) - 1}

    match = re.match(r"^CONCLUIR\s+(\d+)$", value)
    if match:
        return {"mode": "close_only", "index": int(match.group(1)) - 1}

    match = re.match(r"^(\d+)\s+CONCLUIR$", value)
    if match:
        return {"mode": "close_only", "index": int(match.group(1)) - 1}

    confirm_words = {
        "confirmar", "confirma", "confirmo", "ok", "certo", "isso", "fechou"
    }
    cancel_words = {
        "cancelar", "cancela", "cancel", "desconsidera", "esquece"
    }

    if normalized in confirm_words:
        return {"mode": "confirm_final", "index": None}

    if normalized in cancel_words:
        return {"mode": "cancel_final", "index": None}

    return None


def parse_pdf_selection(text: str):
    if not text:
        return []

    ordinal_idx = parse_human_ordinal_reference(text)
    if ordinal_idx is not None:
        return [ordinal_idx]

    raw = text.strip()
    parts = re.split(r"[,\s;]+", raw)

    indexes = []
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if not part.isdigit():
            return None
        indexes.append(int(part) - 1)

    seen = set()
    unique_indexes = []
    for idx in indexes:
        if idx not in seen:
            seen.add(idx)
            unique_indexes.append(idx)

    return unique_indexes


def parse_summary_edit_command(text: str):
    if not text:
        return None

    raw = re.sub(r"\s*\n\s*", " ", text).strip()
    raw = re.sub(r"\s+", " ", raw).strip()

    normalized = normalize_lookup_text(raw)

    patterns = [
        # fenologia
        (r"^(?:alterar|corrigir|corrija|ajustar|mudar|muda)\s+a?\s*fenologia(?:\s+observada)?(?:\s+para)?\s+(.+)$", "fenologia_real"),
        (r"^(?:fenologia|fenologia observada)(?:\s+para)?\s+(.+)$", "fenologia_real"),

        # data
        (r"^(?:alterar|corrigir|corrija|ajustar|mudar|muda)\s+a?\s*data(?:\s+da\s+visita)?(?:\s+para)?\s+(.+)$", "date"),
        (r"^(?:data|data da visita)(?:\s+para)?\s+(.+)$", "date"),

        # observações
        (r"^(?:alterar|corrigir|corrija|ajustar|mudar|muda)\s+a?\s*(?:obs|observacao|observacoes)(?:\s+para)?\s+(.+)$", "recommendation"),
        (r"^(?:obs|observacao|observacoes)(?:\s+para)?\s+(.+)$", "recommendation"),

        # cultura
        (r"^(?:alterar|corrigir|corrija|ajustar|mudar|muda)\s+a?\s*cultura(?:\s+para)?\s+(.+)$", "culture"),
        (r"^cultura(?:\s+para)?\s+(.+)$", "culture"),

        # variedade
        (r"^(?:alterar|corrigir|corrija|ajustar|mudar|muda)\s+a?\s*variedade(?:\s+para)?\s+(.+)$", "variety"),
        (r"^variedade(?:\s+para)?\s+(.+)$", "variety"),
    ]

    for pattern, field_name in patterns:
        match = re.match(pattern, normalized)
        if match:
            value = raw[match.start(1):].strip()
            return {
                "field": field_name,
                "value": value
            }

    return None



def parse_week_visit_action(text: str):
    if not text:
        return None

    normalized = normalize_intent_text(text)

    patterns = [
        (r"^(?:lanca|lancar|vamos lancar|quero lancar|realizar|fazer|abre|abrir)\s+(?:a\s+)?(?:visita\s+)?(\d+)$", "launch_week_visit"),
        (r"^(?:concluir|conclui|finalizar|finaliza|fechar|fecha)\s+(?:a\s+)?(?:visita\s+)?(\d+)$", "complete_week_visit"),
        (r"^(?:visita\s+)?(\d+)$", "launch_week_visit"),
    ]

    for pattern, intent in patterns:
        match = re.match(pattern, normalized)
        if match:
            return {
                "intent": intent,
                "index": int(match.group(1)) - 1
            }

    return None


def is_last_pdf_request(text: str) -> bool:
    normalized = normalize_intent_text(text)

    triggers = [
        "pdf da ultima visita",
        "pdf da última visita",
        "pdf ultima visita",
        "ultimo pdf",
        "último pdf",
        "pdf da visita mais recente",
        "me manda o pdf da ultima visita",
        "me manda o pdf da última visita",
    ]

    return any(trigger in normalized for trigger in triggers)

def is_pdf_request(text: str) -> bool:
    if not text:
        return False

    normalized = normalize_lookup_text(text)

    triggers = [
        "pdf",
        "gerar pdf",
        "me manda o pdf",
        "mande o pdf",
        "pdf da ultima visita",
        "pdf da última visita",
        "pdf das ultimas visitas",
        "pdf das últimas visitas",
        "relatorio pdf",
        "relatório pdf",
    ]

    return any(trigger in normalized for trigger in triggers)



def is_week_schedule_request(text: str) -> bool:
    normalized = normalize_intent_text(text)

    triggers = [
        "agenda da semana",
        "visitas da semana",
        "visitas pendentes da semana",
        "chat agenda da semana",
        "me passa agenda",
        "me passe agenda",
        "me passa as visitas da semana",
        "me passe as visitas da semana",
        "quais visitas tenho essa semana",
        "quais visitas tenho na semana",
        "minha agenda da semana",
    ]

    return any(trigger in normalized for trigger in triggers)


def normalize_culture_input(value: str):
    if not value:
        return None

    raw = value.strip().lower()
    normalized = normalize_lookup_text(raw)

    if normalized == "milho":
        return "Milho"
    if normalized == "soja":
        return "Soja"
    if normalized in ("algodao", "algodão"):
        return "Algodão"

    return None


def is_valid_fenologia(value: str) -> bool:
    if not value:
        return False

    value = value.strip().upper()

    valid_patterns = [
        r"^V\d{1,2}$",   # V1, V4, V10
        r"^R\d{1,2}$",   # R1, R2, R6
        r"^VE$",
        r"^VC$",
        r"^VT$",
    ]

    for pattern in valid_patterns:
        if re.match(pattern, value):
            return True

    return False


def build_name_confirmation_text(entity_label: str, candidates: list) -> str:
    if not candidates:
        return f"Não consegui identificar {entity_label}."

    lines = [f"Encontrei estes {entity_label}s parecidos:"]
    for idx, item in enumerate(candidates[:3], start=1):
        lines.append(f"{idx}. {item.name}")

    lines.append("")
    lines.append("Responda com o número correto.")
    return "\n".join(lines)


def build_pending_visits_confirmation_text(client_name: str, requested_culture: str, suggestions: list, same_culture_found: bool) -> str:
    lines = []

    if suggestions:
        if requested_culture and same_culture_found:
            lines.append(f"📋 Encontrei visitas pendentes de {requested_culture} para {client_name}:")
        elif requested_culture and not same_culture_found:
            lines.append(f"Não encontrei visitas pendentes de {requested_culture} para {client_name}.")
            lines.append("")
            lines.append("📋 Encontrei outras visitas pendentes deste cliente:")
        else:
            lines.append(f"📋 Encontrei visitas pendentes para {client_name}:")

        lines.append("")

        for idx, item in enumerate(suggestions, start=1):
            culture = item.get("culture") or "—"
            variety = item.get("variety") or "—"
            fenologia = item.get("fenologia_real") or "—"
            recommendation = item.get("recommendation") or "—"
            date_value = item.get("date") or "—"

            property_name = item.get("property_name") or "Sem fazenda"
            plot_name = item.get("plot_name") or "Sem talhão"

            lines.append(f"{idx}. {culture} | {variety}")
            lines.append(f"   🏡 Fazenda: {property_name}")
            lines.append(f"   📍 Talhão: {plot_name}")
            lines.append(f"   🌿 Fenologia: {fenologia}")
            lines.append(f"   📝 Evento/Obs: {recommendation}")
            lines.append(f"   📅 Data: {date_value}")
            lines.append("")

        lines.append("Responda com:")
        lines.append("🔢 número da visita para atualizar")
        lines.append("✅ CONCLUIR X para apenas concluir a visita pendente")
        lines.append("🆕 NOVA para criar uma nova visita")
        return "\n".join(lines)

    if requested_culture:
        return (
            f"Não encontrei visitas pendentes de {requested_culture} para {client_name}.\n\n"
            "Você pode responder com NOVA para criar uma nova visita."
        )

    return (
        f"Não encontrei visitas pendentes para {client_name}.\n\n"
        "Você pode responder com NOVA para criar uma nova visita."
    )

def build_pdf_visit_selection_text(visits: list) -> str:
    if not visits:
        return "Não encontrei visitas concluídas recentes para gerar PDF."

    lines = ["📄 Encontrei estas visitas recentes:", ""]

    for idx, v in enumerate(visits, start=1):
        client_name = v.client.name if getattr(v, "client", None) else f"Cliente {v.client_id}"
        visit_date = v.date.strftime("%d/%m/%Y") if v.date else "—"
        culture = v.culture or "—"
        fenologia = v.fenologia_real or "—"

        lines.append(f"{idx}. {visit_date} - {client_name} - {culture} - {fenologia}")

    lines.append("")
    lines.append("Responda com:")
    lines.append("🔢 um número: 1")
    lines.append("🔢 vários números: 1,3 ou 1 3 5")
    lines.append("❌ CANCELAR para sair")

    return "\n".join(lines)


def build_week_schedule_text(consultant_name: str, visits: list) -> str:
    start_date, end_date = get_week_date_range()

    if not visits:
        return (
            f"📅 Agenda da semana de {consultant_name}\n"
            f"Período: {start_date.isoformat()} até {end_date.isoformat()}\n\n"
            "Nenhuma visita pendente encontrada para esta semana."
        )

    lines = [
        f"📅 Agenda da semana de {consultant_name}",
        f"Período: {start_date.isoformat()} até {end_date.isoformat()}",
        "",
    ]

    for idx, visit in enumerate(visits, start=1):
        client_name = visit.client.name if getattr(visit, "client", None) else f"Cliente {visit.client_id}"
        recommendation = (visit.recommendation or "—").strip()
        culture = visit.culture or "—"
        visit_date = visit.date.isoformat() if visit.date else "—"

        lines.append(f"{idx}. {visit_date} - {client_name} - {culture} - {recommendation}")

    lines.append("")
    lines.append("Responda com:")
    lines.append("🔢 LANCAR VISITA X para atualizar uma visita da agenda")
    lines.append("✅ CONCLUIR VISITA X para apenas concluir")
    lines.append("❌ CANCELAR para sair")

    return "\n".join(lines)

def build_visit_summary_text(action: str, final_visit_payload: dict, selected_pending_visit: dict = None, close_only: bool = False) -> str:
    fenologia = final_visit_payload.get("fenologia_real") or "—"
    date_value = final_visit_payload.get("date") or "—"
    observations = final_visit_payload.get("recommendation") or "—"
    client_id = final_visit_payload.get("client_id")

    client_name = "—"
    if client_id:
        client = Client.query.get(client_id)
        if client:
            client_name = client.name

    lines = [f"📝 {bot_phrase('summary_intro', 'Resumo da visita')}", ""]

    if action == "use_existing_pending_visit" and selected_pending_visit:
        lines.append(f"🔧 Tipo: {'Concluir visita pendente' if close_only else 'Atualizar visita pendente'}")
        lines.append(f"🆔 ID da visita: {selected_pending_visit.get('id')}")
        lines.append(f"👤 Cliente: {client_name}")
        lines.append(f"📌 Recomendação pendente: {selected_pending_visit.get('recommendation') or '—'}")
        lines.append(f"🌿 Fenologia observada: {fenologia}")
        lines.append(f"📅 Data da visita: {date_value}")
        lines.append(f"💬 Observações: {observations}")

    elif action == "create_new_visit":
        lines.append("🆕 Tipo: Nova visita")
        lines.append("🆔 ID da visita: nova")
        lines.append(f"👤 Cliente: {client_name}")
        lines.append("📌 Recomendação pendente: —")
        lines.append(f"🌿 Fenologia observada: {fenologia}")
        lines.append(f"📅 Data da visita: {date_value}")
        lines.append(f"💬 Observações: {observations}")

    products = final_visit_payload.get("products") or []
    if products:
        lines.append("")
        lines.append("🧪 Produtos:")
        for p in products:
            product_name = p.get("product_name") or "—"
            dose = p.get("dose") or "—"
            unit = p.get("unit") or ""
            application_date = p.get("application_date") or "sem data"
            lines.append(f"- {product_name} — {dose} {unit} — {application_date}")

    lines.append("")
    lines.append("Responda com:")
    lines.append("✅ CONFIRMAR")
    lines.append("❌ CANCELAR")
    lines.append("✏️ ALTERAR FENOLOGIA V10")
    lines.append("📅 ALTERAR DATA hoje")
    lines.append("💬 ALTERAR OBSERVACAO baixa incidência de pragas")

    return "\n".join(lines)

def build_guided_state_payload(action: str, final_visit_payload: dict, selected_pending_visit: dict = None, close_only: bool = False) -> dict:
    return {
        "action": action,
        "final_visit_payload": final_visit_payload,
        "selected_pending_visit": selected_pending_visit,
        "close_only": close_only,
    }


def resolve_telegram_consultant(chat_message):
    if not chat_message:
        return None

    binding = TelegramContactBinding.query.filter_by(
        telegram_chat_id=str(chat_message.chat_id),
        is_active=True
    ).first()

    if binding and binding.consultant:
        return binding.consultant

    return None


def find_telegram_binding(chat_message):
    if not chat_message:
        return None

    return TelegramContactBinding.query.filter_by(
        telegram_chat_id=str(chat_message.chat_id),
        is_active=True
    ).first()


def bind_telegram_consultant_by_code(chat_message, code: str):
    if not chat_message or not code:
        return None, "dados inválidos"

    normalized_code = code.strip().upper()

    consultant = Consultant.query.filter(
        db.func.upper(Consultant.telegram_link_code) == normalized_code
    ).first()

    if not consultant:
        return None, "código inválido"

    existing = TelegramContactBinding.query.filter_by(
        telegram_chat_id=str(chat_message.chat_id)
    ).first()

    if existing:
        existing.telegram_user_id = str(chat_message.user_id) if chat_message.user_id else existing.telegram_user_id
        existing.telegram_username = chat_message.user_name or existing.telegram_username
        existing.display_name = chat_message.user_name or existing.display_name
        existing.consultant_id = consultant.id
        existing.is_active = True
        db.session.commit()
        return existing, None

    binding = TelegramContactBinding(
        telegram_chat_id=str(chat_message.chat_id),
        telegram_user_id=str(chat_message.user_id) if chat_message.user_id else None,
        telegram_username=chat_message.user_name or None,
        display_name=chat_message.user_name or None,
        consultant_id=consultant.id,
        is_active=True
    )

    db.session.add(binding)
    db.session.commit()
    return binding, None



def find_last_completed_visits_for_consultant(consultant_id: int, limit: int = 6):
    if not consultant_id:
        return []

    visits = (
        Visit.query
        .filter(Visit.consultant_id == consultant_id)
        .filter(Visit.status == "done")
        .order_by(Visit.date.desc().nullslast(), Visit.id.desc())
        .limit(limit)
        .all()
    )
    return visits


def find_consultant_pending_visits_for_week(consultant_id: int, reference_date=None, limit: int = 50):
    if not consultant_id:
        return []

    start_date, end_date = get_week_date_range(reference_date)

    visits = (
        Visit.query
        .filter(Visit.consultant_id == consultant_id)
        .filter(Visit.date >= start_date)
        .filter(Visit.date <= end_date)
        .filter(Visit.status != "done")
        .order_by(Visit.date.asc(), Visit.id.asc())
        .limit(limit)
        .all()
    )

    return visits


def find_client_by_name(client_name: str):
    if not client_name:
        return None, [], False

    target = normalize_lookup_text(client_name)
    target = re.sub(r"\s+", " ", target).strip()

    if not target:
        return None, [], False

    stopwords = {
        "cliente", "fazenda", "faz", "propriedade", "talhao", "talhão",
        "visita", "visitar", "lancar", "lançar", "concluir", "nova",
        "hoje", "amanha", "amanhã", "ontem", "observacao", "observação",
        "fenologia", "cultura"
    }

    target_tokens = [t for t in target.split() if t not in stopwords]
    target_clean = " ".join(target_tokens).strip() or target

    clients = Client.query.all()

    exact = []
    partial = []
    scored = []

    for client in clients:
        current = normalize_lookup_text(client.name)
        current_clean = re.sub(r"\s+", " ", current).strip()

        if current_clean == target_clean:
            exact.append(client)
            continue

        if target_clean and (target_clean in current_clean or current_clean in target_clean):
            partial.append(client)

        score_full = SequenceMatcher(None, target_clean, current_clean).ratio()

        token_hits = 0
        current_words = set(current_clean.split())
        for token in target_clean.split():
            if token in current_words:
                token_hits += 1

        token_score = token_hits / max(len(target_clean.split()), 1)
        final_score = max(score_full, token_score * 0.95)

        scored.append((client, final_score))

    if len(exact) == 1:
        return exact[0], exact, False

    if len(exact) > 1:
        return exact[0], exact[:3], True

    if len(partial) == 1:
        return partial[0], partial, False

    if len(partial) > 1:
        ranked_partial = sorted(
            partial,
            key=lambda c: SequenceMatcher(None, target_clean, normalize_lookup_text(c.name)).ratio(),
            reverse=True
        )
        best = ranked_partial[0]
        best_score = SequenceMatcher(None, target_clean, normalize_lookup_text(best.name)).ratio()
        if best_score >= 0.72:
            return best, ranked_partial[:3], False
        return best, ranked_partial[:3], True

    scored.sort(key=lambda x: x[1], reverse=True)

    if not scored:
        return None, [], True

    best_client, best_score = scored[0]
    top_candidates = [item[0] for item in scored[:3] if item[1] >= 0.45]

    if best_score >= 0.78:
        return best_client, top_candidates, False

    if best_score >= 0.58:
        return best_client, top_candidates, True

    return None, top_candidates, True


def find_known_product_names(limit: int = 300) -> list[str]:
    try:
        rows = (
            db.session.query(VisitProduct.product_name)
            .filter(VisitProduct.product_name.isnot(None))
            .distinct()
            .limit(limit)
            .all()
        )
        names = []
        for row in rows:
            value = row[0]
            if value:
                names.append(value.strip())
        return names
    except Exception:
        return []


def find_similar_product_name(raw_name: str, candidates: list[str] | None = None):
    if not raw_name:
        return None, 0.0

    candidates = candidates or find_known_product_names()
    if not candidates:
        return None, 0.0

    target = normalize_lookup_text(raw_name)
    best_name = None
    best_score = 0.0

    for candidate in candidates:
        current = normalize_lookup_text(candidate)
        score = SequenceMatcher(None, target, current).ratio()
        if score > best_score:
            best_score = score
            best_name = candidate

    return best_name, best_score


def normalize_products_from_parsed(products: list[dict] | None):
    products = products or []
    known_names = find_known_product_names()

    normalized_items = []

    for item in products:
        product_name = (item.get("product_name") or "").strip()
        if not product_name:
            continue

        best_name, score = find_similar_product_name(product_name, known_names)

        normalized_name = product_name
        if best_name and score >= 0.72:
            normalized_name = best_name

        normalized_items.append({
            "product_name": normalized_name,
            "dose": (item.get("dose") or "").strip(),
            "unit": (item.get("unit") or "").strip(),
            "application_date": item.get("application_date"),
        })

    return normalized_items


def is_products_only_update(parsed: dict | None) -> bool:
    parsed = parsed or {}

    has_products = bool(parsed.get("products"))
    has_fenologia = bool((parsed.get("fenologia_real") or "").strip())
    has_date = bool(parsed.get("date"))
    has_recommendation = bool((parsed.get("recommendation") or "").strip())
    has_culture = bool((parsed.get("culture") or "").strip())

    return has_products and not any([
        has_fenologia,
        has_date,
        has_recommendation,
        has_culture,
    ])


def find_property_by_name(property_name: str, client_id: int = None):
    if not property_name:
        return None, [], False

    target = normalize_lookup_text(property_name)
    if not target:
        return None, [], False

    query = Property.query
    if client_id:
        query = query.filter_by(client_id=client_id)

    properties = query.all()

    # 1) match exato
    for prop in properties:
        current = normalize_lookup_text(prop.name)
        if current == target:
            return prop, [prop], False

    # 2) match parcial
    partial_candidates = []
    for prop in properties:
        current = normalize_lookup_text(prop.name)
        if target in current or current in target:
            partial_candidates.append(prop)

    if len(partial_candidates) == 1:
        return partial_candidates[0], partial_candidates, False

    if len(partial_candidates) > 1:
        best_prop = None
        best_score = 0.0

        for prop in partial_candidates:
            current = normalize_lookup_text(prop.name)
            score = SequenceMatcher(None, target, current).ratio()
            if score > best_score:
                best_score = score
                best_prop = prop

        if best_prop and best_score >= 0.86:
            return best_prop, partial_candidates[:3], False

        return best_prop, partial_candidates[:3], True

    # 3) similaridade geral
    scored = []
    for prop in properties:
        current = normalize_lookup_text(prop.name)
        score = SequenceMatcher(None, target, current).ratio()
        scored.append((prop, score))

    scored.sort(key=lambda x: x[1], reverse=True)

    if not scored:
        return None, [], True

    best_prop, best_score = scored[0]
    top_candidates = [item[0] for item in scored[:3] if item[1] >= 0.55]

    if best_prop and best_score >= 0.86:
        return best_prop, top_candidates, False

    if best_prop and best_score >= 0.65:
        return best_prop, top_candidates, True

    return None, top_candidates, True




def find_pending_visits(
    client_id: int,
    property_id: int = None,
    culture: str = None,
    limit: int = 5
):
    """
    Busca visitas pendentes do cliente e, se houver, da propriedade.
    Retorna:
    - visits: lista de visitas
    - same_culture_found: se encontrou visitas da mesma cultura
    """
    base_query = Visit.query.filter(Visit.client_id == client_id)
    base_query = base_query.filter(Visit.status.in_(["planned", "pendente", "planejada", "planejado"]))

    if property_id:
        base_query = base_query.filter(Visit.property_id == property_id)

    if culture:
        same_culture = (
            base_query
            .filter(Visit.culture == culture)
            .order_by(Visit.date.asc().nullslast())
            .limit(limit)
            .all()
        )
        if same_culture:
            return same_culture, True

    fallback = (
        base_query
        .order_by(Visit.date.asc().nullslast())
        .limit(limit)
        .all()
    )
    return fallback, False



def get_current_chatbot_state(platform: str, chat_id: str):
    return ChatbotConversationState.query.filter_by(
        platform=platform,
        chat_id=chat_id
    ).first()



def extract_telegram_audio_info(payload: dict):
    if not payload:
        return None

    message = payload.get("message") or {}

    voice = message.get("voice")
    if voice and voice.get("file_id"):
        return {
            "file_id": voice.get("file_id"),
            "filename": "voice.ogg",
            "suffix": ".ogg",
            "mime_type": voice.get("mime_type") or "audio/ogg",
            "kind": "voice",
        }

    audio = message.get("audio")
    if audio and audio.get("file_id"):
        filename = audio.get("file_name") or "audio.mp3"
        suffix = os.path.splitext(filename)[1] or ".mp3"
        return {
            "file_id": audio.get("file_id"),
            "filename": filename,
            "suffix": suffix,
            "mime_type": audio.get("mime_type") or "audio/mpeg",
            "kind": "audio",
        }

    return None




def convert_audio_bytes_to_wav(audio_bytes: bytes, input_suffix: str = ".ogg"):
    """
    Converte áudio recebido em bytes para WAV usando ffmpeg.
    Retorna: (wav_bytes, erro)
    """
    try:
        with tempfile.NamedTemporaryFile(delete=False, suffix=input_suffix) as src:
            src.write(audio_bytes)
            src_path = src.name

        with tempfile.NamedTemporaryFile(delete=False, suffix=".wav") as dst:
            dst_path = dst.name

        cmd = [
            "ffmpeg",
            "-y",
            "-i", src_path,
            "-ar", "16000",
            "-ac", "1",
            dst_path,
        ]

        result = subprocess.run(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            timeout=120
        )

        if result.returncode != 0:
            return None, result.stderr.decode("utf-8", errors="ignore")

        with open(dst_path, "rb") as f:
            wav_bytes = f.read()

        return wav_bytes, None

    except Exception as e:
        return None, str(e)

    finally:
        try:
            os.remove(src_path)
        except:
            pass
        try:
            os.remove(dst_path)
        except:
            pass

def transcribe_audio_bytes(audio_bytes: bytes, filename: str = "audio.wav"):
    """
    Transcreve áudio usando OpenAI.
    Retorna: (texto_transcrito, erro)
    """
    try:
        api_key = os.getenv("OPENAI_API_KEY")
        if not api_key:
            return None, "OPENAI_API_KEY não configurada"

        client = OpenAI(api_key=api_key)

        audio_file = io.BytesIO(audio_bytes)
        audio_file.name = filename

        transcript = client.audio.transcriptions.create(
            model="gpt-4o-mini-transcribe",
            file=audio_file,
        )

        text = getattr(transcript, "text", None) or ""
        text = text.strip()

        if not text:
            return None, "transcrição vazia"

        return text, None

    except Exception as e:
        print("DEBUG transcribe_audio_bytes exception:", repr(e))
        return None, str(e)



def build_visit_pdf_file(visit_id: int):
    """
    Gera o mesmo PDF da visita e retorna:
    - buffer (BytesIO)
    - filename (str)
    """
    visit = Visit.query.get_or_404(visit_id)
    client = Client.query.get(visit.client_id)
    property_ = Property.query.get(visit.property_id) if visit.property_id else None
    plot = Plot.query.get(visit.plot_id) if visit.plot_id else None

    consultant = Consultant.query.get(visit.consultant_id) if visit.consultant_id else None
    consultant_name = consultant.name if consultant else (f"Consultor {visit.consultant_id}" if visit.consultant_id else "")

    if visit.planting_id:
        visits_to_include = (
            Visit.query
            .filter(Visit.planting_id == visit.planting_id)
            .order_by(Visit.date.desc(), Visit.id.desc())
            .all()
        )
    else:
        q = Visit.query.filter(
            Visit.client_id == visit.client_id,
            Visit.property_id == visit.property_id,
            Visit.plot_id == visit.plot_id,
            Visit.culture == visit.culture,
        )

        base_variety = (visit.variety or "").strip()
        if base_variety:
            q = q.filter(Visit.variety == base_variety)

        visits_to_include = (
            q.order_by(Visit.date.desc(), Visit.id.desc())
            .all()
        )

    filtered = []
    for v in visits_to_include:
        valid = [p for p in getattr(v, "photos", []) if getattr(p, "url", None)]
        if valid:
            v._valid_photos = valid
            filtered.append(v)

    visits_to_include = filtered[:MAX_VISITS]

    def nl2br(text: str) -> str:
        if not text:
            return ""
        t = text.replace("\r\n", "\n").replace("\r", "\n")
        t = html_escape(t)
        return t.replace("\n", "<br/>")

    static_dir = os.path.join(os.path.dirname(__file__), "static")
    nutriverde_logo_path = os.path.join(static_dir, "nutriverde_logo_pdf.png")

    def slugify_variety(name: str) -> str:
        if not name:
            return ""
        s = unicodedata.normalize("NFD", name)
        s = "".join(ch for ch in s if unicodedata.category(ch) != "Mn")
        s = s.strip().lower()
        s = re.sub(r"\s+", "_", s)
        s = re.sub(r"[^a-z0-9_]+", "", s)
        return s

    variety_slug = slugify_variety(visit.variety or "")
    variety_logo_path = os.path.join(static_dir, "variety_logos", f"{variety_slug}.png")

    def draw_footer(canvas, doc):
        canvas.saveState()
        y = 22
        pad = 50

        if variety_slug and os.path.exists(variety_logo_path):
            try:
                img = PILImage.open(variety_logo_path)
                aspect = img.height / float(img.width)
                w = 110
                h = w * aspect
                x = pad
                canvas.drawImage(variety_logo_path, x, y, width=w, height=h, mask="auto")
            except:
                pass

        if os.path.exists(nutriverde_logo_path):
            try:
                img = PILImage.open(nutriverde_logo_path)
                aspect = img.height / float(img.width)
                w = 70
                h = w * aspect
                x = A4[0] - pad - w
                canvas.drawImage(nutriverde_logo_path, x, y, width=w, height=h, mask="auto")
            except:
                pass

        canvas.restoreState()

    buffer = BytesIO()
    temp_jpgs = []

    def smart_params(total_photos_all: int):
        if total_photos_all <= 4:  return (1280, 75)
        if total_photos_all <= 8:  return (1200, 70)
        if total_photos_all <= 16: return (1100, 62)
        return (1000, 55)

    def download_to_temp(url: str, timeout=20, max_bytes=12_000_000):
        tmp = None
        try:
            req = Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urlopen(req, timeout=timeout) as r:
                try:
                    cl = r.headers.get("Content-Length")
                    if cl and int(cl) > max_bytes:
                        return None
                except:
                    pass

                tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".img")
                total = 0

                while True:
                    chunk = r.read(64 * 1024)
                    if not chunk:
                        break
                    total += len(chunk)
                    if total > max_bytes:
                        tmp.close()
                        try:
                            os.remove(tmp.name)
                        except:
                            pass
                        return None
                    tmp.write(chunk)

                tmp.close()
                return tmp.name
        except Exception:
            try:
                if tmp and tmp.name:
                    tmp.close()
                    os.remove(tmp.name)
            except:
                pass
            return None

    def compress_to_jpeg_temp(src_path: str, max_px: int, quality: int):
        try:
            img = PILImage.open(src_path)
            img = ImageOps.exif_transpose(img)
            img.thumbnail((max_px, max_px), PILImage.LANCZOS)

            out = tempfile.NamedTemporaryFile(delete=False, suffix=".jpg")
            out_path = out.name
            out.close()

            img.convert("RGB").save(out_path, "JPEG", optimize=True, quality=quality)
            try:
                img.close()
            except:
                pass
            return out_path
        except Exception:
            return None
        finally:
            try:
                os.remove(src_path)
            except:
                pass

    def draw_dark_background(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(colors.HexColor("#101010"))
        canvas.rect(0, 0, A4[0], A4[1], fill=True, stroke=False)
        canvas.restoreState()
        draw_footer(canvas, doc)

    def draw_cover_background(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(colors.HexColor("#0E0E0E"))
        canvas.rect(0, 0, A4[0], A4[1], fill=True, stroke=False)
        canvas.setFillColor(colors.HexColor("#00E676"))
        canvas.rect(0, 0, 28, A4[1], fill=True, stroke=False)
        canvas.restoreState()
        draw_footer(canvas, doc)

    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        leftMargin=50, rightMargin=40,
        topMargin=60, bottomMargin=40
    )

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="VisitTitleSmall", fontSize=12, leading=14, alignment=TA_CENTER, textColor=colors.HexColor("#BBF7D0"), spaceAfter=8))
    styles.add(ParagraphStyle(name="VisitStageBig", fontSize=22, leading=26, alignment=TA_CENTER, textColor=colors.HexColor("#FFFFFF"), spaceAfter=14))
    styles.add(ParagraphStyle(name="VisitDateCenter", fontSize=12, leading=14, alignment=TA_CENTER, textColor=colors.HexColor("#E0E0E0"), spaceAfter=14))
    styles.add(ParagraphStyle(name="VisitSectionLabel", fontSize=14, leading=16, alignment=TA_CENTER, textColor=colors.HexColor("#A5D6A7"), spaceBefore=10, spaceAfter=4))
    styles.add(ParagraphStyle(name="VisitSectionValue", fontSize=16, leading=20, alignment=TA_CENTER, textColor=colors.HexColor("#FFFFFF"), spaceAfter=14))
    styles.add(ParagraphStyle(name="HrLine", alignment=TA_CENTER, fontSize=10, textColor=colors.HexColor("#333333"), spaceBefore=10, spaceAfter=16))
    styles.add(ParagraphStyle(name="Caption", alignment=TA_CENTER, fontSize=9, textColor=colors.HexColor("#BDBDBD"), spaceBefore=4, spaceAfter=10))
    styles.add(ParagraphStyle(name="Footer", alignment=TA_CENTER, fontSize=9, textColor=colors.HexColor("#9E9E9E"), spaceBefore=20))

    story = []
    story.append(Spacer(1, 80))

    title_style = ParagraphStyle(name="CoverTitle", fontSize=22, leading=26, alignment=TA_CENTER, textColor=colors.HexColor("#E0F2F1"), spaceAfter=6)
    subtitle_style = ParagraphStyle(name="CoverSubtitle", fontSize=14, leading=18, alignment=TA_CENTER, textColor=colors.HexColor("#80CBC4"), spaceAfter=25)

    story.append(Paragraph("RELATÓRIO TÉCNICO DE", title_style))
    story.append(Paragraph("ACOMPANHAMENTO", title_style))
    story.append(Paragraph("Ciclo Fenológico", subtitle_style))

    client_style = ParagraphStyle(name="ClientBig", fontSize=22, leading=28, alignment=TA_CENTER, textColor=colors.HexColor("#FFFFFF"), spaceAfter=35)
    story.append(Paragraph((client.name or "Cliente").strip(), client_style))

    try:
        logo_path = os.path.join(static_dir, "nutricrm_logo_pdf.png")
        if os.path.exists(logo_path):
            img = PILImage.open(logo_path)
            aspect = img.height / float(img.width)
            width = 160
            story.append(Image(logo_path, width=width, height=width * aspect))
            story.append(Spacer(1, 20))
    except:
        pass

    info_label = ParagraphStyle(name="InfoLabel", fontSize=12, alignment=TA_LEFT, textColor=colors.HexColor("#A5D6A7"))
    info_value = ParagraphStyle(name="InfoValue", fontSize=12, alignment=TA_LEFT, textColor=colors.HexColor("#E0E0E0"), spaceAfter=6)

    days_planted_label = ParagraphStyle(
        name="DaysPlantedLabel",
        fontSize=12,
        leading=14,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#A5D6A7"),
        spaceBefore=10,
        spaceAfter=4,
    )

    days_planted_value = ParagraphStyle(
        name="DaysPlantedValue",
        fontSize=18,
        leading=22,
        alignment=TA_CENTER,
        textColor=colors.HexColor("#FFFFFF"),
        spaceAfter=10,
    )

    def add_info(label, value):
        if value:
            story.append(Paragraph(label, info_label))
            story.append(Paragraph(str(value), info_value))

    add_info("Propriedade:", property_.name if property_ else "")
    add_info("Talhão:", plot.name if plot else "")
    add_info("Cultura:", visit.culture or "")
    add_info("Variedade:", visit.variety or "")
    add_info("Consultor:", consultant_name or "")

    planting_date_obj = None

    # 1) tenta usar planting_date da Planting vinculada
    if getattr(visit, "planting_id", None):
        planting_row = Planting.query.get(visit.planting_id)
        if planting_row and getattr(planting_row, "planting_date", None):
            planting_date_obj = planting_row.planting_date

    # 2) fallback: usa a visita mais antiga do ciclo cuja fenologia seja Plantio
    if not planting_date_obj and visits_to_include:
        for cycle_visit in reversed(visits_to_include):
            fenologia_norm = (cycle_visit.fenologia_real or "").strip().lower()
            if fenologia_norm == "plantio" and cycle_visit.date:
                planting_date_obj = cycle_visit.date
                break

    # 3) fallback final: se a própria visita atual for Plantio
    if not planting_date_obj:
        fenologia_norm = (visit.fenologia_real or "").strip().lower()
        if fenologia_norm == "plantio" and visit.date:
            planting_date_obj = visit.date

    if visits_to_include:
        start_date_obj = visits_to_include[-1].date
        end_date_obj = visits_to_include[0].date
    else:
        start_date_obj = visit.date
        end_date_obj = visit.date

    start_date = start_date_obj.strftime("%d/%m/%Y") if start_date_obj else "—"
    end_date = end_date_obj.strftime("%d/%m/%Y") if end_date_obj else "—"

    add_info("Período de acompanhamento:", f"{start_date} → {end_date}")

    # dias de plantado = data de geração do PDF (hoje local) - data do plantio
    if planting_date_obj:
        dias_plantado = (get_local_today() - planting_date_obj).days
        if dias_plantado < 0:
            dias_plantado = 0

        story.append(Spacer(1, 8))
        story.append(Paragraph("Dias de plantado", days_planted_label))
        story.append(Paragraph(f"{dias_plantado} dias", days_planted_value))

    story.append(Spacer(1, 40))
    story.append(PageBreak())

    total_visits = len(visits_to_include)
    for pos, v in enumerate(visits_to_include):
        idx = total_visits - pos

        story.append(Paragraph(f"VISITA {idx}", styles["VisitTitleSmall"]))
        story.append(Paragraph(v.fenologia_real or "—", styles["VisitStageBig"]))

        try:
            dtext = v.date.strftime("%d/%m/%Y")
        except:
            dtext = str(v.date)
        story.append(Paragraph(dtext, styles["VisitDateCenter"]))
        story.append(Spacer(1, 20))

        if v.recommendation:
            story.append(Paragraph("Observações", styles["VisitSectionLabel"]))
            story.append(Paragraph(nl2br(v.recommendation), styles["VisitSectionValue"]))

        story.append(Paragraph("<hr/>", styles["HrLine"]))

        photos = list(getattr(v, "_valid_photos", []) or [])
        if photos:
            photos = photos[:MAX_PHOTOS_V]
            total = len(photos)

            cols = 1 if total <= 3 else (2 if total <= 6 else 3)
            max_width = 220 if cols == 1 else 160
            col_width = (A4[0] - 100) / cols

            total_all = sum(min(len(getattr(x, "photos", []) or []), MAX_PHOTOS_V) for x in visits_to_include)
            max_px, quality = smart_params(total_all)

            row = []
            count = 0

            for i, photo in enumerate(photos, 1):
                photo_url = resolve_photo_url(photo.url)
                if not photo_url:
                    continue

                try:
                    src_path = download_to_temp(photo_url, timeout=20, max_bytes=12_000_000)
                    if not src_path:
                        continue

                    jpg_path = compress_to_jpeg_temp(src_path, max_px=max_px, quality=quality)
                    if not jpg_path:
                        continue

                    temp_jpgs.append(jpg_path)

                    probe = PILImage.open(jpg_path)
                    w, h = probe.size
                    try:
                        probe.close()
                    except:
                        pass

                    aspect = (h / w) if w else 1
                    img_obj = Image(jpg_path, width=max_width, height=max_width * aspect)

                    base_caption = getattr(photo, "caption", "") or ""
                    lat = getattr(photo, "latitude", None)
                    lon = getattr(photo, "longitude", None)

                    gps_caption = ""
                    if lat is not None and lon is not None:
                        gps_caption = f"📍 {lat:.5f}, {lon:.5f}"

                    final_caption = html_escape(base_caption)
                    if gps_caption:
                        final_caption += f"<br/><small>{html_escape(gps_caption)}</small>"

                    caption_par = Paragraph(final_caption, styles["Caption"])

                    row.append([img_obj, caption_par])
                    count += 1

                    if count == cols or i == total:
                        story.append(
                            Table(
                                [row],
                                colWidths=[col_width] * len(row),
                                hAlign="CENTER",
                                style=TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")])
                            )
                        )
                        story.append(Spacer(1, 14))
                        row = []
                        count = 0

                except Exception:
                    continue

            if pos < total_visits - 1:
                story.append(PageBreak())

    story.append(Paragraph("<b>NutriCRM</b>", styles["Footer"]))
    story.append(Paragraph("Relatório cumulativo — ciclo fenológico", styles["Footer"]))

    doc.build(story, onFirstPage=draw_cover_background, onLaterPages=draw_dark_background)
    buffer.seek(0)

    for p in temp_jpgs:
        try:
            os.remove(p)
        except:
            pass

    filename = f"{client.name if client else 'Cliente'} - {visit.variety or ''} - Relatório.pdf"
    return buffer, filename




def interpret_user_message_with_ai(message_text: str, current_state: str = ""):
    """
    Interpreta mensagem livre do usuário com OpenAI.
    Retorna um dict estruturado ou None.
    Muito mais tolerante a erro de digitação e linguagem natural.
    """
    try:
        client = OpenAI()

        cleaned_text = compact_user_text_for_ai(message_text)

        system_prompt = """
Você é um interpretador de intenções para um chatbot agrícola do AgroCRM.

Seu papel é transformar a mensagem do usuário em JSON ESTRITAMENTE válido.
Nunca explique nada.
Nunca escreva texto fora do JSON.
Nunca use markdown.
Nunca use crases.
Nunca escreva comentários.

Você deve ser extremamente tolerante a:
- erros de digitação
- falta de acento
- abreviações
- frases incompletas
- português informal
- ordem bagunçada dos dados

Contexto do sistema:
O chatbot é usado por consultores agrícolas para:
1) consultar agenda semanal
2) lançar visita de uma agenda já listada
3) concluir visita rapidamente
4) criar nova visita
5) corrigir campos do resumo antes de confirmar
6) pedir PDF da última visita
7) pedir lista de PDFs recentes
8) confirmar ou cancelar fluxos

Estados possíveis do chatbot:
- awaiting_week_visit_selection
- awaiting_final_confirmation
- awaiting_confirmation
- awaiting_fenologia
- awaiting_date
- awaiting_observations
- awaiting_pdf_visit_selection
- awaiting_pdf_confirmation
- awaiting_client_confirmation
- awaiting_culture
- awaiting_planting_confirmation
- awaiting_avulsa_confirmation
- none
- today_schedule_request
- daily_routine_request
- pdf_by_client_reference
- contextual_visit_reference



Regras gerais:
- Se o usuário pedir agenda semanal, retorne intent = week_schedule_request
- Se o usuário quiser abrir/lançar/fazer/realizar uma visita da agenda e citar um número, retorne intent = launch_week_visit
- Se o usuário quiser concluir/fechar/finalizar uma visita da agenda e citar um número, retorne intent = complete_week_visit
- Se o usuário quiser PDF da última visita, retorne intent = pdf_last_visit
- Se o usuário quiser lista de PDFs ou gerar PDF sem especificar qual, retorne intent = pdf_recent_visits
- Se o usuário quiser confirmar, retorne intent = confirm
- Se o usuário quiser cancelar, retorne intent = cancel
- Se o usuário quiser alterar um campo do resumo, retorne intent = edit_summary
- Se a mensagem parecer um lançamento completo de visita, retorne intent = create_visit_like_message
- Se não souber, retorne intent = unknown
- "o que tenho hoje", "agenda de hoje", "meu dia", "o que falta hoje" => today_schedule_request ou daily_routine_request
- "prioridades de hoje", "rotina do dia", "resumo do dia" => daily_routine_request
- "pdf do evaristo", "manda o pdf do marcelo", "pdf da ultima do ivan" => pdf_by_client_reference
- "a terceira", "essa", "a do evaristo" => contextual_visit_reference



Campos possíveis no JSON:
- intent
- confidence
- visit_index
- field
- value
- parsed_visit

Campos permitidos em field:
- fenologia_real
- date
- recommendation
- culture
- variety

Campos permitidos em parsed_visit:
- client_name
- property_name
- plot_name
- culture
- fenologia_real
- date
- recommendation

Regras para edit_summary:
- "corrija a fenologia para v10" -> field fenologia_real, value V10
- "muda a data pra hoje" -> field date, value hoje
- "ajusta observacao para baixa incidencia de pragas" -> field recommendation
- "troca cultura pra soja" -> field culture
- "muda variedade para as 1868 pro4" -> field variety

Regras para datas:
- preserve exatamente termos como "hoje", "amanha", "amanhã"
- preserve datas digitadas como "24/02/2026", "24/02", "2026-02-24"
- preserve números simples como "15" se parecerem data do mês

Regras para visita da agenda:
- "visita 7" normalmente significa launch_week_visit
- "lança a 7", "lancar visita 7", "realizar 7", "fazer a 3" => launch_week_visit
- "concluir 7", "fechar visita 7", "finalizar a 4" => complete_week_visit

Regras para PDF:
- "pdf da ultima", "manda o pdf da ultima visita", "ultimo pdf" => pdf_last_visit
- "gera pdf", "me manda um pdf", "quero pdf" => pdf_recent_visits

Regras para agenda:
- "agenda da semana", "minha agenda", "visitas da semana", "o que tenho essa semana" => week_schedule_request

Regras para confirmação:
- "confirmar", "ok", "pode confirmar", "fechou", "isso", "certo" => confirm
- "cancelar", "cancela", "para", "desconsidera", "esquece" => cancel

Regras para mensagens completas de visita:
Se a mensagem parecer um lançamento completo, extraia parsed_visit.
Exemplo:
"cliente Marcelo Alonso soja v4 hoje aplicar fungicida"
=> intent create_visit_like_message

Se houver erro de digitação em cultura ou fenologia, tente inferir o mais provável.
Exemplos:
- "sojja" -> "Soja"
- "milhho" -> "Milho"
- "algodao" -> "Algodão"
- "penduamento" pode aparecer como recomendação/fenologia textual
- "fenolojia v10" -> field fenologia_real value V10

Formato de saída:
Retorne SEMPRE JSON válido.
Confidence deve ser: high, medium ou low.

Exemplos de saída:

{"intent":"week_schedule_request","confidence":"high"}

{"intent":"launch_week_visit","confidence":"high","visit_index":7}

{"intent":"complete_week_visit","confidence":"high","visit_index":3}

{"intent":"pdf_last_visit","confidence":"high"}

{"intent":"pdf_recent_visits","confidence":"high"}

{"intent":"confirm","confidence":"medium"}

{"intent":"cancel","confidence":"high"}

{"intent":"edit_summary","confidence":"high","field":"fenologia_real","value":"V10"}

{"intent":"edit_summary","confidence":"high","field":"date","value":"hoje"}

{"intent":"edit_summary","confidence":"high","field":"recommendation","value":"baixa incidencia de pragas"}

{"intent":"create_visit_like_message","confidence":"medium","parsed_visit":{"client_name":"Marcelo Alonso","property_name":"","plot_name":"","culture":"Soja","fenologia_real":"V4","date":"hoje","recommendation":"aplicar fungicida"}}

{"intent":"unknown","confidence":"low"}

{"intent":"today_schedule_request","confidence":"high"}

{"intent":"daily_routine_request","confidence":"high"}

{"intent":"pdf_by_client_reference","confidence":"medium","parsed_visit":{"client_name":"Evaristo Barzotto"}}

{"intent":"contextual_visit_reference","confidence":"medium","visit_index":3}
""".strip()

        user_prompt = f"""
Estado atual do chatbot: {current_state or "none"}

Mensagem do usuário:
{cleaned_text}
""".strip()

        response = client.responses.create(
            model="gpt-4.1-mini",
            input=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        )

        output_text = (response.output_text or "").strip()
        if not output_text:
            return None

        # tentativa direta
        try:
            data = json.loads(output_text)
        except Exception:
            # fallback: tenta extrair primeiro bloco JSON da resposta
            match = re.search(r"\{.*\}", output_text, re.DOTALL)
            if not match:
                return None
            data = json.loads(match.group(0))

        # saneamento mínimo
        if not isinstance(data, dict):
            return None

        intent = data.get("intent")
        if not intent:
            return None

        allowed_intents = {
            "week_schedule_request",
            "today_schedule_request",
            "daily_routine_request",
            "launch_week_visit",
            "complete_week_visit",
            "pdf_last_visit",
            "pdf_recent_visits",
            "pdf_by_client_reference",
            "contextual_visit_reference",
            "confirm",
            "cancel",
            "edit_summary",
            "create_visit_like_message",
            "unknown",
        }

        if intent not in allowed_intents:
            return None

        confidence = (data.get("confidence") or "low").lower()
        if confidence not in {"high", "medium", "low"}:
            data["confidence"] = "low"

        if "visit_index" in data and data["visit_index"] is not None:
            try:
                data["visit_index"] = int(data["visit_index"])
            except Exception:
                data["visit_index"] = None

        if "field" in data and data["field"] is not None:
            allowed_fields = {"fenologia_real", "date", "recommendation", "culture", "variety"}
            if data["field"] not in allowed_fields:
                data["field"] = None

        return data

    except Exception as e:
        print(f"⚠️ IA fallback falhou: {e}")
        return None



def extract_telegram_photo_info(payload: dict) -> dict | None:
    if not payload:
        return None

    message = payload.get("message") or payload.get("edited_message") or {}
    photos = message.get("photo") or []
    caption = (message.get("caption") or "").strip()

    if not photos:
        return None

    best = photos[-1] if photos else None
    if not best:
        return None

    return {
        "file_id": best.get("file_id"),
        "file_unique_id": best.get("file_unique_id"),
        "width": best.get("width"),
        "height": best.get("height"),
        "file_size": best.get("file_size"),
        "caption": caption,
        "mime_group": "photo",
        "media_group_id": message.get("media_group_id"),
        "telegram_message_id": message.get("message_id"),
    }


def download_telegram_file_bytes(file_id: str) -> tuple[bytes | None, str | None]:
    """
    Baixa um arquivo do Telegram a partir do file_id.
    Retorna: (bytes, erro)
    """
    try:
        import os
        import requests

        bot_token = os.getenv("TELEGRAM_BOT_TOKEN")
        if not bot_token:
            return None, "TELEGRAM_BOT_TOKEN não configurado"

        get_file_url = f"https://api.telegram.org/bot{bot_token}/getFile"
        resp = requests.get(get_file_url, params={"file_id": file_id}, timeout=20)
        if resp.status_code != 200:
            return None, f"falha getFile status={resp.status_code}"

        data = resp.json()
        if not data.get("ok"):
            return None, f"getFile retornou ok=False: {data}"

        file_path = data.get("result", {}).get("file_path")
        if not file_path:
            return None, "file_path ausente no getFile"

        download_url = f"https://api.telegram.org/file/bot{bot_token}/{file_path}"
        file_resp = requests.get(download_url, timeout=40)
        if file_resp.status_code != 200:
            return None, f"falha download status={file_resp.status_code}"

        return file_resp.content, None

    except Exception as e:
        return None, str(e)


def guess_telegram_photo_filename(photo_info: dict | None) -> str:
    if not photo_info:
        return "telegram_photo.jpg"

    file_unique_id = photo_info.get("file_unique_id") or "photo"
    return f"telegram_{file_unique_id}.jpg"




def attach_photo_to_visit_from_telegram(
    visit,
    photo_bytes: bytes | None,
    filename: str | None,
    caption: str | None = None,
):
    """
    Faz upload da foto recebida do Telegram para o Cloudflare R2
    e cria o registro Photo vinculado à visita.
    """
    if not visit:
        return None, "visita ausente"

    if not photo_bytes:
        return None, "foto ausente"

    try:
        bucket = os.environ.get("R2_BUCKET")
        public_base = (os.environ.get("R2_PUBLIC_BASE_URL") or "").rstrip("/")

        if not bucket or not public_base:
            return None, "R2 não configurado: faltam variáveis de ambiente"

        r2 = get_r2_client()

        original = secure_filename(filename or "telegram_photo.jpg")
        if "." not in original:
            original = f"{original}.jpg"

        unique = uuid.uuid4().hex
        key = f"visits/{visit.id}/{unique}_{original}"

        import io
        file_obj = io.BytesIO(photo_bytes)

        r2.upload_fileobj(
            Fileobj=file_obj,
            Bucket=bucket,
            Key=key,
            ExtraArgs={
                "ContentType": "image/jpeg",
            },
        )

        url = f"{public_base}/{key}"

        photo = Photo(
            visit_id=visit.id,
            url=url,
            caption=(caption or "").strip() or None
        )

        db.session.add(photo)
        db.session.commit()

        return photo, None

    except Exception as e:
        db.session.rollback()
        return None, str(e)


def try_extract_client_from_free_text(message_text: str):
    if not message_text:
        return None

    normalized = normalize_lookup_text(message_text)
    normalized = re.sub(r"\s+", " ", normalized).strip()

    if not normalized:
        return None

    blocked_prefixes = (
        "/start",
        "/vincular",
        "pdf",
        "agenda da semana",
        "visitas da semana",
        "me passa agenda",
        "gerar pdf",
        "confirmar",
        "cancelar",
        "concluir visita",
        "lancar visita",
        "lançar visita",
    )

    if any(normalized.startswith(p) for p in blocked_prefixes):
        return None

    client, candidates, needs_confirmation = find_client_by_name(normalized)
    if client:
        return {
            "client": client,
            "candidates": candidates,
            "needs_confirmation": needs_confirmation,
        }

    return None



def get_month_date_range(reference_date=None):
    today = reference_date or _date.today()
    start = today.replace(day=1)

    if today.month == 12:
        next_month = today.replace(year=today.year + 1, month=1, day=1)
    else:
        next_month = today.replace(month=today.month + 1, day=1)

    end = next_month - _timedelta(days=1)
    return start, end


def is_month_visits_request(text: str) -> bool:
    if not text:
        return False

    normalized = normalize_lookup_text(text)

    triggers = [
        "visitas do mes",
        "visitas do mês",
        "minhas visitas do mes",
        "minhas visitas do mês",
        "visitas do mes concluidas",
        "visitas do mês concluidas",
        "visitas do mes concluídas",
        "visitas do mês concluídas",
    ]

    return any(trigger in normalized for trigger in triggers)


def parse_month_visit_filter(text: str) -> str:
    normalized = normalize_lookup_text(text)

    if "atrasad" in normalized:
        return "overdue"

    if "concluid" in normalized or "concluíd" in normalized or "feitas" in normalized:
        return "done"

    return "all"


def find_consultant_visits_for_month(consultant_id: int, filter_mode: str = "all", reference_date=None, limit: int = 100):
    if not consultant_id:
        return []

    start_date, end_date = get_month_date_range(reference_date)
    today = reference_date or _date.today()

    q = (
        Visit.query
        .filter(Visit.consultant_id == consultant_id)
        .filter(Visit.date >= start_date)
        .filter(Visit.date <= end_date)
    )

    if filter_mode == "done":
        q = q.filter(Visit.status == "done")
    elif filter_mode == "overdue":
        q = q.filter(Visit.status != "done").filter(Visit.date < today)

    visits = (
        q.order_by(Visit.date.asc().nullslast(), Visit.id.asc())
        .limit(limit)
        .all()
    )

    return visits


def build_month_visits_text(consultant_name: str, visits: list, filter_mode: str = "all") -> str:
    filter_label_map = {
        "all": "todas",
        "done": "concluídas",
        "overdue": "atrasadas",
    }

    filter_label = filter_label_map.get(filter_mode, "todas")

    if not visits:
        return (
            f"📋 Visitas do mês para {consultant_name}\n"
            f"Filtro: {filter_label}\n\n"
            f"Não encontrei visitas nesse filtro."
        )

    lines = [
        f"📋 Visitas do mês para {consultant_name}",
        f"Filtro: {filter_label}",
        ""
    ]

    for i, v in enumerate(visits, start=1):
        client_name = v.client.name if v.client else f"Cliente {v.client_id}"
        date_label = v.date.isoformat() if v.date else "sem data"
        status_label = v.status or "planned"
        culture_label = v.culture or (v.planting.culture if getattr(v, "planting", None) else "")
        stage_label = v.fenologia_real or ""

        extra = " - ".join([x for x in [culture_label, stage_label] if x])
        if extra:
            lines.append(f"{i}. {client_name} - {date_label} - {status_label} - {extra}")
        else:
            lines.append(f"{i}. {client_name} - {date_label} - {status_label}")

    lines.extend([
        "",
        "Responda com:",
        "🔢 número da visita para editar",
        "📄 PDF X para gerar o PDF",
        "❌ CANCELAR"
    ])

    return "\n".join(lines)


def parse_month_visit_action(text: str):
    normalized = normalize_lookup_text(text)

    if normalized in {"cancelar", "cancel"}:
        return {"mode": "cancel"}

    pdf_match = re.fullmatch(r"pdf\s+(\d+)", normalized)
    if pdf_match:
        return {"mode": "pdf", "index": int(pdf_match.group(1)) - 1}

    if normalized.isdigit():
        return {"mode": "select", "index": int(normalized) - 1}
    ordinal_idx = parse_human_ordinal_reference(text)
    if ordinal_idx is not None:
        return {"mode": "select", "index": ordinal_idx}

    return None





def extract_prefill_from_message_text(message_text: str):
    if not message_text:
        return {
            "date": None,
            "culture": "",
            "fenologia_real": None,
            "recommendation": "",
            "products": [],
        }

    parsed = parse_chatbot_message(message_text) or {}

    recommendation = extract_recommendation_fallback(message_text)

    if not recommendation:
        recommendation = (parsed.get("recommendation") or "").strip()

    return {
        "date": parsed.get("date"),
        "culture": parsed.get("culture") or "",
        "fenologia_real": parsed.get("fenologia_real"),
        "recommendation": recommendation,
        "products": normalize_products_from_parsed(parsed.get("products") or []),
    }




def extract_recommendation_fallback(message_text: str) -> str:
    if not message_text:
        return ""

    raw = message_text.strip()

    # normaliza quebras de linha, mas preserva o conteúdo
    raw_single = re.sub(r"\s*\n\s*", " ", raw).strip()

    patterns = [
        r"(?:observacoes|observação|observacao|obs)\s*[:,\-]?\s*([\s\S]+)$",
    ]

    for pattern in patterns:
        match = re.search(pattern, raw_single, re.IGNORECASE)
        if match:
            value = match.group(1).strip(" .,-;:")
            if value:
                return value

    return ""



def get_local_today() -> date:
    """
    Retorna a data local do usuário/negócio.
    Para Lucas do Rio Verde (MT), use America/Cuiaba.
    """
    return _dt.now(ZoneInfo("America/Cuiaba")).date()




def get_pending_media_dir(chat_id: str) -> Path:
    base_dir = Path("/tmp") / "telegram_pending_media" / str(chat_id)
    base_dir.mkdir(parents=True, exist_ok=True)
    return base_dir


def get_pending_media_manifest_path(chat_id: str) -> Path:
    return get_pending_media_dir(chat_id) / "pending_media.json"


def load_pending_media_manifest(chat_id: str) -> dict:
    manifest_path = get_pending_media_manifest_path(chat_id)
    if not manifest_path.exists():
        return {"photos": []}

    try:
        return json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        return {"photos": []}


def save_pending_media_manifest(chat_id: str, data: dict) -> None:
    manifest_path = get_pending_media_manifest_path(chat_id)
    manifest_path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )


def save_pending_telegram_photo(chat_id: str, photo_bytes: bytes, filename: str, caption: str = "") -> dict:
    media_dir = get_pending_media_dir(chat_id)
    ext = Path(filename).suffix or ".jpg"
    unique_name = f"{uuid.uuid4().hex}{ext}"
    file_path = media_dir / unique_name

    with open(file_path, "wb") as f:
        f.write(photo_bytes)

    manifest = load_pending_media_manifest(chat_id)
    row = {
        "filename": unique_name,
        "original_filename": filename,
        "caption": caption or "",
    }
    manifest["photos"].append(row)
    save_pending_media_manifest(chat_id, manifest)
    print("DEBUG save_pending_telegram_photo manifest_count:", len(manifest.get("photos", [])))
    return row


def get_pending_telegram_photos(chat_id: str) -> list:
    manifest = load_pending_media_manifest(chat_id)
    media_dir = get_pending_media_dir(chat_id)

    photos = []
    for item in manifest.get("photos", []):
        file_path = media_dir / item["filename"]
        if file_path.exists():
            photos.append({
                "path": str(file_path),
                "filename": item["original_filename"] or item["filename"],
                "caption": item.get("caption") or "",
            })
    print("DEBUG get_pending_telegram_photos count:", len(photos))
    return photos


def clear_pending_telegram_photos(chat_id: str) -> None:
    media_dir = get_pending_media_dir(chat_id)
    if not media_dir.exists():
        return

    for child in media_dir.iterdir():
        try:
            if child.is_file():
                child.unlink()
        except Exception:
            pass

    try:
        media_dir.rmdir()
    except Exception:
        pass


def attach_pending_telegram_photos_to_visit(chat_id: str, visit):
    pending = get_pending_telegram_photos(chat_id)
    print("DEBUG attach_pending_telegram_photos_to_visit visit_id:", getattr(visit, "id", None))
    print("DEBUG attach_pending_telegram_photos_to_visit pending_count:", len(pending))
    attached = 0
    errors = []

    for item in pending:
        try:
            with open(item["path"], "rb") as f:
                photo_bytes = f.read()

            _, attach_error = attach_photo_to_visit_from_telegram(
                visit=visit,
                photo_bytes=photo_bytes,
                filename=item["filename"],
                caption=item["caption"],
            )

            if attach_error:
                errors.append(attach_error)
            else:
                attached += 1

        except Exception as e:
            errors.append(str(e))

    if attached > 0:
        clear_pending_telegram_photos(chat_id)
    print("DEBUG attach_pending_telegram_photos_to_visit attached:", attached)
    print("DEBUG attach_pending_telegram_photos_to_visit errors:", errors)
    return attached, errors




def get_pending_media_group_marker_path(chat_id: str, media_group_id: str) -> Path:
    media_dir = get_pending_media_dir(chat_id)
    return media_dir / f"media_group_{media_group_id}.marker"


def should_send_photo_prompt(chat_id: str, photo_info: dict | None) -> bool:
    if not photo_info:
        return False

    media_group_id = photo_info.get("media_group_id")
    if not media_group_id:
        return True

    marker = get_pending_media_group_marker_path(chat_id, str(media_group_id))
    if marker.exists():
        return False

    marker.write_text("sent", encoding="utf-8")
    return True



def resolve_audio_message_text(chat_message, payload, current_text: str):
    """
    Se a mensagem não tiver texto/caption mas tiver áudio,
    baixa, converte, transcreve e devolve o texto final.
    Se falhar, já responde ao usuário e devolve None.
    """
    message_text = (current_text or "").strip()

    if message_text:
        return message_text

    audio_info = extract_telegram_audio_info(payload)
    if not audio_info:
        return message_text

    audio_bytes, download_error = download_telegram_file_bytes(audio_info["file_id"])
    if download_error or not audio_bytes:
        send_telegram_message(
            chat_id=chat_message.chat_id,
            text="Recebi seu áudio, mas não consegui baixar para transcrever."
        )
        return None

    wav_bytes, convert_error = convert_audio_bytes_to_wav(
        audio_bytes=audio_bytes,
        input_suffix=audio_info["suffix"],
    )
    if convert_error or not wav_bytes:
        send_telegram_message(
            chat_id=chat_message.chat_id,
            text="Recebi seu áudio, mas não consegui converter para transcrição."
        )
        return None

    transcript_text, transcript_error = transcribe_audio_bytes(
        audio_bytes=wav_bytes,
        filename="audio.wav",
    )
    if transcript_error or not transcript_text:
        send_telegram_message(
            chat_id=chat_message.chat_id,
            text=bot_phrase(
                "audio_fail",
                "Recebi seu áudio, mas não consegui transcrever. Tente novamente ou envie em texto."
            )
        )
        return None

    message_text = transcript_text.strip()

    send_telegram_message(
        chat_id=chat_message.chat_id,
        text=f"🎤 Áudio transcrito:\n{message_text}"
    )

    return message_text


def resolve_pending_photo_for_message(chat_message, payload, current_text: str):
    """
    Salva foto recebida do Telegram em armazenamento temporário.
    Se a mensagem vier só com foto e sem texto, responde pedindo contexto
    e devolve o marcador '__PHOTO_ONLY_WAITING_CONTEXT__'.
    """
    message_text = (current_text or "").strip()
    photo_info = extract_telegram_photo_info(payload)

    if not photo_info:
        return None

    downloaded_photo_name = guess_telegram_photo_filename(photo_info)
    downloaded_photo_bytes, photo_download_error = download_telegram_file_bytes(photo_info["file_id"])

    if photo_download_error:
        print("DEBUG telegram photo download error:", photo_download_error)
        downloaded_photo_bytes = None

    if downloaded_photo_bytes:
        try:
            saved_media = save_pending_telegram_photo(
                chat_id=chat_message.chat_id,
                photo_bytes=downloaded_photo_bytes,
                filename=downloaded_photo_name,
                caption=photo_info.get("caption") or "",
            )
            print("DEBUG pending telegram photo saved:", saved_media)
        except Exception as e:
            print("DEBUG save_pending_telegram_photo error:", str(e))

    if not message_text:
        if should_send_photo_prompt(chat_message.chat_id, photo_info):
            send_telegram_message(
                chat_id=chat_message.chat_id,
                text=(
                    bot_phrase(
                        "photo_saved_waiting_context",
                        "Recebi sua foto e já deixei ela separada para esta conversa."
                    )
                    + "\n\n"
                    + "Exemplo:\n"
                    + "- Cliente Rogério Remor, fenologia observada V11, data hoje, observações plantas sadias\n"
                    + "- Ou envie um áudio com essas informações"
                )
            )
        return "__PHOTO_ONLY_WAITING_CONTEXT__"

    return photo_info


def build_final_visit_payload(
    base_preview: dict,
    selected_pending_visit: dict | None,
    resolved_consultant_id: int,
    close_only: bool = False
):
    """
    Consolida o payload final antes de persistir.
    """
    base_preview = base_preview or {}
    selected_pending_visit = selected_pending_visit or {}

    linked_pending_visit_id = (
        base_preview.get("linked_pending_visit_id")
        or selected_pending_visit.get("id")
    )

    client_id = (
        base_preview.get("client_id")
        or selected_pending_visit.get("client_id")
    )

    property_id = (
        base_preview.get("property_id")
        if base_preview.get("property_id") is not None
        else selected_pending_visit.get("property_id")
    )

    plot_id = (
        base_preview.get("plot_id")
        if base_preview.get("plot_id") is not None
        else selected_pending_visit.get("plot_id")
    )

    consultant_id = (
        base_preview.get("consultant_id")
        or resolved_consultant_id
    )

    culture = (
        (base_preview.get("culture") or "").strip()
        or (selected_pending_visit.get("culture") or "").strip()
    )

    variety = (
        (base_preview.get("variety") or "").strip()
        or (selected_pending_visit.get("variety") or "").strip()
    )

    fenologia_real = (base_preview.get("fenologia_real") or "").strip() or None

    recommendation = base_preview.get("recommendation")
    if recommendation is None:
        recommendation = selected_pending_visit.get("recommendation") or ""
    recommendation = (recommendation or "").strip()

    payload = {
        "linked_pending_visit_id": linked_pending_visit_id,
        "client_id": client_id,
        "property_id": property_id,
        "plot_id": plot_id,
        "consultant_id": consultant_id,
        "date": base_preview.get("date") or selected_pending_visit.get("date"),
        "status": "done",
        "culture": culture,
        "variety": variety,
        "fenologia_real": fenologia_real,
        "recommendation": recommendation,
        "products": normalize_products_from_parsed(base_preview.get("products") or []),
        "latitude": base_preview.get("latitude"),
        "longitude": base_preview.get("longitude"),
        "source": "chatbot",
        "update_only_products": bool(base_preview.get("update_only_products")),
        "close_only": bool(close_only),
    }

    return payload


def apply_payload_to_existing_visit(visit, final_visit_payload: dict, close_only: bool = False):
    """
    Aplica atualização em visita já existente.

    REGRA IMPORTANTE:
    Quando o bot atualiza uma visita já existente, ele preserva os campos estruturais:
    - client_id
    - property_id
    - plot_id
    - consultant_id
    - culture
    - variety

    O bot deve alterar apenas:
    - date
    - status
    - fenologia_real
    - recommendation
    - latitude
    - longitude
    - products
    """
    if not visit:
        raise ValueError("Visita não encontrada para atualização")

    date_value = final_visit_payload.get("date")
    parsed_date = None

    if date_value:
        try:
            parsed_date = _date.fromisoformat(date_value)
        except Exception:
            parsed_date = parse_human_date(date_value)

    update_only_products = bool(final_visit_payload.get("update_only_products"))

    # ✅ Atualiza data apenas se veio válida
    if parsed_date:
        visit.date = parsed_date

    # ✅ Se for apenas atualização de produtos, não mexe no resto
    if update_only_products:
        replace_visit_products_from_payload(visit, final_visit_payload)
        db.session.add(visit)
        db.session.commit()
        return visit

    # ✅ Em visita existente editada pelo bot, preservar campos estruturais
    visit.status = "done"

    if final_visit_payload.get("fenologia_real"):
        visit.fenologia_real = final_visit_payload.get("fenologia_real")

    if close_only:
        if final_visit_payload.get("recommendation"):
            visit.recommendation = final_visit_payload.get("recommendation")
    else:
        # se veio recommendation, substitui; se não veio, preserva o valor atual
        if "recommendation" in final_visit_payload:
            new_rec = (final_visit_payload.get("recommendation") or "").strip()
            if new_rec:
                visit.recommendation = new_rec

    if final_visit_payload.get("latitude") is not None:
        visit.latitude = final_visit_payload.get("latitude")

    if final_visit_payload.get("longitude") is not None:
        visit.longitude = final_visit_payload.get("longitude")

    visit.source = final_visit_payload.get("source") or "chatbot"

    replace_visit_products_from_payload(visit, final_visit_payload)

    db.session.add(visit)
    db.session.commit()
    return visit


def create_visit_from_payload(final_visit_payload: dict):
    """
    Cria nova visita a partir do payload final.
    """
    client_id = final_visit_payload.get("client_id")
    if not client_id:
        raise ValueError("client_id é obrigatório para criar nova visita")

    date_value = final_visit_payload.get("date")
    parsed_date = None

    if date_value:
        try:
            parsed_date = _date.fromisoformat(date_value)
        except Exception:
            parsed_date = parse_human_date(date_value)

    visit = Visit(
        client_id=client_id,
        property_id=final_visit_payload.get("property_id"),
        plot_id=final_visit_payload.get("plot_id"),
        consultant_id=final_visit_payload.get("consultant_id"),
        date=parsed_date,
        recommendation=final_visit_payload.get("recommendation") or "",
        status="done",
        culture=final_visit_payload.get("culture") or "",
        variety=final_visit_payload.get("variety") or "",
        fenologia_real=final_visit_payload.get("fenologia_real") or None,
        latitude=final_visit_payload.get("latitude"),
        longitude=final_visit_payload.get("longitude"),
        source=final_visit_payload.get("source") or "chatbot",
    )

    if not visit.culture and visit.plot_id:
        planting = (
            Planting.query
            .filter_by(plot_id=visit.plot_id)
            .order_by(Planting.id.desc())
            .first()
        )
        if planting:
            visit.culture = planting.culture
            visit.variety = visit.variety or planting.variety

    db.session.add(visit)
    db.session.commit()

    replace_visit_products_from_payload(visit, final_visit_payload)
    db.session.commit()

    return visit


def build_same_cycle_visit_query(base_visit):
    q = Visit.query.filter(Visit.id != base_visit.id)

    # ✅ critério principal
    if getattr(base_visit, "planting_id", None):
        q = q.filter(Visit.planting_id == base_visit.planting_id)

        base_recommendation = (getattr(base_visit, "recommendation", None) or "").strip()
        if base_recommendation:
            q = q.filter(Visit.recommendation == base_recommendation)

        return q

    # ✅ sem planting_id, só segue se houver variedade
    base_variety = (getattr(base_visit, "variety", None) or "").strip()
    if not base_variety:
        return None

    q = q.filter(Visit.client_id == base_visit.client_id)

    if getattr(base_visit, "property_id", None) is not None:
        q = q.filter(Visit.property_id == base_visit.property_id)
    else:
        q = q.filter(Visit.property_id.is_(None))

    if getattr(base_visit, "plot_id", None) is not None:
        q = q.filter(Visit.plot_id == base_visit.plot_id)
    else:
        q = q.filter(Visit.plot_id.is_(None))

    base_culture = (getattr(base_visit, "culture", None) or "").strip()
    if base_culture:
        q = q.filter(Visit.culture == base_culture)

    q = q.filter(Visit.variety == base_variety)

    base_recommendation = (getattr(base_visit, "recommendation", None) or "").strip()
    if base_recommendation:
        q = q.filter(Visit.recommendation == base_recommendation)

    return q


def auto_close_previous_cycle_visits(current_visit):
    '''
    Fecha automaticamente visitas anteriores do mesmo ciclo/evento.

    REGRA NOVA:
    - usa planting_id como prioridade
    - se NÃO houver planting_id, só fecha se houver variedade preenchida
    - se não houver query segura, não fecha nada
    - só fecha visitas com data <= data atual
    - só fecha visitas ainda não concluídas
    '''
    if not current_visit:
        return []

    if not getattr(current_visit, "date", None):
        return []

    q = build_same_cycle_visit_query(current_visit)
    if q is None:
        return []

    previous_visits = (
        q.filter(Visit.date.isnot(None))
         .filter(Visit.date <= current_visit.date)
         .filter(Visit.status != "done")
         .order_by(Visit.date.asc(), Visit.id.asc())
         .all()
    )

    closed_ids = []

    for visit in previous_visits:
        visit.status = "done"
        db.session.add(visit)
        closed_ids.append(visit.id)

    return closed_ids


def start_new_visit_direct_confirmation(
    state,
    chat_message,
    visit_preview: dict,
    matched_client,
    matched_property=None,
):
    """
    Quando o bot já identificou que é uma nova visita e não há pendências,
    pula a etapa de pedir NOVA e vai direto para confirmação final.
    """
    action = "create_new_visit"

    final_visit_payload = {
        **(visit_preview or {}),
        "linked_pending_visit_id": None,
    }

    summary_text = build_visit_summary_text(
        action=action,
        final_visit_payload=final_visit_payload,
        selected_pending_visit=None,
        close_only=False,
    )

    state.pending_visit_suggestions_json = json.dumps([], ensure_ascii=False)
    state.visit_preview_json = json.dumps(
        build_guided_state_payload(
            action=action,
            final_visit_payload=final_visit_payload,
            selected_pending_visit=None,
            close_only=False,
        ),
        ensure_ascii=False
    )
    state.confirmation_text = summary_text
    state.status = "awaiting_final_confirmation"
    db.session.commit()

    intro_lines = [
        f"Não encontrei visitas pendentes para {matched_client.name if matched_client else 'este cliente'}.",
        "Vou considerar isso como NOVA visita.",
        "",
        "Confira o resumo abaixo e responda com:",
        "✅ CONFIRMAR",
        "❌ CANCELAR",
        "",
        summary_text
    ]

    send_result = send_telegram_message(
        chat_id=chat_message.chat_id,
        text="\n".join(intro_lines)
    )

    return jsonify({
        "ok": True,
        "message": "nova visita enviada direto para confirmação final",
        "confirmation_text": "\n".join(intro_lines),
        "send_result": send_result,
    }), 200





def handle_final_confirmation(chat_message, message_text: str, photo_info=None):
    """
    Trata o estado awaiting_final_confirmation.
    """
    state = ChatbotConversationState.query.filter_by(
        platform="telegram",
        chat_id=chat_message.chat_id,
        status="awaiting_final_confirmation"
    ).first()

    if not state:
        return None

    reply = parse_pending_reply(message_text)

    if not reply:
        edit_command = parse_summary_edit_command(message_text)
        if not edit_command:
            send_telegram_message(
                chat_id=chat_message.chat_id,
                text=(
                    "Não entendi sua resposta.\n\n"
                    "Responda com:\n"
                    "✅ CONFIRMAR\n"
                    "❌ CANCELAR\n"
                    "✏️ ALTERAR FENOLOGIA V10\n"
                    "📅 ALTERAR DATA hoje\n"
                    "💬 ALTERAR OBSERVACAO baixa incidência de pragas"
                )
            )
            return jsonify({
                "ok": True,
                "message": "aguardando confirmação final"
            }), 200

        try:
            preview_data = json.loads(state.visit_preview_json or "{}")
        except Exception:
            preview_data = {}

        final_visit_payload = preview_data.get("final_visit_payload") or {}
        selected_pending_visit = preview_data.get("selected_pending_visit") or {}
        action = preview_data.get("action") or "create_new_visit"
        close_only = bool(preview_data.get("close_only"))

        field = edit_command.get("field")
        value = (edit_command.get("value") or "").strip()

        if field == "date":
            parsed_iso = parse_date_flexible(value)
            if not parsed_iso:
                send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text="Não consegui interpretar a data. Exemplo: hoje, ontem, 24/02/2026 ou 2026-02-24."
                )
                return jsonify({
                    "ok": True,
                    "message": "data inválida no resumo final"
                }), 200
            final_visit_payload["date"] = parsed_iso

        elif field == "fenologia_real":
            if not is_valid_fenologia(value):
                send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text="Fenologia inválida. Exemplo: V4, V10, VT, R1."
                )
                return jsonify({
                    "ok": True,
                    "message": "fenologia inválida no resumo final"
                }), 200
            final_visit_payload["fenologia_real"] = value.upper()

        elif field == "recommendation":
            final_visit_payload["recommendation"] = value

        elif field == "culture":
            normalized_culture = normalize_culture_input(value)
            if not normalized_culture:
                send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text="Cultura inválida. Use: Milho, Soja ou Algodão."
                )
                return jsonify({
                    "ok": True,
                    "message": "cultura inválida no resumo final"
                }), 200
            final_visit_payload["culture"] = normalized_culture

        elif field == "variety":
            final_visit_payload["variety"] = value

        summary_text = build_visit_summary_text(
            action=action,
            final_visit_payload=final_visit_payload,
            selected_pending_visit=selected_pending_visit,
            close_only=close_only
        )

        state.visit_preview_json = json.dumps({
            "action": action,
            "final_visit_payload": final_visit_payload,
            "selected_pending_visit": selected_pending_visit,
            "close_only": close_only,
        }, ensure_ascii=False)
        state.confirmation_text = summary_text
        db.session.commit()

        send_telegram_message(
            chat_id=chat_message.chat_id,
            text=summary_text
        )

        return jsonify({
            "ok": True,
            "message": "resumo final atualizado"
        }), 200

    if reply["mode"] == "cancel_final":
        db.session.delete(state)
        db.session.commit()

        # ✅ limpa fotos pendentes ligadas a essa tentativa
        clear_pending_telegram_photos(chat_message.chat_id)

        send_telegram_message(
            chat_id=chat_message.chat_id,
            text="Operação cancelada."
        )

        return jsonify({
            "ok": True,
            "message": "confirmação final cancelada"
        }), 200

    if reply["mode"] != "confirm_final":
        send_telegram_message(
            chat_id=chat_message.chat_id,
            text=bot_phrase("confirm_or_cancel", "Para esta etapa, responda com CONFIRMAR ou CANCELAR.")
        )
        return jsonify({
            "ok": True,
            "message": "resposta inválida para confirmação final"
        }), 200

    try:
        preview_data = json.loads(state.visit_preview_json or "{}")
    except Exception:
        preview_data = {}

    action = preview_data.get("action") or "create_new_visit"
    base_preview = preview_data.get("final_visit_payload") or {}
    selected_pending_visit = preview_data.get("selected_pending_visit") or {}
    close_only = bool(preview_data.get("close_only"))

    final_visit_payload = build_final_visit_payload(
        base_preview=base_preview,
        selected_pending_visit=selected_pending_visit,
        resolved_consultant_id=base_preview.get("consultant_id") or 1,
        close_only=close_only,
    )

    update_only_products = bool(final_visit_payload.get("update_only_products"))

    try:
        visit = None

        if action == "use_existing_pending_visit":
            visit_id = (
                final_visit_payload.get("linked_pending_visit_id")
                or selected_pending_visit.get("id")
            )
            if not visit_id:
                raise ValueError("ID da visita pendente não encontrado")

            visit = Visit.query.get(visit_id)
            if not visit:
                raise ValueError(f"Visita {visit_id} não encontrada")

            visit = apply_payload_to_existing_visit(
                visit=visit,
                final_visit_payload=final_visit_payload,
                close_only=close_only,
            )

        elif action == "create_new_visit":
            visit = create_visit_from_payload(final_visit_payload)

        else:
            raise ValueError(f"Ação final inválida: {action}")

        auto_closed_ids = auto_close_previous_cycle_visits(visit)

        attached_count, attach_errors = attach_pending_telegram_photos_to_visit(
            chat_id=chat_message.chat_id,
            visit=visit
        )

        db.session.delete(state)
        db.session.commit()

        success_lines = []

        if update_only_products:
            success_lines.append(f"✅ Produtos atualizados na visita {visit.id}.")
        elif action == "use_existing_pending_visit":
            success_lines.append(f"✅ Visita {visit.id} atualizada e concluída com sucesso.")
        else:
            success_lines.append(f"✅ Nova visita criada com sucesso. ID {visit.id}.")

        if auto_closed_ids:
            success_lines.append(
                f"🔄 Também concluí automaticamente {len(auto_closed_ids)} visita(s) anterior(es) do mesmo ciclo."
            )

        if attached_count > 0:
            success_lines.append(f"📸 {attached_count} foto(s) vinculada(s) à visita.")

        if attach_errors:
            print("DEBUG attach_pending errors:", attach_errors)
            success_lines.append("⚠️ Algumas fotos não conseguiram ser anexadas.")

        send_telegram_message(
            chat_id=chat_message.chat_id,
            text="\n".join(success_lines)
        )

        return jsonify({
            "ok": True,
            "message": "confirmação final concluída",
            "visit_id": visit.id,
            "attached_photos": attached_count,
            "attach_errors": attach_errors,
        }), 200

    except Exception as e:
        db.session.rollback()
        print("❌ Erro ao confirmar visita final:", str(e))

        send_telegram_message(
            chat_id=chat_message.chat_id,
            text="Ocorreu um erro ao salvar a visita. Tente novamente."
        )

        return jsonify({
            "ok": False,
            "message": "erro ao confirmar visita final",
            "error": str(e),
        }), 200



def replace_visit_products_from_payload(visit, final_visit_payload: dict):
    from models import VisitProduct

    VisitProduct.query.filter_by(visit_id=visit.id).delete()

    for p in final_visit_payload.get("products") or []:
        raw_date = p.get("application_date")
        application_date = None

        if raw_date:
            try:
                application_date = _date.fromisoformat(raw_date)
            except Exception:
                application_date = None

        vp = VisitProduct(
            visit_id=visit.id,
            product_name=(p.get("product_name") or "").strip(),
            dose=(p.get("dose") or "").strip(),
            unit=(p.get("unit") or "").strip(),
            application_date=application_date,
        )
        db.session.add(vp)




def find_visit_by_explicit_id(visit_id: int, consultant_id: int | None = None):
    if not visit_id:
        return None

    q = Visit.query.filter_by(id=visit_id)

    if consultant_id:
        q = q.filter_by(consultant_id=consultant_id)

    return q.first()



def handle_priority_stateful_actions(chat_message, consultant, message_text: str):
    """
    Trata primeiro estados que dependem de resposta curta/numérica,
    para evitar conflito entre PDF, agenda semanal, visitas do mês etc.
    """
    active_state = ChatbotConversationState.query.filter_by(
        platform="telegram",
        chat_id=chat_message.chat_id
    ).first()

    if not active_state:
        return None

    # prioridade máxima: seleção de PDF
    if active_state.status in ("awaiting_pdf_visit_selection", "awaiting_pdf_confirmation"):
        return handle_pdf_flow(
            chat_message=chat_message,
            consultant=consultant,
            message_text=message_text,
        )

    # depois seleção de visitas do mês
    if active_state.status == "awaiting_month_visit_selection":
        return handle_month_visits_flow(
            chat_message=chat_message,
            consultant=consultant,
            message_text=message_text,
        )

    # depois confirmação final
    if active_state.status == "awaiting_final_confirmation":
        return handle_final_confirmation(
            chat_message=chat_message,
            message_text=message_text,
        )

    # depois agenda semanal
    if active_state.status == "awaiting_week_visit_selection":
        return handle_week_schedule_flow(
            chat_message=chat_message,
            consultant=consultant,
            resolved_consultant_id=consultant.id if consultant else 1,
            message_text=message_text,
        )

    return None


# ============================================================
# 🌾 CULTURAS, VARIEDADES, CONSULTOR
# ============================================================

@bp.route('/cultures', methods=['GET'])
def list_cultures():
    CULTURES = [
        {"id": 1, "name": "Milho"},
        {"id": 2, "name": "Soja"},
        {"id": 3, "name": "Algodão"},
    ]
    return jsonify(CULTURES), 200


@bp.route('/whatsapp/bindings', methods=['GET'])
def list_whatsapp_bindings():
    rows = WhatsAppContactBinding.query.order_by(WhatsAppContactBinding.id.asc()).all()
    return jsonify([r.to_dict() for r in rows]), 200


@bp.route('/whatsapp/bindings', methods=['POST'])
def create_whatsapp_binding():
    data = request.get_json() or {}

    phone_number = normalize_phone_number((data.get("phone_number") or "").strip())
    consultant_id = data.get("consultant_id")
    display_name = (data.get("display_name") or "").strip() or None

    if not phone_number:
        return jsonify(message="phone_number is required"), 400

    if consultant_id is None:
        return jsonify(message="consultant_id is required"), 400

    try:
        consultant_id = int(consultant_id)
    except (TypeError, ValueError):
        return jsonify(message="consultant_id must be an integer"), 400

    if consultant_id not in CONSULTANT_IDS:
        return jsonify(message="consultant not found"), 404

    existing = WhatsAppContactBinding.query.filter_by(phone_number=phone_number).first()
    if existing:
        return jsonify(message="phone_number already linked"), 409

    row = WhatsAppContactBinding(
        phone_number=phone_number,
        consultant_id=consultant_id,
        display_name=display_name,
        is_active=True,
    )

    db.session.add(row)
    db.session.commit()

    return jsonify(message="binding created", binding=row.to_dict()), 201


    

@bp.route('/whatsapp/webhook', methods=['GET'])
def whatsapp_webhook_verify():
    verify_token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    mode = request.args.get("hub.mode")

    expected_token = os.environ.get("WHATSAPP_VERIFY_TOKEN", "agrocrm_verify_token")

    if mode == "subscribe" and verify_token == expected_token:
        return challenge, 200

    return jsonify({"error": "verification failed"}), 403


@bp.route('/whatsapp/webhook', methods=['POST'])
def whatsapp_webhook_receive():
    payload = request.get_json(silent=True) or {}

    try:
        entry_list = payload.get("entry", [])
        saved = 0

        for entry in entry_list:
            changes = entry.get("changes", [])
            for change in changes:
                value = change.get("value", {})
                contacts = value.get("contacts", []) or []
                messages = value.get("messages", []) or []

                contact_name = None
                wa_from = None

                if contacts:
                    contact_name = contacts[0].get("profile", {}).get("name")
                    wa_from = contacts[0].get("wa_id")

                for msg in messages:
                    message_type = msg.get("type", "unknown")
                    wa_message_id = msg.get("id")
                    from_number = msg.get("from") or wa_from

                    text_content = None
                    media_id = None
                    mime_type = None

                    if message_type == "text":
                        text_content = (msg.get("text") or {}).get("body")

                    elif message_type == "image":
                        image_obj = msg.get("image") or {}
                        media_id = image_obj.get("id")
                        mime_type = image_obj.get("mime_type")

                    elif message_type == "audio":
                        audio_obj = msg.get("audio") or {}
                        media_id = audio_obj.get("id")
                        mime_type = audio_obj.get("mime_type")

                    existing = WhatsAppInboundMessage.query.filter_by(
                        wa_message_id=wa_message_id
                    ).first()

                    if existing:
                        continue

                    row = WhatsAppInboundMessage(
                        wa_message_id=wa_message_id,
                        phone_number=from_number or "",
                        contact_name=contact_name,
                        message_type=message_type,
                        text_content=text_content,
                        media_id=media_id,
                        mime_type=mime_type,
                        raw_payload=str(payload),
                        processing_status="received",
                    )
                    db.session.add(row)
                    saved += 1

        db.session.commit()
        return jsonify({"status": "ok", "saved": saved}), 200

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro no webhook WhatsApp: {e}")
        return jsonify({"error": str(e)}), 500



@bp.route('/telegram/test-send', methods=['POST'])
def telegram_test_send():
    try:
        data = request.get_json(silent=True) or {}
        chat_id = str(data.get("chat_id") or "").strip()
        text = (data.get("text") or "Teste do AgroCRM no Telegram").strip()

        if not chat_id:
            return jsonify({
                "ok": False,
                "error": "chat_id is required"
            }), 400

        result = send_telegram_message(chat_id=chat_id, text=text)

        return jsonify({
            "ok": True,
            "send_result": result
        }), 200

    except Exception as e:
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500       

@bp.route("/reports/monthly.xlsx", methods=["GET"])
def report_monthly_xlsx():
    """
    Relatório XLSX formatado.
    Aceita:
      - ?month=YYYY-MM   (ex: 2026-02)
    ou
      - ?start=YYYY-MM-DD&end=YYYY-MM-DD  (end inclusive)
    """
    try:
        from io import BytesIO
        from datetime import date as _date
        import datetime as _dt
        from collections import Counter, defaultdict

        from flask import request, jsonify, send_file

        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import LineChart, BarChart, PieChart, Reference

        # =========================
        # 1) Define intervalo
        # =========================
        month = request.args.get("month")
        start = request.args.get("start")
        end = request.args.get("end")

        if month:
            y, m = [int(x) for x in month.split("-")]
            start_date = _date(y, m, 1)
            if m == 12:
                next_month = _date(y + 1, 1, 1)
            else:
                next_month = _date(y, m + 1, 1)
            end_date = next_month - _dt.timedelta(days=1)
        else:
            if not start or not end:
                return jsonify(message="Informe ?month=YYYY-MM ou ?start=YYYY-MM-DD&end=YYYY-MM-DD"), 400
            start_date = _date.fromisoformat(start)
            end_date = _date.fromisoformat(end)

        # =========================
        # 2) Busca visitas
        # =========================
        visits = (
            Visit.query
            .filter(Visit.date >= start_date)
            .filter(Visit.date <= end_date)
            .order_by(Visit.date.asc().nullslast())
            .all()
        )

        client_ids = sorted({v.client_id for v in visits if v.client_id})
        prop_ids   = sorted({v.property_id for v in visits if v.property_id})
        plot_ids   = sorted({v.plot_id for v in visits if v.plot_id})

        clients_map = {c.id: c.name for c in Client.query.filter(Client.id.in_(client_ids)).all()} if client_ids else {}
        props_map   = {p.id: p.name for p in Property.query.filter(Property.id.in_(prop_ids)).all()} if prop_ids else {}
        plots_map   = {pl.id: pl.name for pl in Plot.query.filter(Plot.id.in_(plot_ids)).all()} if plot_ids else {}

        # =========================
        # 3) Cria workbook / sheets
        # =========================
        wb = Workbook()

        ws = wb.active
        ws.title = "Visitas"

        ws_dash = wb.create_sheet("Dashboard", 0)  # primeira aba
        ws2 = wb.create_sheet("Produtos")

        # =========================
        # 4) Estilos
        # =========================
        header_fill = PatternFill("solid", fgColor="14532D")
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(bold=True, size=14, color="14532D")
        bold_font = Font(bold=True)

        subheader_fill = PatternFill("solid", fgColor="E8F5E9")  # verde bem claro
        zebra_fill  = PatternFill("solid", fgColor="F8FAFC")
        zebra_fill2 = PatternFill("solid", fgColor="EEF2F7")
        # e ajuste as fontes das linhas para preto:
        row_font = Font(color="111827")



        status_planned = PatternFill("solid", fgColor="F59E0B")  # amarelo
        status_done = PatternFill("solid", fgColor="22C55E")     # verde
        status_canceled = PatternFill("solid", fgColor="EF4444") # vermelho
        status_font_dark = Font(color="0B1F17", bold=True)
        status_font_light = Font(color="FFFFFF", bold=True)

        dash_section_fill = PatternFill("solid", fgColor="E5E7EB")  # cinza claro
        dash_header_fill  = PatternFill("solid", fgColor="D1D5DB")  # cinza um pouco mais escuro
        dash_font = Font(bold=True, color="111827")                 # quase preto
        dash_center = Alignment(horizontal="center", vertical="center", wrap_text=True)



        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left = Alignment(horizontal="left", vertical="top", wrap_text=True)

        # BORDAS MAIS SUAVES (troque o seu thin/border por estes)
        thin_header = Side(style="thin", color="1F3A33")   # header um pouco mais visível
        hair_data   = Side(style="hair", color="16312B")   # linhas bem discretas

        border_header = Border(left=thin_header, right=thin_header, top=thin_header, bottom=thin_header)
        border_data   = Border(bottom=hair_data)          # só linha inferior, fica limpo


        kpi_fill = PatternFill("solid", fgColor="0B3A2E")
        kpi_label_fill = PatternFill("solid", fgColor="0F5132")
        kpi_font = Font(color="FFFFFF", bold=True, size=12)
        kpi_value_font = Font(color="FFFFFF", bold=True, size=18)

        def style_header(sheet, row_idx, max_col):
            for col in range(1, max_col + 1):
                cell = sheet.cell(row=row_idx, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center
                cell.border = border_header


        def br_date(d):
            if not d:
                return ""
            if isinstance(d, str):
                try:
                    d2 = _date.fromisoformat(d[:10])
                    return d2.strftime("%d/%m/%Y")
                except:
                    return d[:10]
            try:
                return d.strftime("%d/%m/%Y")
            except:
                return str(d)

        period_label = f"{start_date.strftime('%d/%m/%Y')} a {end_date.strftime('%d/%m/%Y')}"

        # ==========================================================
        # 5) DASHBOARD (KPIs + Gráficos) — Layout executivo
        # ==========================================================

        # -------------------------
        # Setup visual geral
        # -------------------------
        ws_dash.sheet_view.showGridLines = False
        ws_dash.sheet_view.zoomScale = 110
        ws_dash.page_setup.orientation = "landscape"
        ws_dash.page_setup.fitToWidth = 1
        ws_dash.page_setup.fitToHeight = 0

        # largura de colunas (painel usa A..N)
        for col in range(1, 15):  # A..N
            ws_dash.column_dimensions[get_column_letter(col)].width = 16

        # alturas (melhora leitura)
        ws_dash.row_dimensions[1].height = 30
        ws_dash.row_dimensions[2].height = 18
        ws_dash.row_dimensions[4].height = 18
        ws_dash.row_dimensions[5].height = 28
        ws_dash.row_dimensions[6].height = 28
        ws_dash.row_dimensions[7].height = 28
        ws_dash.row_dimensions[8].height = 28

        # -------------------------
        # Novos estilos do Dashboard
        # -------------------------
        dash_title_fill = PatternFill("solid", fgColor="0F5132")
        dash_title_font = Font(color="FFFFFF", bold=True, size=16)
        dash_sub_font = Font(color="1F2937", bold=True, size=11)

        kpi_fill = PatternFill("solid", fgColor="0B3A2E")
        kpi_label_fill = PatternFill("solid", fgColor="14532D")
        kpi_font = Font(color="FFFFFF", bold=True, size=11)
        kpi_value_font = Font(color="FFFFFF", bold=True, size=18)

        box_border = Border(
            left=Side(style="thin", color="1F3A33"),
            right=Side(style="thin", color="1F3A33"),
            top=Side(style="thin", color="1F3A33"),
            bottom=Side(style="thin", color="1F3A33"),
        )

        muted = Font(color="6B7280")
        num_font = Font(color="111827", bold=True)

        # -------------------------
        # Cabeçalho (banner)
        # -------------------------
        ws_dash.merge_cells("A1:N1")
        ws_dash["A1"] = "Painel Gerencial — NutriCRM (Relatório de Visitas)"
        ws_dash["A1"].fill = dash_title_fill
        ws_dash["A1"].font = dash_title_font
        ws_dash["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_dash["A1"].border = box_border

        ws_dash["A2"] = "Período:"
        ws_dash["A2"].font = dash_sub_font
        ws_dash["B2"] = period_label
        ws_dash["B2"].font = Font(color="111827", bold=True)

        ws_dash["D2"] = "Regra:"
        ws_dash["D2"].font = dash_sub_font
        ws_dash.merge_cells("E2:N2")
        ws_dash["E2"] = "Visita concluída - Meta = 5 visitas por cliente."
        ws_dash["E2"].font = muted

        # -------------------------
        # KPIs (linha de cards)
        # -------------------------
        total_visits = len(visits)

        def has_valid_photo(v) -> bool:
            photos = getattr(v, "photos", []) or []
            return any(getattr(p, "url", None) for p in photos)

        visits_with_photo = sum(1 for v in visits if has_valid_photo(v))
        real_completion = (visits_with_photo / total_visits) if total_visits else 0.0

        unique_clients = len({v.client_id for v in visits if v.client_id})
        total_clients = Client.query.count()
        coverage = (unique_clients / total_clients) if total_clients else 0

        unique_consultants = len({v.consultant_id for v in visits if v.consultant_id})
        avg_visits_per_consultant = (total_visits / unique_consultants) if unique_consultants else 0

        META_VISITAS_CLIENTE = 5

        # visitas válidas (com foto) por cliente
        photo_visits_by_client = Counter()
        for v in visits:
            if v.client_id and has_valid_photo(v):
                photo_visits_by_client[v.client_id] += 1

        done_units = sum(min(cnt, META_VISITAS_CLIENTE) for cnt in photo_visits_by_client.values())
        target_units = (total_clients or 0) * META_VISITAS_CLIENTE
        portfolio_progress = (done_units / target_units) if target_units else 0.0

        # Card helper com largura fixa de 2 colunas
        def make_kpi_block(col1, col2, row_top, title, value, fmt=None):
            # título
            ws_dash.merge_cells(f"{col1}{row_top}:{col2}{row_top}")
            tcell = ws_dash[f"{col1}{row_top}"]
            tcell.value = title
            tcell.fill = kpi_label_fill
            tcell.font = kpi_font
            tcell.alignment = Alignment(horizontal="center", vertical="center")
            tcell.border = box_border

            # valor
            ws_dash.merge_cells(f"{col1}{row_top+1}:{col2}{row_top+3}")
            vcell = ws_dash[f"{col1}{row_top+1}"]
            vcell.value = value
            vcell.fill = kpi_fill
            vcell.font = kpi_value_font
            vcell.alignment = Alignment(horizontal="center", vertical="center")
            vcell.border = box_border

            # aplicar formato (percent / number)
            if fmt:
                vcell.number_format = fmt

            # borda/fill nas células mescladas
            c1 = ord(col1) - 64
            c2 = ord(col2) - 64
            for rr in range(row_top+1, row_top+4):
                for cc in range(c1, c2+1):
                    cell = ws_dash.cell(rr, cc)
                    cell.fill = kpi_fill
                    cell.border = box_border

        # KPIs (7 cards) — 2 colunas cada
        # A-B / C-D / E-F / G-H / I-J / K-L / M-N
        make_kpi_block("A","B",4,"Total de visitas (todas)", total_visits, fmt="#,##0")
        make_kpi_block("C","D",4,"Visitas concluídas", visits_with_photo, fmt="#,##0")
        make_kpi_block("E","F",4,"Taxa de conclusão real", real_completion, fmt="0.0%")
        make_kpi_block("G","H",4,"Clientes atendidos (período)", unique_clients, fmt="#,##0")
        make_kpi_block("I","J",4,"Cobertura da carteira", coverage, fmt="0.0%")
        make_kpi_block("K","L",4,"Média visitas/consultor", round(avg_visits_per_consultant,1), fmt="0.0")
        make_kpi_block("M","N",4,"Meta 5 visitas (carteira)", portfolio_progress, fmt="0.0%")

        # trava cabeçalho
        ws_dash.freeze_panes = "A10"

        # -------------------------
        # Seções (títulos)
        # -------------------------
        def section_title(cell_ref, text):
            cell = ws_dash[cell_ref]
            cell.value = text
            cell.font = Font(bold=True, color="111827", size=12)
            cell.fill = PatternFill("solid", fgColor="E5E7EB")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = box_border

        # -------------------------
        # Tabela: Visitas por dia (A..B)
        # -------------------------
        section_title("A10", "Visitas por dia")
        ws_dash["A12"] = "Data"
        ws_dash["B12"] = "Visitas"
        for cell in ws_dash["A12:B12"][0]:
            cell.fill = dash_header_fill
            cell.font = dash_font
            cell.alignment = dash_center
            cell.border = border_header

        day_counts = defaultdict(int)
        for v in visits:
            if v.date:
                day_counts[v.date] += 1

        days_sorted = sorted(day_counts.keys())
        r = 13
        for d in days_sorted:
            ws_dash[f"A{r}"] = d.strftime("%d/%m/%Y")
            ws_dash[f"B{r}"] = day_counts[d]
            ws_dash[f"A{r}"].alignment = left
            ws_dash[f"B{r}"].alignment = dash_center
            ws_dash[f"B{r}"].number_format = "#,##0"
            r += 1
        end_row_days = r - 1

        # -------------------------
        # Tabela: Visitas por consultor (D..E)
        # -------------------------
        section_title("D10", "Visitas por consultor")
        ws_dash["D12"] = "Consultor"
        ws_dash["E12"] = "Visitas"
        for cell in ws_dash["D12:E12"][0]:
            cell.fill = dash_header_fill
            cell.font = dash_font
            cell.alignment = dash_center
            cell.border = border_header

        try:
            consultants_map = {c["id"]: c["name"] for c in CONSULTANTS}
        except:
            consultants_map = {}

        cons_counts = Counter(v.consultant_id for v in visits if v.consultant_id)
        r2 = 13
        for cid, cnt in cons_counts.most_common():
            ws_dash[f"D{r2}"] = consultants_map.get(cid, f"ID {cid}")
            ws_dash[f"E{r2}"] = cnt
            ws_dash[f"D{r2}"].alignment = left
            ws_dash[f"E{r2}"].alignment = dash_center
            ws_dash[f"E{r2}"].number_format = "#,##0"
            r2 += 1
        end_row_cons = r2 - 1

        # -------------------------
        # Tabela: Visitas por cultura (G..H)
        # -------------------------
        section_title("G10", "Visitas por cultura")
        ws_dash["G12"] = "Cultura"
        ws_dash["H12"] = "Visitas"
        for cell in ws_dash["G12:H12"][0]:
            cell.fill = dash_header_fill
            cell.font = dash_font
            cell.alignment = dash_center
            cell.border = border_header

        cult_counts = Counter()
        for v in visits:
            culture = v.culture or (v.planting.culture if getattr(v, "planting", None) else None)
            culture = (culture or "—").strip()
            cult_counts[culture] += 1

        r3 = 13
        for culture, cnt in cult_counts.most_common():
            ws_dash[f"G{r3}"] = culture
            ws_dash[f"H{r3}"] = cnt
            ws_dash[f"G{r3}"].alignment = left
            ws_dash[f"H{r3}"].alignment = dash_center
            ws_dash[f"H{r3}"].number_format = "#,##0"
            r3 += 1
        end_row_cult = r3 - 1

        # -------------------------
        # Tabela: Top 5 clientes (visitas com foto) (J..K)
        # -------------------------
        section_title("J10", "Top 5 clientes (visitas com foto)")
        ws_dash["J12"] = "Cliente"
        ws_dash["K12"] = "Concluídas"
        for cell in ws_dash["J12:K12"][0]:
            cell.fill = dash_header_fill
            cell.font = dash_font
            cell.alignment = dash_center
            cell.border = border_header

        client_counts = Counter()
        for v in visits:
            if v.client_id and has_valid_photo(v):
                client_counts[v.client_id] += 1

        top5 = client_counts.most_common(5)
        r4 = 13
        for cid, cnt in top5:
            ws_dash[f"J{r4}"] = clients_map.get(cid, f"Cliente {cid}")
            ws_dash[f"K{r4}"] = cnt
            ws_dash[f"J{r4}"].alignment = left
            ws_dash[f"K{r4}"].alignment = dash_center
            ws_dash[f"K{r4}"].number_format = "#,##0"
            r4 += 1
        end_row_top5 = r4 - 1


        # ==========================================================
        # 📊 Progresso meta por cliente (5 visitas com foto)
        # ==========================================================
        # bloco mais "executivo": Cliente | Concluídas | % | Barra
        start_meta_title_row = 40
        section_title(f"A{start_meta_title_row}", "Progresso da meta por cliente (5 visitas com foto)")
        ws_dash.merge_cells(f"A{start_meta_title_row}:N{start_meta_title_row}")

        ws_dash[f"A{start_meta_title_row+2}"] = "Cliente"
        ws_dash[f"B{start_meta_title_row+2}"] = "Concluídas"
        ws_dash[f"C{start_meta_title_row+2}"] = "% da meta"
        ws_dash[f"D{start_meta_title_row+2}"] = "Barra"

        for cell in ws_dash[f"A{start_meta_title_row+2}:D{start_meta_title_row+2}"][0]:
            cell.fill = dash_header_fill
            cell.font = dash_font
            cell.alignment = dash_center
            cell.border = border_header

        # escreve dados
        rmeta = start_meta_title_row + 3
        for cid, cnt in sorted(photo_visits_by_client.items(), key=lambda x: x[1], reverse=True):
            ws_dash[f"A{rmeta}"] = clients_map.get(cid, f"Cliente {cid}")
            ws_dash[f"B{rmeta}"] = cnt
            pct = min(cnt, META_VISITAS_CLIENTE) / META_VISITAS_CLIENTE
            ws_dash[f"C{rmeta}"] = pct
            ws_dash[f"C{rmeta}"].number_format = "0%"

            # estilos
            ws_dash[f"A{rmeta}"].alignment = left
            ws_dash[f"B{rmeta}"].alignment = dash_center
            ws_dash[f"C{rmeta}"].alignment = dash_center
            ws_dash[f"B{rmeta}"].number_format = "#,##0"

            for col in range(1, 5):
                ws_dash.cell(rmeta, col).border = border_data

            rmeta += 1

        end_row_meta = rmeta - 1

        # barra de progresso (DataBar)
        if end_row_meta >= (start_meta_title_row + 3):
            rule = DataBarRule(
                start_type="num", start_value=0,
                end_type="num", end_value=1,
                color="2DD36F", showValue=False
            )
            ws_dash.conditional_formatting.add(
                f"D{start_meta_title_row+3}:D{end_row_meta}", rule
            )

            # coluna D recebe o mesmo pct só pra barra funcionar bem
            for rr in range(start_meta_title_row+3, end_row_meta+1):
                ws_dash[f"D{rr}"] = ws_dash[f"C{rr}"].value
                ws_dash[f"D{rr}"].number_format = "0%"

        # ajuste largura das colunas do bloco meta
        ws_dash.column_dimensions["A"].width = 34
        ws_dash.column_dimensions["B"].width = 14
        ws_dash.column_dimensions["C"].width = 12
        ws_dash.column_dimensions["D"].width = 22



        # ==========================================================
        # 6) ABA VISITAS
        # ==========================================================
        ws["A1"] = "Relatório Mensal — Visitas Técnicas"
        ws["A1"].font = title_font

        ws["A2"] = "Período:"
        ws["A2"].font = bold_font
        ws["B2"] = period_label

        ws["A3"] = "Total de visitas:"
        ws["A3"].font = bold_font
        ws["B3"] = total_visits

        ws["A4"] = "Clientes atendidos:"
        ws["A4"].font = bold_font
        ws["B4"] = unique_clients


        # 🎨 Destaque do resumo (A1:B4)
        for r in range(1, 5):
            for c in range(1, 3):  # A..B
                cell = ws.cell(r, c)
                cell.border = border_data
                if r == 1:
                    cell.fill = subheader_fill
                    cell.font = Font(bold=True, size=14, color="14532D")
                else:
                    cell.fill = subheader_fill
                    if c == 1:
                        cell.font = bold_font
                    cell.alignment = left


        ws["A6"] = "Detalhamento de visitas"
        ws["A6"].font = bold_font

        headers = [
            "Data","Cliente","Propriedade","Talhão","Consultor",
            "Cultura","Variedade","Fenologia (observada)","Status","Observações"
        ]
        header_row = 6

        for i, h in enumerate(headers, start=1):
            ws.cell(row=header_row, column=i).value = h

        style_header(ws, header_row, len(headers))
        ws.freeze_panes = "A7"

        row_idx = header_row
        for v in visits:
            row_idx += 1

            client_name = clients_map.get(v.client_id, f"Cliente {v.client_id}")
            prop_name   = props_map.get(v.property_id, "") if v.property_id else ""
            plot_name   = plots_map.get(v.plot_id, "") if v.plot_id else ""

            culture = v.culture or (v.planting.culture if getattr(v, "planting", None) else "")
            variety = v.variety or (v.planting.variety if getattr(v, "planting", None) else "")

            try:
                consultant_name = next((c["name"] for c in CONSULTANTS if c["id"] == v.consultant_id), "")
            except:
                consultant_name = ""

            fenologia_obs = v.fenologia_real or ""
            status = (v.status or "").strip()
            obs = (v.recommendation or "").strip()

            ws.append([
                br_date(v.date),
                client_name,
                prop_name,
                plot_name,
                consultant_name,
                culture or "",
                variety or "",
                fenologia_obs,
                status,
                obs,
            ])

            # 🎨 Zebra (alternando fundo)
            row_fill = zebra_fill if (row_idx % 2 == 0) else zebra_fill2

            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = border_data
                cell.fill = row_fill
                cell.alignment = left if col in (2, 3, 4, 10) else center

            # 🎨 Status com cor (coluna 9)
            st = (status or "").lower().strip()
            status_cell = ws.cell(row=row_idx, column=9)

            if st in ("planned", "planejada", "pendente"):
                status_cell.fill = status_planned
                status_cell.font = status_font_dark
            elif st in ("done", "realizada", "concluida", "concluída"):
                status_cell.fill = status_done
                status_cell.font = status_font_dark
            elif st in ("canceled", "cancelada", "cancelado"):
                status_cell.fill = status_canceled
                status_cell.font = status_font_light
            else:
                status_cell.font = Font(color="FFFFFF", bold=True)


        # ✅ filtro cobrindo até a última linha de dados
        last_row_visits = ws.max_row
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row_visits}"

        # ==========================================================
        # 7) ABA PRODUTOS
        # ==========================================================
        ws2["A1"] = "Produtos aplicados nas visitas"
        ws2["A1"].font = title_font
        ws2["A2"] = "Período:"
        ws2["A2"].font = bold_font
        ws2["B2"] = period_label

        prod_headers = ["Data visita", "Cliente", "Produto", "Dose", "Unidade", "Data aplicação"]
        ws2_header_row = 4


        for i, h in enumerate(prod_headers, start=1):
            ws2.cell(row=ws2_header_row, column=i).value = h

        style_header(ws2, ws2_header_row, len(prod_headers))
        ws2.freeze_panes = "A5"

        prod_row = ws2_header_row
        for v in visits:
            client_name = clients_map.get(v.client_id, f"Cliente {v.client_id}")
            v_date = br_date(v.date)

            prods = getattr(v, "products", []) or []
            for p in prods:
                prod_row += 1
                ws2.append([
                    v_date,
                    client_name,
                    getattr(p, "product_name", "") or "",
                    getattr(p, "dose", "") or "",
                    getattr(p, "unit", "") or "",
                    br_date(getattr(p, "application_date", None)),
                ])

                row_fill = zebra_fill if (prod_row % 2 == 0) else zebra_fill2

                for col in range(1, len(prod_headers) + 1):
                    cell = ws2.cell(row=prod_row, column=col)
                    cell.border = border_data
                    cell.fill = row_fill
                    cell.alignment = left if col in (2, 3) else center


        # ✅ filtro cobrindo até a última linha de dados
        last_row_prod = ws2.max_row
        ws2.auto_filter.ref = f"A{ws2_header_row}:{get_column_letter(len(prod_headers))}{last_row_prod}"

        # ==========================================================
        # 8) Ajusta largura colunas
        # ==========================================================
        def autosize(sheet, max_col, min_w=12, max_w=44):
            for col in range(1, max_col + 1):
                letter = get_column_letter(col)
                max_len = 0
                for cell in sheet[letter]:
                    val = "" if cell.value is None else str(cell.value)
                    max_len = max(max_len, len(val))
                sheet.column_dimensions[letter].width = max(min_w, min(max_w, max_len + 2))

        autosize(ws, len(headers))
        autosize(ws2, len(prod_headers))

        # ==========================================================
        # 9) Exporta arquivo
        # ==========================================================
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"relatorio_visitas_{start_date.isoformat()}_a_{end_date.isoformat()}.xlsx"
        return send_file(
            bio,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print("⚠️ erro report_monthly_xlsx:", e)
        return jsonify(error=str(e)), 500





@bp.route('/varieties', methods=['GET'])
def list_varieties():
    # opcional: filtrar por culture_id (mais seguro e rápido)
    culture_id = request.args.get("culture_id", type=int)

    q = Variety.query
    if culture_id:
        q = q.filter(Variety.culture_id == culture_id)

    rows = q.order_by(Variety.id.asc()).all()

    # devolve também o nome da cultura (útil no front)
    culture_map = {c.id: c.name for c in Culture.query.all()}

    return jsonify([
        {
            "id": v.id,
            "culture_id": v.culture_id,
            "culture": culture_map.get(v.culture_id, ""),
            "name": v.name,
        }
        for v in rows
    ]), 200



@bp.route('/telegram/setup-link-codes', methods=['POST'])
def setup_telegram_link_codes():
    try:
        mapping = {
            "Jhonatan": "JHONATAN123",
            "Felipe": "FELIPE123",
            "Everton": "EVERTON123",
            "Pedro": "PEDRO123",
            "Alexandre": "ALEXANDRE123",
        }

        consultants = Consultant.query.all()
        updated = []

        for consultant in consultants:
            code = mapping.get((consultant.name or "").strip())
            if not code:
                continue

            consultant.telegram_link_code = code
            updated.append({
                "id": consultant.id,
                "name": consultant.name,
                "telegram_link_code": consultant.telegram_link_code,
            })

        db.session.commit()

        return jsonify({
            "ok": True,
            "updated": updated,
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({
            "ok": False,
            "error": str(e),
        }), 500




@bp.route('/consultants', methods=['GET'])
def list_consultants():
    return jsonify(CONSULTANTS), 200


# ============================================================
# 🔵 Endpoint de teste usado pelo APK (detectar conexão real)
# ============================================================
@bp.route("/ping", methods=["GET"])
def ping():
    return "pong", 200

def is_confirmation_reply(text: str) -> bool:
    if not text:
        return False

    value = text.strip().upper()
    return value == "NOVA" or value.isdigit()







def parse_date_flexible(value: str):
    if not value:
        return None

    raw = value.strip()
    normalized = normalize_lookup_text(raw)
    today = get_local_today()

    if normalized == "hoje":
        return today.isoformat()

    if normalized in ("amanha", "amanhã"):
        return (today + _timedelta(days=1)).isoformat()

    # só dia do mês atual, ex: "15"
    if re.match(r"^\d{1,2}$", raw):
        day = int(raw)
        try:
            return _date(today.year, today.month, day).isoformat()
        except ValueError:
            return None

    # formato DD/MM/YYYY
    match_br = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", raw)
    if match_br:
        dd, mm, yyyy = match_br.groups()
        try:
            return _date(int(yyyy), int(mm), int(dd)).isoformat()
        except ValueError:
            return None

    # formato DD/MM (assume ano atual)
    match_br_short = re.match(r"^(\d{1,2})/(\d{1,2})$", raw)
    if match_br_short:
        dd, mm = match_br_short.groups()
        try:
            return _date(today.year, int(mm), int(dd)).isoformat()
        except ValueError:
            return None

    # formato YYYY-MM-DD
    match_iso = re.match(r"^(20\d{2})-(\d{2})-(\d{2})$", raw)
    if match_iso:
        try:
            return _date.fromisoformat(raw).isoformat()
        except ValueError:
            return None

    return None




def visit_has_valid_photo(visit) -> bool:
    if not visit:
        return False

    photos = getattr(visit, "photos", []) or []
    return any(getattr(p, "url", None) for p in photos)




def is_stale_clients_request(text: str) -> bool:
    if not text:
        return False

    normalized = normalize_lookup_text(text)

    triggers = [
        "clientes mais atrasados",
        "clientes atrasados",
        "clientes sem visita",
        "clientes ha mais tempo sem visita",
        "clientes há mais tempo sem visita",
        "clientes sem visita com foto",
        "ranking de clientes atrasados",
        "ranking clientes atrasados",
        "me mostra os clientes mais atrasados",
        "quais clientes estao mais atrasados",
        "quais clientes estão mais atrasados",
        "visitas do mes atrasadas",
        "visitas do mês atrasadas",
        "visitas atrasadas",
    ]

    return any(trigger in normalized for trigger in triggers)



def find_stale_clients_ranking(consultant_id: int | None = None, limit: int = 15):
    """
    Ranking do mais atrasado para o menos atrasado.

    Regra:
    - entra no ranking apenas cliente que já tenha pelo menos 1 visita lançada
    - conta como última visita válida apenas visita COM FOTO
    - cliente sem visita com foto fica fora do ranking
    """
    today = get_local_today()

    clients = Client.query.order_by(Client.name.asc()).all()
    ranking = []

    for client in clients:
        q = Visit.query.filter(Visit.client_id == client.id)

        if consultant_id:
            q = q.filter(Visit.consultant_id == consultant_id)

        visits = q.order_by(Visit.date.desc().nullslast(), Visit.id.desc()).all()

        # ✅ cliente sem nenhuma visita lançada não entra no ranking
        if not visits:
            continue

        valid_photo_visit = None
        for visit in visits:
            if visit_has_valid_photo(visit):
                valid_photo_visit = visit
                break

        # ✅ cliente com visitas lançadas, mas sem nenhuma visita com foto, fica fora do ranking
        if not valid_photo_visit:
            continue

        if not valid_photo_visit.date:
            continue

        days_without = (today - valid_photo_visit.date).days

        ranking.append({
            "client_id": client.id,
            "client_name": client.name,
            "last_valid_visit_date": valid_photo_visit.date,
            "days_without_valid_visit": days_without,
            "last_valid_culture": (
                valid_photo_visit.culture if valid_photo_visit.culture else "—"
            ),
            "last_valid_visit_id": valid_photo_visit.id,
        })

    ranking.sort(
        key=lambda x: (
            x["days_without_valid_visit"],
            x["client_name"] or ""
        ),
        reverse=True
    )

    return ranking[:limit]



def build_stale_clients_ranking_text(consultant_name: str, items: list) -> str:
    if not items:
        return (
            f"📊 Clientes há mais tempo sem visita válida\n"
            f"Consultor: {consultant_name}\n\n"
            f"Nenhum cliente com visita válida encontrada."
        )

    lines = [
        "📊 Clientes há mais tempo sem visita válida",
        f"Consultor: {consultant_name}",
        "Critério: conta apenas visita realizada com foto.",
        ""
    ]

    for idx, item in enumerate(items, start=1):
        client_name = item.get("client_name") or f"Cliente {item.get('client_id')}"
        days_without = item.get("days_without_valid_visit")
        culture = item.get("last_valid_culture") or "—"

        lines.append(
            f"{idx}. {client_name} - {days_without} dias desde a última visita - {culture}"
        )

    return "\n".join(lines)




def handle_week_schedule_flow(chat_message, consultant, resolved_consultant_id: int, message_text: str):
    """
    Isola:
    - seleção numérica / lançar visita da agenda
    - conclusão simples da agenda
    - pedido de agenda da semana
    """
    active_state = ChatbotConversationState.query.filter_by(
        platform="telegram",
        chat_id=chat_message.chat_id
    ).first()

    if active_state and active_state.status in (
        "awaiting_confirmation",
        "awaiting_client_confirmation",
        "awaiting_month_visit_selection",
    ):
        week_action = None
    else:
        week_action = parse_week_visit_action(message_text)

    if week_action:
        state = ChatbotConversationState.query.filter_by(
            platform="telegram",
            chat_id=chat_message.chat_id,
            status="awaiting_week_visit_selection"
        ).first()

        if not state:
            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text="Não encontrei uma agenda semanal ativa. Peça primeiro sua agenda da semana."
            )
            return jsonify({
                "ok": True,
                "message": "nenhum estado ativo para seleção da agenda",
                "send_result": send_result,
            }), 200

        week_candidates = json.loads(state.pending_visit_suggestions_json or "[]")
        idx = week_action["index"]

        if idx < 0 or idx >= len(week_candidates):
            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text="Número inválido da agenda. Revise a lista e tente novamente."
            )
            return jsonify({
                "ok": True,
                "message": "índice inválido da agenda",
                "send_result": send_result,
            }), 200

        selected_visit = week_candidates[idx]

        if not selected_visit:
            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text="Não consegui identificar a visita selecionada."
            )
            return jsonify({
                "ok": False,
                "message": "visita selecionada inválida",
                "send_result": send_result,
            }), 400

        action = "use_existing_pending_visit"
        final_visit_payload = {
            "client_id": selected_visit.get("client_id"),
            "property_id": selected_visit.get("property_id"),
            "plot_id": selected_visit.get("plot_id"),
            "consultant_id": resolved_consultant_id,
            "date": selected_visit.get("date"),
            "status": "done",
            "culture": selected_visit.get("culture") or "",
            "variety": selected_visit.get("variety") or "",
            "fenologia_real": selected_visit.get("fenologia_real"),
            "recommendation": selected_visit.get("recommendation") or "",
            "products": [],
            "latitude": None,
            "longitude": None,
            "generate_schedule": False,
            "source": "chatbot",
            "linked_pending_visit_id": selected_visit.get("id"),
        }

        if week_action["intent"] == "complete_week_visit":
            summary_text = build_visit_summary_text(
                action=action,
                final_visit_payload=final_visit_payload,
                selected_pending_visit=selected_visit,
                close_only=True
            )

            state.visit_preview_json = json.dumps(
                build_guided_state_payload(
                    action=action,
                    final_visit_payload=final_visit_payload,
                    selected_pending_visit=selected_visit,
                    close_only=True,
                ),
                ensure_ascii=False
            )
            state.confirmation_text = summary_text
            state.status = "awaiting_final_confirmation"
            db.session.commit()

            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text=summary_text
            )

            return jsonify({
                "ok": True,
                "message": "resumo enviado para conclusão da visita da agenda",
                "summary_text": summary_text,
                "send_result": send_result,
            }), 200

        state.visit_preview_json = json.dumps(
            build_guided_state_payload(
                action=action,
                final_visit_payload=final_visit_payload,
                selected_pending_visit=selected_visit,
                close_only=False,
            ),
            ensure_ascii=False
        )
        state.status = "awaiting_fenologia"
        db.session.commit()

        send_result = send_telegram_message(
            chat_id=chat_message.chat_id,
            text="🌿 Informe a fenologia observada.\nExemplo: V4, V5, R1"
        )

        return jsonify({
            "ok": True,
            "message": "iniciado fluxo para atualizar visita da agenda",
            "send_result": send_result,
        }), 200

    if is_week_schedule_request(message_text):
        if not consultant:
            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text=(
                    "Seu Telegram ainda não está vinculado a um consultor do AgroCRM.\n"
                    "Use /start e depois /vincular SEU_CODIGO."
                )
            )
            return jsonify({
                "ok": False,
                "message": "consultor não vinculado",
                "send_result": send_result,
            }), 400

        week_visits = find_consultant_pending_visits_for_week(
            consultant_id=consultant.id
        )

        response_text = build_week_schedule_text(
            consultant_name=consultant.name,
            visits=week_visits
        )

        state = ChatbotConversationState.query.filter_by(
            platform="telegram",
            chat_id=chat_message.chat_id
        ).first()

        if not state:
            state = ChatbotConversationState(
                platform="telegram",
                chat_id=chat_message.chat_id,
            )
            db.session.add(state)

        state.pending_visit_suggestions_json = json.dumps(
            [
                {
                    "id": v.id,
                    "client_id": v.client_id,
                    "property_id": v.property_id,
                    "plot_id": v.plot_id,
                    "culture": v.culture,
                    "variety": v.variety,
                    "date": v.date.isoformat() if v.date else None,
                    "recommendation": v.recommendation or "",
                    "fenologia_real": v.fenologia_real,
                    "status": v.status,
                }
                for v in week_visits
            ],
            ensure_ascii=False
        )
        state.confirmation_text = response_text
        state.status = "awaiting_week_visit_selection"
        db.session.commit()

        send_result = send_telegram_message(
            chat_id=chat_message.chat_id,
            text=response_text
        )

        return jsonify({
            "ok": True,
            "message": "agenda semanal enviada",
            "consultant": {
                "id": consultant.id,
                "name": consultant.name,
            },
            "visits_count": len(week_visits),
            "response_text": response_text,
            "send_result": send_result,
        }), 200

    return None


def handle_stale_clients_ranking_flow(chat_message, consultant, message_text: str):
    if not is_stale_clients_request(message_text):
        return None

    if not consultant:
        send_result = send_telegram_message(
            chat_id=chat_message.chat_id,
            text=(
                "Seu Telegram ainda não está vinculado a um consultor do AgroCRM.\n"
                "Use /start e depois /vincular SEU_CODIGO."
            )
        )
        return jsonify({
            "ok": False,
            "message": "consultor não vinculado",
            "send_result": send_result,
        }), 400

    ranking = find_stale_clients_ranking(
        consultant_id=consultant.id,
        limit=15
    )

    response_text = build_stale_clients_ranking_text(
        consultant_name=consultant.name,
        items=ranking
    )

    send_result = send_telegram_message(
        chat_id=chat_message.chat_id,
        text=response_text
    )

    return jsonify({
        "ok": True,
        "message": "ranking de clientes atrasados enviado",
        "items_count": len(ranking),
        "response_text": response_text,
        "send_result": send_result,
    }), 200



def handle_pdf_flow(chat_message, consultant, message_text: str):
    """
    Isola:
    - PDF da última visita
    - pedido manual de PDF
    - escolha da(s) visita(s) para gerar PDF
    - confirmação para gerar PDF da última visita concluída
    """
    client_reference = parse_pdf_client_reference(message_text)
    if client_reference:
        if not consultant:
            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text=bot_phrase(
                    "consultant_not_bound",
                    "Seu Telegram ainda não está vinculado a um consultor do AgroCRM."
                )
            )
            return jsonify({
                "ok": False,
                "message": "consultor não vinculado",
                "send_result": send_result,
            }), 400

        visit = find_last_completed_visit_for_client_reference(
            consultant_id=consultant.id,
            client_name=client_reference,
        )

        if not visit:
            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text=f"Não achei PDF recente para {client_reference}."
            )
            return jsonify({
                "ok": True,
                "message": "pdf por cliente não encontrado",
                "send_result": send_result,
            }), 200

        try:
            buffer, filename = build_visit_pdf_file(visit.id)
            pdf_bytes = buffer.getvalue()

            send_result = send_telegram_document(
                chat_id=chat_message.chat_id,
                file_bytes=pdf_bytes,
                filename=filename,
                caption=f"📄 PDF da visita {visit.id}"
            )

            return jsonify({
                "ok": True,
                "message": "pdf por cliente enviado",
                "visit_id": visit.id,
                "send_result": send_result,
            }), 200

        except Exception as e:
            return jsonify({
                "ok": False,
                "error": str(e)
            }), 500

    # =========================================================
    # PDF da última visita
    # =========================================================
    if is_last_pdf_request(message_text):
        if not consultant:
            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text=(
                    "Seu Telegram ainda não está vinculado a um consultor do AgroCRM.\n"
                    "Use /start e depois /vincular SEU_CODIGO."
                )
            )
            return jsonify({
                "ok": False,
                "message": "consultor não vinculado",
                "send_result": send_result,
            }), 400

        recent_visits = find_last_completed_visits_for_consultant(
            consultant.id,
            limit=1
        )
        last_visit = recent_visits[0] if recent_visits else None

        if not last_visit:
            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text="Não encontrei nenhuma visita concluída recente para gerar PDF."
            )
            return jsonify({
                "ok": True,
                "message": "nenhuma visita concluída encontrada",
                "send_result": send_result,
            }), 200

        try:
            buffer, filename = build_visit_pdf_file(last_visit.id)
            pdf_bytes = buffer.getvalue()

            send_result = send_telegram_document(
                chat_id=chat_message.chat_id,
                file_bytes=pdf_bytes,
                filename=filename,
                caption=f"📄 PDF da última visita ({last_visit.id})"
            )

            return jsonify({
                "ok": True,
                "message": "pdf da última visita enviado",
                "visit_id": last_visit.id,
                "send_result": send_result,
            }), 200

        except Exception as e:
            return jsonify({
                "ok": False,
                "error": str(e)
            }), 500

    # =========================================================
    # Pedido manual de PDF
    # =========================================================
    if is_pdf_request(message_text):
        if not consultant:
            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text=(
                    "Seu Telegram ainda não está vinculado a um consultor do AgroCRM.\n"
                    "Use /start e depois /vincular SEU_CODIGO."
                )
            )
            return jsonify({
                "ok": False,
                "message": "consultor não vinculado",
                "send_result": send_result,
            }), 400

        recent_visits = find_last_completed_visits_for_consultant(
            consultant_id=consultant.id,
            limit=6
        )

        response_text = build_pdf_visit_selection_text(recent_visits)

        state = ChatbotConversationState.query.filter_by(
            platform="telegram",
            chat_id=chat_message.chat_id
        ).first()

        if not state:
            state = ChatbotConversationState(
                platform="telegram",
                chat_id=chat_message.chat_id,
            )
            db.session.add(state)

        state.pending_visit_suggestions_json = json.dumps(
            [{"id": v.id} for v in recent_visits],
            ensure_ascii=False
        )
        state.confirmation_text = response_text
        state.status = "awaiting_pdf_visit_selection"
        db.session.commit()

        send_result = send_telegram_message(
            chat_id=chat_message.chat_id,
            text=response_text
        )

        return jsonify({
            "ok": True,
            "message": "lista de visitas para pdf enviada",
            "visits_count": len(recent_visits),
            "response_text": response_text,
            "send_result": send_result,
        }), 200

    # =========================================================
    # Escolha da(s) visita(s) para gerar PDF
    # =========================================================
    state = ChatbotConversationState.query.filter_by(
        platform="telegram",
        chat_id=chat_message.chat_id,
        status="awaiting_pdf_visit_selection"
    ).first()

    if state:
        selected_indexes = parse_pdf_selection(message_text)

        if selected_indexes is None or not selected_indexes:
            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text="Opção inválida. Responda com um número ou vários, como 1,3 ou 1 3 5."
            )
            return jsonify({
                "ok": True,
                "message": "opção inválida para pdf",
                "send_result": send_result,
            }), 200

        pdf_candidates = json.loads(state.pending_visit_suggestions_json or "[]")

        invalid = [idx for idx in selected_indexes if idx < 0 or idx >= len(pdf_candidates)]
        if invalid:
            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text="Uma ou mais opções são inválidas. Revise os números e tente novamente."
            )
            return jsonify({
                "ok": True,
                "message": "índices inválidos para pdf",
                "send_result": send_result,
            }), 200

        results = []
        sent_count = 0

        for idx in selected_indexes:
            selected = pdf_candidates[idx]
            visit_id = selected.get("id")

            try:
                buffer, filename = build_visit_pdf_file(visit_id)
                pdf_bytes = buffer.getvalue()

                send_result = send_telegram_document(
                    chat_id=chat_message.chat_id,
                    file_bytes=pdf_bytes,
                    filename=filename,
                    caption=f"📄 PDF da visita {visit_id}"
                )

                results.append({
                    "visit_id": visit_id,
                    "send_result": send_result,
                })

                if send_result.get("ok"):
                    sent_count += 1

            except Exception as e:
                results.append({
                    "visit_id": visit_id,
                    "error": str(e),
                })

        state.status = "completed"
        db.session.commit()

        return jsonify({
            "ok": True,
            "message": "pdf(s) processado(s)",
            "requested_count": len(selected_indexes),
            "sent_count": sent_count,
            "results": results,
        }), 200

    # =========================================================
    # Confirmação para gerar PDF da última visita concluída
    # =========================================================
    state = ChatbotConversationState.query.filter_by(
        platform="telegram",
        chat_id=chat_message.chat_id,
        status="awaiting_pdf_confirmation"
    ).first()

    if state:
        yes_no = parse_yes_no(message_text)

        if yes_no is None:
            send_telegram_message(
                chat_id=chat_message.chat_id,
                text="📄 Resposta inválida. Responda com SIM ou NÃO."
            )
            return jsonify({
                "ok": True,
                "message": "resposta inválida para pdf"
            }), 200

        if yes_no is False:
            state.status = "completed"
            db.session.commit()

            send_telegram_message(
                chat_id=chat_message.chat_id,
                text="Certo. PDF não gerado."
            )
            return jsonify({
                "ok": True,
                "message": "pdf não gerado"
            }), 200

        stored_data = json.loads(state.visit_preview_json or "{}")
        visit_id = stored_data.get("last_completed_visit_id")

        if not visit_id:
            state.status = "completed"
            db.session.commit()

            send_telegram_message(
                chat_id=chat_message.chat_id,
                text="Não consegui identificar a última visita para gerar o PDF."
            )
            return jsonify({
                "ok": True,
                "message": "visit_id ausente para pdf"
            }), 200

        try:
            buffer, filename = build_visit_pdf_file(visit_id)
            pdf_bytes = buffer.getvalue()

            send_result = send_telegram_document(
                chat_id=chat_message.chat_id,
                file_bytes=pdf_bytes,
                filename=filename,
                caption=f"📄 PDF da visita {visit_id}"
            )

            state.status = "completed"
            db.session.commit()

            return jsonify({
                "ok": True,
                "message": "pdf enviado",
                "visit_id": visit_id,
                "send_result": send_result,
            }), 200

        except Exception as e:
            send_telegram_message(
                chat_id=chat_message.chat_id,
                text=f"Erro ao gerar o PDF da visita {visit_id}."
            )
            return jsonify({
                "ok": False,
                "error": str(e)
            }), 500

    return None


def handle_month_visits_flow(chat_message, consultant, message_text: str):
    """
    Isola:
    - pedido de visitas do mês
    - seleção de visita da lista do mês
    - ação PDF X dentro da lista do mês
    - abertura da visita mensal para edição / resumo
    """

    # =========================================================
    # Visitas do mês
    # =========================================================
    if is_month_visits_request(message_text):
        if not consultant:
            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text=(
                    "Seu Telegram ainda não está vinculado a um consultor do AgroCRM.\n"
                    "Use /start e depois /vincular SEU_CODIGO."
                )
            )
            return jsonify({
                "ok": False,
                "message": "consultor não vinculado",
                "send_result": send_result,
            }), 400

        filter_mode = parse_month_visit_filter(message_text)

        month_visits = find_consultant_visits_for_month(
            consultant_id=consultant.id,
            filter_mode=filter_mode
        )

        response_text = build_month_visits_text(
            consultant_name=consultant.name,
            visits=month_visits,
            filter_mode=filter_mode
        )

        state = ChatbotConversationState.query.filter_by(
            platform="telegram",
            chat_id=chat_message.chat_id
        ).first()

        if not state:
            state = ChatbotConversationState(
                platform="telegram",
                chat_id=chat_message.chat_id,
            )
            db.session.add(state)

        state.pending_visit_suggestions_json = json.dumps(
            [
                {
                    "id": v.id,
                    "client_id": v.client_id,
                    "property_id": v.property_id,
                    "plot_id": v.plot_id,
                    "culture": v.culture,
                    "variety": v.variety,
                    "date": v.date.isoformat() if v.date else None,
                    "recommendation": v.recommendation or "",
                    "fenologia_real": v.fenologia_real,
                    "status": v.status,
                }
                for v in month_visits
            ],
            ensure_ascii=False
        )
        state.confirmation_text = response_text
        state.status = "awaiting_month_visit_selection"
        db.session.commit()

        send_result = send_telegram_message(
            chat_id=chat_message.chat_id,
            text=response_text
        )

        return jsonify({
            "ok": True,
            "message": "visitas do mês enviadas",
            "filter_mode": filter_mode,
            "visits_count": len(month_visits),
            "response_text": response_text,
            "send_result": send_result,
        }), 200

    # =========================================================
    # Seleção de visita da lista do mês
    # =========================================================
    month_action = parse_month_visit_action(message_text)

    if month_action:
        state = ChatbotConversationState.query.filter_by(
            platform="telegram",
            chat_id=chat_message.chat_id,
            status="awaiting_month_visit_selection"
        ).first()

        if state:
            if month_action["mode"] == "cancel":
                state.status = "cancelled"
                db.session.commit()

                send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text="Operação cancelada com sucesso."
                )
                return jsonify({
                    "ok": True,
                    "message": "seleção mensal cancelada"
                }), 200

            candidates = json.loads(state.pending_visit_suggestions_json or "[]")
            idx = month_action.get("index", -1)

            if idx < 0 or idx >= len(candidates):
                send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text="Número inválido da lista do mês. Revise a lista e tente novamente."
                )
                return jsonify({
                    "ok": True,
                    "message": "índice inválido da lista do mês"
                }), 200

            selected = candidates[idx]
            visit = Visit.query.get(selected["id"])

            if not visit:
                send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text="Não consegui localizar essa visita."
                )
                return jsonify({
                    "ok": True,
                    "message": "visita da lista mensal não encontrada"
                }), 200

            if month_action["mode"] == "pdf":
                try:
                    buffer, filename = build_visit_pdf_file(visit.id)
                    pdf_bytes = buffer.getvalue()
                except Exception as e:
                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text="Não consegui gerar o PDF dessa visita agora."
                    )
                    return jsonify({
                        "ok": True,
                        "message": "falha ao gerar pdf da visita mensal",
                        "error": str(e),
                    }), 200

                send_result = send_telegram_document(
                    chat_id=chat_message.chat_id,
                    file_bytes=pdf_bytes,
                    filename=filename,
                    caption=f"📄 PDF da visita {visit.id}"
                )

                return jsonify({
                    "ok": True,
                    "message": "pdf enviado para visita da lista mensal",
                    "visit_id": visit.id,
                    "send_result": send_result,
                }), 200

            summary_text = build_visit_summary_text(
                action="use_existing_pending_visit",
                final_visit_payload={
                    "linked_pending_visit_id": visit.id,
                    "client_id": visit.client_id,
                    "property_id": visit.property_id,
                    "plot_id": visit.plot_id,
                    "consultant_id": visit.consultant_id,
                    "date": visit.date.isoformat() if visit.date else None,
                    "status": visit.status,
                    "culture": visit.culture or "",
                    "variety": visit.variety or "",
                    "fenologia_real": visit.fenologia_real,
                    "recommendation": visit.recommendation or "",
                    "products": [p.to_dict() for p in (visit.products or [])],
                    "latitude": visit.latitude,
                    "longitude": visit.longitude,
                    "source": "chatbot",
                },
                selected_pending_visit=selected,
                close_only=False
            )

            state.visit_preview_json = json.dumps({
                "action": "use_existing_pending_visit",
                "final_visit_payload": {
                    "linked_pending_visit_id": visit.id,
                    "client_id": visit.client_id,
                    "property_id": visit.property_id,
                    "plot_id": visit.plot_id,
                    "consultant_id": visit.consultant_id,
                    "date": visit.date.isoformat() if visit.date else None,
                    "status": visit.status,
                    "culture": visit.culture or "",
                    "variety": visit.variety or "",
                    "fenologia_real": visit.fenologia_real,
                    "recommendation": visit.recommendation or "",
                    "products": [p.to_dict() for p in (visit.products or [])],
                    "latitude": visit.latitude,
                    "longitude": visit.longitude,
                    "source": "chatbot",
                },
                "selected_pending_visit": selected,
                "close_only": False,
            }, ensure_ascii=False)
            state.confirmation_text = summary_text
            state.status = "awaiting_final_confirmation"
            db.session.commit()

            send_telegram_message(
                chat_id=chat_message.chat_id,
                text=summary_text
            )

            return jsonify({
                "ok": True,
                "message": "visita da lista mensal carregada para edição",
                "visit": visit.to_dict()
            }), 200

    return None


BOT_TONE_MESSAGES = {
    "week_not_found": [
        "Ainda não tenho uma agenda aberta aqui. Me pede primeiro sua agenda da semana.",
        "Não achei uma agenda ativa agora. Primeiro me chama com agenda da semana.",
        "Beleza, mas antes preciso abrir sua agenda da semana."
    ],
    "consultant_not_bound": [
        "Seu Telegram ainda não está vinculado a um consultor. Usa /start e depois /vincular SEU_CODIGO.",
        "Ainda não consegui te ligar a um consultor do sistema. Faz /start e depois /vincular SEU_CODIGO.",
    ],
    "audio_fail": [
        "Recebi seu áudio, mas não consegui transcrever. Tenta de novo ou me manda em texto.",
        "Peguei seu áudio aqui, mas a transcrição falhou. Pode reenviar ou mandar escrito.",
    ],
    "photo_saved_waiting_context": [
        "Fechado. Já deixei sua foto separada. Agora me manda o texto ou áudio da visita.",
        "Perfeito, foto recebida. Agora só preciso do texto ou áudio da visita.",
    ],
    "invalid_option": [
        "Essa opção não bateu aqui. Confere e me manda de novo.",
        "Não consegui usar essa opção. Dá uma revisada e tenta outra vez.",
    ],
    "confirm_or_cancel": [
        "Agora só me responde com CONFIRMAR ou CANCELAR.",
        "Fechou. Nessa etapa me responde só com CONFIRMAR ou CANCELAR.",
    ],
    "operation_cancelled": [
        "Fechado, cancelei por aqui.",
        "Certo, operação cancelada.",
        "Beleza, parei esse fluxo aqui.",
    ],
    "help_intro": [
        "🤖 Posso te ajudar com:",
    ],
    "summary_intro": [
        "Fechou, entendi assim:",
        "Beleza, ficou assim:",
        "Montei esse resumo aqui:",
    ],
}

def bot_phrase(key: str, default: str = "") -> str:
    options = BOT_TONE_MESSAGES.get(key) or []
    if not options:
        return default
    return random.choice(options)



def is_today_schedule_request(text: str) -> bool:
    normalized = normalize_lookup_text(text or "")

    triggers = [
        "agenda de hoje",
        "visitas de hoje",
        "o que tenho hoje",
        "o que eu tenho hoje",
        "hoje pra mim",
    ]
    return any(t in normalized for t in triggers)


def is_daily_routine_request(text: str) -> bool:
    normalized = normalize_lookup_text(text or "")

    triggers = [
        "rotina do dia",
        "meu dia",
        "prioridades de hoje",
        "resumo do dia",
        "o que falta hoje",
    ]
    return any(t in normalized for t in triggers)


def find_consultant_visits_for_day(consultant_id: int, reference_date=None, limit: int = 50):
    if not consultant_id:
        return []

    day_ref = reference_date or get_local_today()

    visits = (
        Visit.query
        .filter(Visit.consultant_id == consultant_id)
        .filter(Visit.date == day_ref)
        .order_by(Visit.date.asc(), Visit.id.asc())
        .limit(limit)
        .all()
    )
    return visits


def build_today_schedule_text(consultant_name: str, visits: list, reference_date=None) -> str:
    today = reference_date or get_local_today()

    if not visits:
        return (
            f"📅 Agenda de hoje\n"
            f"Consultor: {consultant_name}\n"
            f"Data: {today.strftime('%d/%m/%Y')}\n\n"
            f"Hoje você não tem visitas cadastradas."
        )

    lines = [
        "📅 Agenda de hoje",
        f"Consultor: {consultant_name}",
        f"Data: {today.strftime('%d/%m/%Y')}",
        "",
    ]

    for idx, visit in enumerate(visits, start=1):
        client_name = visit.client.name if getattr(visit, "client", None) else f"Cliente {visit.client_id}"
        culture = visit.culture or "—"
        stage = visit.fenologia_real or visit.recommendation or "—"
        status = visit.status or "planned"
        lines.append(f"{idx}. {client_name} - {culture} - {stage} - {status}")

    return "\n".join(lines)


def build_daily_routine_text(consultant_name: str, today_visits: list, overdue_month_visits: list, stale_clients: list) -> str:
    lines = [
        "☀️ Rotina do dia",
        f"Consultor: {consultant_name}",
        "",
        f"📅 Visitas de hoje: {len(today_visits)}",
        f"⏰ Visitas atrasadas no mês: {len(overdue_month_visits)}",
        "",
    ]

    if today_visits:
        lines.append("Hoje:")
        for v in today_visits[:5]:
            client_name = v.client.name if getattr(v, "client", None) else f"Cliente {v.client_id}"
            culture = v.culture or "—"
            lines.append(f"- {client_name} - {culture}")
        lines.append("")

    if stale_clients:
        lines.append("Mais atrasados:")
        for item in stale_clients[:5]:
            client_name = item.get("client_name") or "—"
            days_without = item.get("days_without_valid_visit")
            culture = item.get("last_valid_culture") or "—"
            lines.append(f"- {client_name} - {days_without} dias - {culture}")

    return "\n".join(lines)


def handle_daily_routine_flow(chat_message, consultant, message_text: str):
    if not consultant:
        return None

    if is_today_schedule_request(message_text):
        today_visits = find_consultant_visits_for_day(consultant.id)
        response_text = build_today_schedule_text(consultant.name, today_visits)

        send_result = send_telegram_message(
            chat_id=chat_message.chat_id,
            text=response_text
        )

        return jsonify({
            "ok": True,
            "message": "agenda de hoje enviada",
            "response_text": response_text,
            "send_result": send_result,
        }), 200

    if is_daily_routine_request(message_text):
        today_visits = find_consultant_visits_for_day(consultant.id)
        overdue_month_visits = find_consultant_visits_for_month(
            consultant_id=consultant.id,
            filter_mode="overdue"
        )
        stale_clients = find_stale_clients_ranking(
            consultant_id=consultant.id,
            limit=5
        )

        response_text = build_daily_routine_text(
            consultant_name=consultant.name,
            today_visits=today_visits,
            overdue_month_visits=overdue_month_visits,
            stale_clients=stale_clients,
        )

        send_result = send_telegram_message(
            chat_id=chat_message.chat_id,
            text=response_text
        )

        return jsonify({
            "ok": True,
            "message": "rotina do dia enviada",
            "response_text": response_text,
            "send_result": send_result,
        }), 200

    return None

def parse_pdf_client_reference(text: str) -> str | None:
    raw = (text or "").strip()
    normalized = normalize_lookup_text(raw)

    patterns = [
        r"^pdf do cliente (.+)$",
        r"^pdf do (.+)$",
        r"^manda o pdf do (.+)$",
        r"^pdf da ultima do (.+)$",
        r"^pdf da última do (.+)$",
    ]

    for pattern in patterns:
        match = re.match(pattern, normalized)
        if match:
            value = raw[match.start(1):].strip(" .,-")
            if value:
                return value

    return None


def find_last_completed_visit_for_client_reference(consultant_id: int, client_name: str):
    matched_client, _, _ = find_client_by_name(client_name)
    if not matched_client:
        return None

    visit = (
        Visit.query
        .filter(Visit.consultant_id == consultant_id)
        .filter(Visit.client_id == matched_client.id)
        .filter(Visit.status == "done")
        .order_by(Visit.date.desc().nullslast(), Visit.id.desc())
        .first()
    )
    return visit


def parse_human_ordinal_reference(text: str):
    normalized = normalize_lookup_text(text or "")

    mapping = {
        "primeira": 0,
        "segunda": 1,
        "terceira": 2,
        "quarta": 3,
        "quinta": 4,
        "sexta": 5,
        "setima": 6,
        "sétima": 6,
        "oitava": 7,
        "nona": 8,
        "decima": 9,
        "décima": 9,
    }

    for word, idx in mapping.items():
        if normalized == word or normalized == f"a {word}" or normalized == f"visita {word}":
            return idx

    return None



def resolve_single_active_reference(chat_message, message_text: str):
    normalized = normalize_lookup_text(message_text or "").strip()
    if normalized not in {"essa", "essa ai", "essa aí", "essa mesma"}:
        return None

    state = ChatbotConversationState.query.filter_by(
        platform="telegram",
        chat_id=chat_message.chat_id
    ).first()

    if not state:
        return None

    candidates = json.loads(state.pending_visit_suggestions_json or "[]")
    if len(candidates) == 1:
        return "1"

    return None


def is_week_organization_request(text: str) -> bool:
    normalized = normalize_lookup_text(text or "")
    normalized = re.sub(r"\s+", " ", normalized).strip()

    triggers = [
        "organiza minha semana",
        "organizar minha semana",
        "organize minha semana",
        "organize a minha semana",
        "organiza a minha semana",
        "monta minha semana",
        "monte minha semana",
        "monta a minha semana",
        "planeja minha semana",
        "planejar minha semana",
        "planeje minha semana",
        "como organizar minha semana",
        "me ajuda a organizar minha semana",
        "organizar semana",
    ]

    return any(t in normalized for t in triggers)

def resolve_visit_region_label(visit) -> str:
    if not visit:
        return "Sem região definida"

    prop = getattr(visit, "property", None)

    if prop:
        city_state = (getattr(prop, "city_state", None) or "").strip()
        if city_state:
            return city_state

        prop_name = (getattr(prop, "name", None) or "").strip()
        if prop_name:
            return prop_name

    return "Sem região definida"


def resolve_visit_client_name(visit) -> str:
    if not visit:
        return "Cliente indefinido"

    client = getattr(visit, "client", None)
    if client and getattr(client, "name", None):
        return client.name

    return f"Cliente {getattr(visit, 'client_id', '—')}"



def choose_best_visit_per_client(visits: list):
    # Mantém apenas 1 visita por cliente.
    # Regra:
    # - se houver visitas atrasadas do cliente, pega a mais antiga
    # - senão, pega a próxima visita mais próxima da semana
    grouped = {}

    for visit in visits:
        client_id = getattr(visit, "client_id", None)
        if not client_id:
            continue
        grouped.setdefault(client_id, []).append(visit)

    chosen = []

    for client_id, client_visits in grouped.items():
        client_visits.sort(
            key=lambda v: (
                v.date or _date.max,
                v.id or 0,
            )
        )
        chosen.append(client_visits[0])

    return chosen



def build_week_priority_items(consultant_id: int, reference_date=None):
    today = reference_date or get_local_today()
    start_date, end_date = get_week_date_range(today)

    base_visits = (
        Visit.query
        .filter(Visit.consultant_id == consultant_id)
        .filter(Visit.status != "done")
        .filter(
            db.or_(
                Visit.date < today,
                db.and_(Visit.date >= start_date, Visit.date <= end_date)
            )
        )
        .order_by(Visit.date.asc().nullslast(), Visit.id.asc())
        .all()
    )

    # mantém só 1 visita por cliente
    selected_visits = choose_best_visit_per_client(base_visits)

    stale_clients = find_stale_clients_ranking(
        consultant_id=consultant_id,
        limit=500
    )
    stale_map = {
        item["client_id"]: item["days_without_valid_visit"]
        for item in stale_clients
        if item.get("client_id") is not None
    }

    combined = []

    for visit in selected_visits:
        visit_date = visit.date
        is_overdue = bool(visit_date and visit_date < today)

        stale_days = stale_map.get(visit.client_id)
        if stale_days is None:
            # cliente sem visita válida com foto fica no fim
            stale_days = -1

        combined.append({
            "visit": visit,
            "visit_id": visit.id,
            "client_id": visit.client_id,
            "client_name": resolve_visit_client_name(visit),
            "region_label": resolve_visit_region_label(visit),
            "culture": visit.culture or "—",
            "fenologia_real": visit.fenologia_real or "—",
            "recommendation": (visit.recommendation or "").strip() or "—",
            "date": visit_date,
            "is_overdue": is_overdue,
            "stale_days": stale_days,
        })

    combined.sort(
        key=lambda item: (
            0 if item["is_overdue"] else 1,
            -(item["stale_days"] if item["stale_days"] >= 0 else -99999),
            item["region_label"] or "",
            item["date"] or _date.max,
            item["client_name"] or "",
        )
    )

    return combined

def group_week_items_by_region(items: list):
    grouped = {}

    for item in items:
        region = item.get("region_label") or "Sem região definida"
        grouped.setdefault(region, []).append(item)

    return grouped


def build_week_organization_text(consultant_name: str, items: list, reference_date=None) -> str:
    if not items:
        return (
            f"🗓️ Organização da semana\n"
            f"Consultor: {consultant_name}\n\n"
            f"Não encontrei visitas pendentes/atrasadas para organizar."
        )

    today = reference_date or get_local_today()
    agenda = distribute_items_across_week(items, today)

    overdue_count = sum(1 for item in items if item.get("is_overdue"))

    lines = [
        "🗓️ Organização sugerida da semana",
        f"Consultor: {consultant_name}",
        "Critério: 1 prioridade por cliente, considerando atraso da visita e dias desde a última visita válida com foto.",
        "",
        f"📌 Total priorizado: {len(items)}",
        f"⏰ Clientes com pendência atrasada: {overdue_count}",
        "",
    ]

    for day, day_items in agenda.items():
        lines.append(f"📅 {format_weekday_br(day)} - {day.strftime('%d/%m/%Y')}")

        if not day_items:
            lines.append("- sem prioridade sugerida")
            lines.append("")
            continue

        for item in day_items:
            client_name = item["client_name"]
            culture = item["culture"]
            region = item["region_label"] or "Sem região definida"
            stale_days = item.get("stale_days", -1)

            if stale_days >= 0:
                stale_label = f"{stale_days} dias desde a última visita válida"
            else:
                stale_label = "sem visita válida com foto"

            lines.append(
                f"- {client_name} | {culture} | {stale_label} | região: {region}"
            )

        lines.append("")

    lines.append("Sugestão prática:")
    lines.append("- começar pelos clientes mais atrasados")
    lines.append("- aproveitar a mesma região no mesmo dia")
    lines.append("- ajustar manualmente os casos sem região definida")

    return "\n".join(lines)

def handle_week_organization_flow(chat_message, consultant, message_text: str):
    if not is_week_organization_request(message_text):
        return None

    if not consultant:
        send_result = send_telegram_message(
            chat_id=chat_message.chat_id,
            text=bot_phrase(
                "consultant_not_bound",
                "Seu Telegram ainda não está vinculado a um consultor do AgroCRM."
            )
        )
        return jsonify({
            "ok": False,
            "message": "consultor não vinculado",
            "send_result": send_result,
        }), 400

    items = build_week_priority_items(consultant_id=consultant.id)

    response_text = build_week_organization_text(
        consultant_name=consultant.name,
        items=items,
        reference_date=get_local_today(),
    )

    send_result = send_telegram_message(
        chat_id=chat_message.chat_id,
        text=response_text
    )

    return jsonify({
        "ok": True,
        "message": "organização da semana enviada",
        "items_count": len(items),
        "response_text": response_text,
        "send_result": send_result,
    }), 200

def get_business_week_days(reference_date=None):
    today = reference_date or get_local_today()
    start_date, _ = get_week_date_range(today)

    # segunda a sexta
    return [start_date + _timedelta(days=i) for i in range(5)]


def distribute_items_across_week(items: list, reference_date=None):
    # Distribui itens ao longo dos 5 dias úteis da semana.
    # Estratégia simples:
    # - percorre os itens já priorizados
    # - vai distribuindo em round-robin entre segunda e sexta
    week_days = get_business_week_days(reference_date)
    agenda = {day: [] for day in week_days}

    if not items:
        return agenda

    idx = 0
    total_days = len(week_days)

    for item in items:
        target_day = week_days[idx % total_days]
        agenda[target_day].append(item)
        idx += 1

    return agenda


def format_weekday_br(d: date) -> str:
    names = {
        0: "Segunda",
        1: "Terça",
        2: "Quarta",
        3: "Quinta",
        4: "Sexta",
        5: "Sábado",
        6: "Domingo",
    }
    return names.get(d.weekday(), d.strftime("%A"))













@bp.route('/telegram/webhook', methods=['POST'])
def telegram_webhook():
    """
    Webhook inicial do Telegram.
    Nesta etapa:
    - recebe o payload bruto
    - normaliza a mensagem
    - interpreta texto/caption
    - busca cliente e pendências
    - envia resposta automática no Telegram
    """
    try:
        payload = request.get_json(silent=True) or {}

        chatbot_service = ChatbotService()
        chat_message = chatbot_service.normalize_telegram_update(payload)

        if not chat_message:
            return jsonify({
                "ok": True,
                "message": "update sem mensagem utilizável"
            }), 200

        consultant = resolve_telegram_consultant(chat_message)
        resolved_consultant_id = consultant.id if consultant else 1

        message_text = (chat_message.text or chat_message.caption or "").strip()

        message_text = resolve_audio_message_text(
            chat_message=chat_message,
            payload=payload,
            current_text=message_text,
        )

        if message_text is None:
            return jsonify({
                "ok": False,
                "message": "falha no processamento do áudio"
            }), 200

        message_text = (message_text or "").strip()
        message_text_lower = message_text.lower()

        single_reference = resolve_single_active_reference(
            chat_message=chat_message,
            message_text=message_text,
        )
        if single_reference:
            message_text = single_reference
            message_text_lower = message_text.lower()

        photo_info = resolve_pending_photo_for_message(
            chat_message=chat_message,
            payload=payload,
            current_text=message_text,
        )

        if photo_info == "__PHOTO_ONLY_WAITING_CONTEXT__":
            return jsonify({
                "ok": True,
                "message": "foto recebida e salva temporariamente"
            }), 200

        # =========================================================
        # Auto-vínculo Telegram
        # =========================================================
        current_binding = find_telegram_binding(chat_message)

        if message_text_lower == "/start":
            if current_binding:
                send_result = send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text=f"Olá, {current_binding.consultant.name}! Seu Telegram já está vinculado ao AgroCRM ✅"
                )
                return jsonify({
                    "ok": True,
                    "message": "telegram já vinculado",
                    "binding": current_binding.to_dict(),
                    "send_result": send_result,
                }), 200

            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text=(
                    "Olá! Seu Telegram ainda não está vinculado ao AgroCRM.\n\n"
                    "Envie o comando:\n"
                    "/vincular SEU_CODIGO\n\n"
                    "Exemplo:\n"
                    "/vincular JHONAT-AB12"
                )
            )
            return jsonify({
                "ok": True,
                "message": "aguardando vínculo",
                "send_result": send_result,
            }), 200

        if message_text_lower.startswith("/vincular "):
            code = message_text[10:].strip()

            binding, error = bind_telegram_consultant_by_code(chat_message, code)

            if error:
                send_result = send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text="Código inválido. Verifique o código recebido e tente novamente."
                )
                return jsonify({
                    "ok": False,
                    "message": error,
                    "send_result": send_result,
                }), 400

            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text=f"Telegram vinculado com sucesso ✅\nConsultor: {binding.consultant.name}"
            )
            return jsonify({
                "ok": True,
                "message": "vínculo realizado com sucesso",
                "binding": binding.to_dict(),
                "send_result": send_result,
            }), 200

        if not message_text:
            send_telegram_message(
                chat_id=chat_message.chat_id,
                text="Recebi sua mensagem, mas ainda não consegui interpretar esse formato."
            )
            return jsonify({
                "ok": True,
                "message": "mensagem sem texto/caption"
            }), 200

        # =========================================================
        # Cancelar a qualquer momento
        # =========================================================
        if normalize_lookup_text(message_text) in ("cancelar", "cancela", "cancel"):
            state = ChatbotConversationState.query.filter_by(
                platform="telegram",
                chat_id=chat_message.chat_id
            ).first()

            if state:
                state.status = "cancelled"
                db.session.commit()

            # ✅ limpa fotos pendentes da conversa cancelada
            clear_pending_telegram_photos(chat_message.chat_id)

            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text=bot_phrase("operation_cancelled", "Operação cancelada com sucesso.")
            )

            return jsonify({
                "ok": True,
                "message": "operação cancelada globalmente",
                "send_result": send_result,
            }), 200


        priority_state_response = handle_priority_stateful_actions(
            chat_message=chat_message,
            consultant=consultant,
            message_text=message_text,
        )
        if priority_state_response:
            return priority_state_response


        week_schedule_response = handle_week_schedule_flow(
            chat_message=chat_message,
            consultant=consultant,
            resolved_consultant_id=resolved_consultant_id,
            message_text=message_text,
        )
        if week_schedule_response:
            return week_schedule_response

        stale_clients_response = handle_stale_clients_ranking_flow(
            chat_message=chat_message,
            consultant=consultant,
            message_text=message_text,
        )
        if stale_clients_response:
            return stale_clients_response

        daily_routine_response = handle_daily_routine_flow(
            chat_message=chat_message,
            consultant=consultant,
            message_text=message_text,
        )
        if daily_routine_response:
            return daily_routine_response

        week_organization_response = handle_week_organization_flow(
            chat_message=chat_message,
            consultant=consultant,
            message_text=message_text,
        )
        if week_organization_response:
            return week_organization_response


        month_visits_response = handle_month_visits_flow(
            chat_message=chat_message,
            consultant=consultant,
            message_text=message_text,
        )
        if month_visits_response:
            return month_visits_response

        pdf_flow_response = handle_pdf_flow(
            chat_message=chat_message,
            consultant=consultant,
            message_text=message_text,
        )
        if pdf_flow_response:
            return pdf_flow_response

        final_confirmation_response = handle_final_confirmation(
            chat_message=chat_message,
            message_text=message_text,
            photo_info=photo_info,
        )
        if final_confirmation_response:
            return final_confirmation_response

        # =====================================================================================
        # Fluxo guiado: cultura -> plantio/avulsa -> fenologia/data -> observações -> resumo
        # =====================================================================================
        state = ChatbotConversationState.query.filter_by(
            platform="telegram",
            chat_id=chat_message.chat_id
        ).first()

        if state and state.status in (
            "awaiting_culture",
            "awaiting_planting_confirmation",
            "awaiting_avulsa_confirmation",
            "awaiting_fenologia",
            "awaiting_date",
            "awaiting_observations",
        ):
            stored_data = json.loads(state.visit_preview_json or "{}")
            action = stored_data.get("action")
            final_visit_payload = stored_data.get("final_visit_payload") or {}
            selected_pending_visit = stored_data.get("selected_pending_visit")
            close_only = stored_data.get("close_only", False)

            if state.status == "awaiting_culture":
                culture_input = normalize_culture_input(message_text)

                if not culture_input:
                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text="🌱 Cultura inválida.\nEnvie algo como: Milho, Soja ou Algodão."
                    )
                    return jsonify({
                        "ok": True,
                        "message": "cultura inválida"
                    }), 200

                final_visit_payload["culture"] = culture_input

                state.visit_preview_json = json.dumps(
                    build_guided_state_payload(
                        action=action,
                        final_visit_payload=final_visit_payload,
                        selected_pending_visit=selected_pending_visit,
                        close_only=close_only,
                    ),
                    ensure_ascii=False
                )
                state.status = "awaiting_planting_confirmation"
                db.session.commit()

                send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text="🌱 É visita de plantio?\nResponda com SIM ou NÃO."
                )

                return jsonify({
                    "ok": True,
                    "message": "cultura recebida, aguardando confirmação de plantio"
                }), 200

            if state.status == "awaiting_planting_confirmation":
                yes_no = parse_yes_no(message_text)

                if yes_no is None:
                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text="🌱 Resposta inválida.\nResponda apenas com SIM ou NÃO."
                    )
                    return jsonify({
                        "ok": True,
                        "message": "resposta inválida para plantio"
                    }), 200

                if yes_no is True:
                    final_visit_payload["fenologia_real"] = "Plantio"

                    state.visit_preview_json = json.dumps(
                        build_guided_state_payload(
                            action=action,
                            final_visit_payload=final_visit_payload,
                            selected_pending_visit=selected_pending_visit,
                            close_only=close_only,
                        ),
                        ensure_ascii=False
                    )
                    state.status = "awaiting_date"
                    db.session.commit()

                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text="📅 Informe a data da visita.\nExemplos: hoje, amanhã, 15, 24/02/2026"
                    )

                    return jsonify({
                        "ok": True,
                        "message": "visita de plantio confirmada"
                    }), 200

                state.visit_preview_json = json.dumps(
                    build_guided_state_payload(
                        action=action,
                        final_visit_payload=final_visit_payload,
                        selected_pending_visit=selected_pending_visit,
                        close_only=close_only,
                    ),
                    ensure_ascii=False
                )
                state.status = "awaiting_avulsa_confirmation"
                db.session.commit()

                send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text="📌 É uma visita avulsa?\nResponda com SIM ou NÃO."
                )

                return jsonify({
                    "ok": True,
                    "message": "aguardando confirmação de visita avulsa"
                }), 200

            if state.status == "awaiting_avulsa_confirmation":
                yes_no = parse_yes_no(message_text)

                if yes_no is None:
                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text="📌 Resposta inválida.\nResponda apenas com SIM ou NÃO."
                    )
                    return jsonify({
                        "ok": True,
                        "message": "resposta inválida para visita avulsa"
                    }), 200

                # ✅ Se for visita avulsa, FORÇA nova visita
                if yes_no is True:
                    forced_payload = {
                        **(final_visit_payload or {}),
                        "linked_pending_visit_id": None,
                    }

                    state.visit_preview_json = json.dumps(
                        build_guided_state_payload(
                            action="create_new_visit",
                            final_visit_payload=forced_payload,
                            selected_pending_visit=None,
                            close_only=False,
                        ),
                        ensure_ascii=False
                    )
                    state.status = "awaiting_fenologia"
                    db.session.commit()

                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text="🌿 Perfeito. Como é visita avulsa, vou criar uma NOVA visita.\nAgora me informe a fenologia observada.\nExemplo: V4, V5, VE, VT, R1"
                    )

                    return jsonify({
                        "ok": True,
                        "message": "visita avulsa forçada como nova visita"
                    }), 200

                send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text=(
                        "No momento consigo seguir com dois tipos de nova visita:\n"
                        "- visita de plantio\n"
                        "- visita avulsa\n\n"
                        "Se quiser, me envie novamente os dados da visita e eu sigo por esse caminho."
                    )
                )

                state.status = "cancelled"
                db.session.commit()

                return jsonify({
                    "ok": True,
                    "message": "fluxo encerrado por não ser plantio nem avulsa"
                }), 200

            if state.status == "awaiting_fenologia":
                fenologia_input = message_text.strip().upper()

                if not is_valid_fenologia(fenologia_input):
                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text="🌿 Fenologia inválida.\nEnvie algo como: V4, V5, VE, VT, R1."
                    )
                    return jsonify({
                        "ok": True,
                        "message": "fenologia inválida"
                    }), 200

                final_visit_payload["fenologia_real"] = fenologia_input

                state.visit_preview_json = json.dumps(
                    build_guided_state_payload(
                        action=action,
                        final_visit_payload=final_visit_payload,
                        selected_pending_visit=selected_pending_visit,
                        close_only=close_only,
                    ),
                    ensure_ascii=False
                )
                state.status = "awaiting_date"
                db.session.commit()

                send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text="📅 Informe a data da visita.\nExemplo: 24/02/2026"
                )

                return jsonify({
                    "ok": True,
                    "message": "fenologia recebida"
                }), 200

            if state.status == "awaiting_date":
                parsed_date_obj = parse_human_date(
                    message_text.strip(),
                    base_date=get_local_today()
                )
                if not parsed_date_obj:
                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text="Data inválida. Envie algo como: hoje, amanhã, ontem, 2 dias atrás, 15, 24/02/2026 ou 2026-02-24."
                    )
                    return jsonify({
                        "ok": True,
                        "message": "data inválida"
                    }), 200

                final_visit_payload["date"] = parsed_date_obj.isoformat()

                state.visit_preview_json = json.dumps(
                    build_guided_state_payload(
                        action=action,
                        final_visit_payload=final_visit_payload,
                        selected_pending_visit=selected_pending_visit,
                        close_only=close_only,
                    ),
                    ensure_ascii=False
                )
                state.status = "awaiting_observations"
                db.session.commit()

                send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text="💬 Informe as observações da visita."
                )

                return jsonify({
                    "ok": True,
                    "message": "data recebida"
                }), 200

            if state.status == "awaiting_observations":
                observations_input = message_text.strip()

                if observations_input == "":
                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text="💬 Envie algo nas observações, mesmo que seja apenas ."
                    )
                    return jsonify({
                        "ok": True,
                        "message": "observação vazia"
                    }), 200

                final_visit_payload["recommendation"] = observations_input

                summary_text = build_visit_summary_text(
                    action=action,
                    final_visit_payload=final_visit_payload,
                    selected_pending_visit=selected_pending_visit,
                    close_only=close_only
                )

                state.visit_preview_json = json.dumps(
                    build_guided_state_payload(
                        action=action,
                        final_visit_payload=final_visit_payload,
                        selected_pending_visit=selected_pending_visit,
                        close_only=close_only,
                    ),
                    ensure_ascii=False
                )
                state.confirmation_text = summary_text
                state.status = "awaiting_final_confirmation"
                db.session.commit()

                send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text=summary_text
                )

                return jsonify({
                    "ok": True,
                    "message": "observações recebidas e resumo enviado"
                }), 200

        # =========================================================
        # Se a mensagem for resposta para escolha de cliente parecido
        # =========================================================
        if message_text.strip().isdigit():
            state = ChatbotConversationState.query.filter_by(
                platform="telegram",
                chat_id=chat_message.chat_id,
                status="awaiting_client_confirmation"
            ).first()

            if state:
                client_candidates = json.loads(state.pending_visit_suggestions_json or "[]")
                idx = int(message_text.strip()) - 1

                if idx < 0 or idx >= len(client_candidates):
                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text="Opção inválida. Responda com o número correto do cliente."
                    )
                    return jsonify({
                        "ok": True,
                        "message": "opção inválida para cliente"
                    }), 200

                selected_client = client_candidates[idx]
                selected_client_id = selected_client.get("id")

                matched_client = Client.query.get(selected_client_id)
                if not matched_client:
                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text="Não consegui localizar o cliente escolhido."
                    )
                    return jsonify({
                        "ok": True,
                        "message": "cliente escolhido não encontrado"
                    }), 200

                original_message = state.last_message or ""
                parsed = parse_chatbot_message(original_message)

                parsed_recommendation = extract_recommendation_fallback(original_message)

                if not parsed_recommendation:
                    parsed_recommendation = (parsed.get("recommendation") or "").strip()

                parsed_products = normalize_products_from_parsed(parsed.get("products") or [])

                matched_property, property_candidates, property_needs_confirmation = find_property_by_name(
                    parsed.get("property_name"),
                    matched_client.id if matched_client else None
                )

                pending_visits = []
                same_culture_found = False

                if matched_client:
                    pending_visits, same_culture_found = find_pending_visits(
                        client_id=matched_client.id,
                        property_id=matched_property.id if matched_property else None,
                        culture=parsed.get("culture"),
                        limit=5
                    )

                suggestions = []
                for visit in pending_visits:
                    suggestions.append({
                        "id": visit.id,
                        "date": visit.date.isoformat() if visit.date else None,
                        "status": visit.status,
                        "culture": visit.culture,
                        "variety": visit.variety,
                        "fenologia_real": visit.fenologia_real,
                        "recommendation": (visit.recommendation or "").strip(),
                        "client_id": visit.client_id,
                        "property_id": visit.property_id,
                        "plot_id": visit.plot_id,
                        "property_name": visit.property.name if getattr(visit, "property", None) else "",
                        "plot_name": visit.plot.name if getattr(visit, "plot", None) else "",
                        "display_text": visit.to_dict().get("display_text"),
                    })

                visit_preview = {
                    "client_id": matched_client.id if matched_client else None,
                    "property_id": matched_property.id if matched_property else None,
                    "plot_id": None,
                    "consultant_id": resolved_consultant_id,
                    "date": parsed.get("date"),
                    "status": parsed.get("status", "planned"),
                    "culture": parsed.get("culture") or "",
                    "variety": "",
                    "fenologia_real": parsed.get("fenologia_real"),
                    "recommendation": parsed_recommendation,
                    "products": parsed_products,
                    "latitude": None,
                    "longitude": None,
                    "generate_schedule": False,
                    "source": parsed.get("source", "chatbot"),
                }

                # ✅ NOVO COMPORTAMENTO:
                # se não há pendências, pula o "NOVA" e já vai para confirmação final
                if not suggestions:
                    return start_new_visit_direct_confirmation(
                        state=state,
                        chat_message=chat_message,
                        visit_preview=visit_preview,
                        matched_client=matched_client,
                        matched_property=matched_property,
                    )

                confirmation_text = build_pending_visits_confirmation_text(
                    client_name=matched_client.name,
                    requested_culture=parsed.get("culture"),
                    suggestions=suggestions,
                    same_culture_found=same_culture_found
                )

                state.pending_visit_suggestions_json = json.dumps(suggestions, ensure_ascii=False)
                state.visit_preview_json = json.dumps(visit_preview, ensure_ascii=False)
                state.confirmation_text = confirmation_text
                state.status = "awaiting_confirmation"

                db.session.commit()

                send_result = send_telegram_message(
                    chat_id=chat_message.chat_id,
                    text=confirmation_text
                )

                return jsonify({
                    "ok": True,
                    "message": "cliente confirmado e fluxo continuado",
                    "matched_client": matched_client.to_dict() if matched_client else None,
                    "matched_property": matched_property.to_dict() if matched_property else None,
                    "pending_visit_suggestions": suggestions,
                    "same_culture_found": same_culture_found,
                    "confirmation_text": confirmation_text,
                    "send_result": send_result,
                }), 200

        # =========================================================
        # Se a mensagem for resposta para visitas pendentes / nova visita
        # =========================================================
        parsed_reply = parse_pending_reply(message_text)

        if parsed_reply and parsed_reply["mode"] in ("update_existing", "close_only", "create_new"):
            state = ChatbotConversationState.query.filter_by(
                platform="telegram",
                chat_id=chat_message.chat_id,
                status="awaiting_confirmation"
            ).first()

            if state:
                pending_visit_suggestions = json.loads(state.pending_visit_suggestions_json or "[]")
                visit_preview = json.loads(state.visit_preview_json or "{}")

                mode = parsed_reply["mode"]
                selected_pending_visit = None
                final_visit_payload = {}
                action = None
                close_only = False

                if mode == "create_new":
                    action = "create_new_visit"
                    final_visit_payload = {
                        **visit_preview,
                        "culture": "",
                        "fenologia_real": None,
                        "date": None,
                        "recommendation": "",
                    }

                    state.visit_preview_json = json.dumps(
                        build_guided_state_payload(
                            action=action,
                            final_visit_payload=final_visit_payload,
                            selected_pending_visit=None,
                            close_only=False,
                        ),
                        ensure_ascii=False
                    )
                    state.status = "awaiting_culture"
                    db.session.commit()

                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text="🌱 Informe a cultura da visita.\nExemplo: Milho, Soja, Algodão"
                    )

                    return jsonify({
                        "ok": True,
                        "message": "aguardando cultura para nova visita",
                        "action": action,
                    }), 200

                else:
                    idx = parsed_reply.get("index", -1)

                    if idx < 0 or idx >= len(pending_visit_suggestions):
                        send_telegram_message(
                            chat_id=chat_message.chat_id,
                            text="Número inválido da lista. Revise as opções e tente novamente."
                        )
                        return jsonify({
                            "ok": True,
                            "message": "índice inválido"
                        }), 200

                    selected_pending_visit = pending_visit_suggestions[idx]

                    action = "use_existing_pending_visit"
                    close_only = (mode == "close_only")

                    final_visit_payload = {
                        **visit_preview,
                        "client_id": selected_pending_visit.get("client_id") or visit_preview.get("client_id"),
                        "property_id": selected_pending_visit.get("property_id") or visit_preview.get("property_id"),
                        "plot_id": selected_pending_visit.get("plot_id") or visit_preview.get("plot_id"),
                        "consultant_id": resolved_consultant_id or visit_preview.get("consultant_id"),
                        "linked_pending_visit_id": selected_pending_visit.get("id"),
                        "status": "done",
                        "culture": visit_preview.get("culture") or selected_pending_visit.get("culture") or "",
                        "variety": selected_pending_visit.get("variety") or visit_preview.get("variety") or "",
                        "fenologia_real": visit_preview.get("fenologia_real"),
                        "date": visit_preview.get("date"),
                        "recommendation": visit_preview.get("recommendation") or "",
                        "products": visit_preview.get("products") or [],
                        "latitude": visit_preview.get("latitude"),
                        "longitude": visit_preview.get("longitude"),
                        "source": "chatbot",
                    }

                    if close_only:
                        summary_text = build_visit_summary_text(
                            action=action,
                            final_visit_payload=final_visit_payload,
                            selected_pending_visit=selected_pending_visit,
                            close_only=True
                        )

                        state.visit_preview_json = json.dumps(
                            build_guided_state_payload(
                                action=action,
                                final_visit_payload=final_visit_payload,
                                selected_pending_visit=selected_pending_visit,
                                close_only=True,
                            ),
                            ensure_ascii=False
                        )
                        state.confirmation_text = summary_text
                        state.status = "awaiting_final_confirmation"
                        db.session.commit()

                        send_telegram_message(
                            chat_id=chat_message.chat_id,
                            text=summary_text
                        )

                        return jsonify({
                            "ok": True,
                            "message": "resumo enviado para conclusão simples",
                            "summary_text": summary_text
                        }), 200

                    has_prefilled_fenologia = bool((final_visit_payload.get("fenologia_real") or "").strip())
                    has_prefilled_date = bool(final_visit_payload.get("date"))
                    has_prefilled_observation = bool((final_visit_payload.get("recommendation") or "").strip())

                    if has_prefilled_fenologia and has_prefilled_date and has_prefilled_observation:
                        summary_text = build_visit_summary_text(
                            action=action,
                            final_visit_payload=final_visit_payload,
                            selected_pending_visit=selected_pending_visit,
                            close_only=False
                        )

                        state.visit_preview_json = json.dumps(
                            build_guided_state_payload(
                                action=action,
                                final_visit_payload=final_visit_payload,
                                selected_pending_visit=selected_pending_visit,
                                close_only=False,
                            ),
                            ensure_ascii=False
                        )
                        state.confirmation_text = summary_text
                        state.status = "awaiting_final_confirmation"
                        db.session.commit()

                        send_telegram_message(
                            chat_id=chat_message.chat_id,
                            text=summary_text
                        )

                        return jsonify({
                            "ok": True,
                            "message": "resumo final enviado com pré-preenchimento",
                            "summary_text": summary_text
                        }), 200

                    if not has_prefilled_fenologia:
                        next_status = "awaiting_fenologia"
                        next_message = "🌿 Informe a fenologia observada.\nExemplo: V4, V5, R1"
                    elif not has_prefilled_date:
                        next_status = "awaiting_date"
                        next_message = "📅 Informe a data da visita.\nExemplo: hoje, ontem ou 24/02/2026"
                    else:
                        next_status = "awaiting_observations"
                        next_message = "💬 Informe as observações da visita."

                    state.visit_preview_json = json.dumps(
                        build_guided_state_payload(
                            action=action,
                            final_visit_payload=final_visit_payload,
                            selected_pending_visit=selected_pending_visit,
                            close_only=False,
                        ),
                        ensure_ascii=False
                    )
                    state.status = next_status
                    db.session.commit()

                    send_telegram_message(
                        chat_id=chat_message.chat_id,
                        text=next_message
                    )

                    return jsonify({
                        "ok": True,
                        "message": "fluxo guiado continuado com pré-preenchimento",
                        "next_status": next_status
                    }), 200

        # =========================================================
        # IA fallback para interpretação livre
        # =========================================================
        current_state_row = get_current_chatbot_state("telegram", chat_message.chat_id)
        current_state = current_state_row.status if current_state_row else ""

        ai_result = interpret_user_message_with_ai(
            message_text=message_text,
            current_state=current_state
        ) or {}

        if ai_result and ai_result.get("confidence") in ("high", "medium"):
            ai_intent = ai_result.get("intent")

            if ai_intent == "week_schedule_request":
                message_text = "agenda da semana"

            elif ai_intent == "pdf_last_visit":
                message_text = "pdf da última visita"

            elif ai_intent == "pdf_recent_visits":
                message_text = "gerar pdf"

            elif ai_intent == "confirm":
                message_text = "CONFIRMAR"

            elif ai_intent == "cancel":
                message_text = "CANCELAR"

            elif ai_intent == "edit_summary":
                field = ai_result.get("field")
                value = (ai_result.get("value") or "").strip()

                if field and value:
                    field_map = {
                        "fenologia_real": "ALTERAR FENOLOGIA",
                        "date": "ALTERAR DATA",
                        "recommendation": "ALTERAR OBSERVACAO",
                        "culture": "ALTERAR CULTURA",
                        "variety": "ALTERAR VARIEDADE",
                    }

                    prefix = field_map.get(field)
                    if prefix:
                        message_text = f"{prefix} {value}"

            elif ai_intent == "launch_week_visit":
                visit_index = ai_result.get("visit_index")
                if visit_index:
                    message_text = f"LANCAR VISITA {visit_index}"

            elif ai_intent == "complete_week_visit":
                visit_index = ai_result.get("visit_index")
                if visit_index:
                    message_text = f"CONCLUIR VISITA {visit_index}"

            elif ai_intent == "create_visit_like_message":
                parsed_visit = ai_result.get("parsed_visit") or {}

                ai_parts = []
                if parsed_visit.get("client_name"):
                    ai_parts.append(f"cliente {parsed_visit['client_name']}")
                if parsed_visit.get("property_name"):
                    ai_parts.append(f"fazenda {parsed_visit['property_name']}")
                if parsed_visit.get("plot_name"):
                    ai_parts.append(f"talhao {parsed_visit['plot_name']}")
                if parsed_visit.get("culture"):
                    ai_parts.append(str(parsed_visit["culture"]))
                if parsed_visit.get("fenologia_real"):
                    ai_parts.append(str(parsed_visit["fenologia_real"]))
                if parsed_visit.get("date"):
                    ai_parts.append(str(parsed_visit["date"]))
                if parsed_visit.get("recommendation"):
                    ai_parts.append(str(parsed_visit["recommendation"]))

                if ai_parts:
                    message_text = " ".join(ai_parts)
            if ai_intent == "today_schedule_request":
                message_text = "agenda de hoje"

            elif ai_intent == "daily_routine_request":
                message_text = "rotina do dia"

            elif ai_intent == "pdf_by_client_reference":
                parsed_visit = ai_result.get("parsed_visit") or {}
                client_name = (parsed_visit.get("client_name") or "").strip()
                if client_name:
                    message_text = f"pdf do cliente {client_name}"

            elif ai_intent == "contextual_visit_reference":
                visit_index = ai_result.get("visit_index")
                if visit_index:
                    message_text = str(visit_index)

        free_client_guess = try_extract_client_from_free_text(message_text)

        if free_client_guess and not any([
            is_week_schedule_request(message_text),
            is_pdf_request(message_text),
            is_last_pdf_request(message_text),
            parse_pending_reply(message_text),
            parse_summary_edit_command(message_text),
        ]):
            guessed_client = free_client_guess["client"]

            pending_visits, same_culture_found = find_pending_visits(
                client_id=guessed_client.id,
                property_id=None,
                culture=None,
                limit=5
            )

            suggestions = []
            for visit in pending_visits:
                suggestions.append({
                    "id": visit.id,
                    "date": visit.date.isoformat() if visit.date else None,
                    "status": visit.status,
                    "culture": visit.culture,
                    "variety": visit.variety,
                    "fenologia_real": visit.fenologia_real,
                    "recommendation": (visit.recommendation or "").strip(),
                    "client_id": visit.client_id,
                    "property_id": visit.property_id,
                    "plot_id": visit.plot_id,
                    "property_name": visit.property.name if getattr(visit, "property", None) else "",
                    "plot_name": visit.plot.name if getattr(visit, "plot", None) else "",
                    "display_text": visit.to_dict().get("display_text"),
                })

            state = ChatbotConversationState.query.filter_by(
                platform="telegram",
                chat_id=chat_message.chat_id
            ).first()

            if not state:
                state = ChatbotConversationState(
                    platform="telegram",
                    chat_id=chat_message.chat_id,
                )
                db.session.add(state)

            state.last_message = f"cliente {guessed_client.name}"
            state.pending_visit_suggestions_json = json.dumps(suggestions, ensure_ascii=False)

            prefill = extract_prefill_from_message_text(message_text)

            visit_preview = {
                "client_id": guessed_client.id,
                "property_id": None,
                "plot_id": None,
                "consultant_id": resolved_consultant_id,
                "date": prefill.get("date"),
                "status": "planned",
                "culture": prefill.get("culture") or "",
                "variety": "",
                "fenologia_real": prefill.get("fenologia_real"),
                "recommendation": prefill.get("recommendation") or "",
                "products": prefill.get("products") or [],
                "latitude": None,
                "longitude": None,
                "generate_schedule": False,
                "source": "chatbot",
            }

            # ✅ se não há pendência, já vai direto para confirmação final
            if not suggestions:
                return start_new_visit_direct_confirmation(
                    state=state,
                    chat_message=chat_message,
                    visit_preview=visit_preview,
                    matched_client=guessed_client,
                    matched_property=None,
                )

            confirmation_text = build_pending_visits_confirmation_text(
                client_name=guessed_client.name,
                requested_culture=None,
                suggestions=suggestions,
                same_culture_found=same_culture_found
            )

            state.visit_preview_json = json.dumps(visit_preview, ensure_ascii=False)
            state.confirmation_text = confirmation_text
            state.status = "awaiting_confirmation"
            db.session.commit()

            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text=confirmation_text
            )

            return jsonify({
                "ok": True,
                "message": "cliente identificado por texto livre",
                "matched_client": guessed_client.to_dict(),
                "pending_visit_suggestions": suggestions,
                "confirmation_text": confirmation_text,
                "send_result": send_result,
            }), 200


        normalized_help = normalize_lookup_text(message_text)

        if normalized_help in {
            "ajuda",
            "menu",
            "o que voce faz",
            "o que você faz",
            "comandos",
        }:
            help_text = (
                "🤖 Posso te ajudar com:\n\n"
                "1. Agenda semanal\n"
                "- Ex.: agenda da semana\n"
                "- Ex.: lança a 3\n"
                "- Ex.: conclui a 2\n\n"
                "2. Rotina do dia\n"
                "- Ex.: meu dia\n"
                "- Ex.: agenda de hoje\n"
                "- Ex.: prioridades de hoje\n\n"
                "3. Clientes mais atrasados\n"
                "- Ex.: clientes mais atrasados\n"
                "- Ex.: visitas do mês atrasadas\n\n"
                "4. PDFs\n"
                "- Ex.: pdf\n"
                "- Ex.: pdf da ultima visita\n"
                "- Ex.: pdf do Evaristo\n\n"
                "5. Lançar visita em linguagem natural\n"
                "- Ex.: cliente Marcelo Alonso soja v4 hoje aplicar fungicida\n"
                "- Ex.: a terceira\n"
                "- Ex.: essa\n\n"
                "6. Organização da semana\n"
                "- Ex.: organiza minha semana\n"
                "- Ex.: organize a minha semana\n"
                "- Ex.: monta minha semana\n"
                "- Ex.: planeja minha semana\n"
            )

            send_telegram_message(
                chat_id=chat_message.chat_id,
                text=help_text
            )

            return jsonify({
                "ok": True,
                "message": "ajuda enviada"
            }), 200

        parsed = parse_chatbot_message(message_text)

        matched_client, client_candidates, client_needs_confirmation = find_client_by_name(
            parsed.get("client_name")
        )

        matched_property, property_candidates, property_needs_confirmation = find_property_by_name(
            parsed.get("property_name"),
            matched_client.id if matched_client else None
        )

        if client_needs_confirmation:
            confirmation_text = build_name_confirmation_text("cliente", client_candidates)

            state = ChatbotConversationState.query.filter_by(
                platform="telegram",
                chat_id=chat_message.chat_id
            ).first()

            if not state:
                state = ChatbotConversationState(
                    platform="telegram",
                    chat_id=chat_message.chat_id,
                )
                db.session.add(state)

            state.last_message = message_text
            state.pending_visit_suggestions_json = json.dumps(
                [{"id": c.id, "name": c.name} for c in client_candidates[:3]],
                ensure_ascii=False
            )
            state.visit_preview_json = json.dumps({}, ensure_ascii=False)
            state.confirmation_text = confirmation_text
            state.status = "awaiting_client_confirmation"

            db.session.commit()

            send_result = send_telegram_message(
                chat_id=chat_message.chat_id,
                text=confirmation_text
            )

            return jsonify({
                "ok": True,
                "confirmation_text": confirmation_text,
                "client_candidates": [c.to_dict() for c in client_candidates[:3]],
                "send_result": send_result,
            }), 200

        pending_visits = []
        same_culture_found = False

        if matched_client:
            pending_visits, same_culture_found = find_pending_visits(
                client_id=matched_client.id,
                property_id=matched_property.id if matched_property else None,
                culture=parsed.get("culture"),
                limit=5
            )

        suggestions = []
        for visit in pending_visits:
            suggestions.append({
                "id": visit.id,
                "date": visit.date.isoformat() if visit.date else None,
                "status": visit.status,
                "culture": visit.culture,
                "variety": visit.variety,
                "fenologia_real": visit.fenologia_real,
                "recommendation": (visit.recommendation or "").strip(),
                "client_id": visit.client_id,
                "property_id": visit.property_id,
                "plot_id": visit.plot_id,
                "property_name": visit.property.name if getattr(visit, "property", None) else "",
                "plot_name": visit.plot.name if getattr(visit, "plot", None) else "",
                "display_text": visit.to_dict().get("display_text"),
            })

        parsed_recommendation = (parsed.get("recommendation") or "").strip()
        if not parsed_recommendation:
            parsed_recommendation = extract_recommendation_fallback(message_text)

        parsed_products = normalize_products_from_parsed(parsed.get("products") or [])

        if matched_client:
            confirmation_text = build_pending_visits_confirmation_text(
                client_name=matched_client.name,
                requested_culture=parsed.get("culture"),
                suggestions=suggestions,
                same_culture_found=same_culture_found
            )
        else:
            confirmation_text = (
                "Não consegui localizar o cliente informado.\n\n"
                "Tente enviar no formato:\n"
                "cliente NOME_CLIENTE fazenda NOME_FAZENDA cultura estágio observação"
            )

        visit_preview = {
            "client_id": matched_client.id if matched_client else None,
            "property_id": matched_property.id if matched_property else None,
            "plot_id": None,
            "consultant_id": resolved_consultant_id,
            "date": parsed.get("date"),
            "status": parsed.get("status", "planned"),
            "culture": parsed.get("culture") or "",
            "variety": "",
            "fenologia_real": parsed.get("fenologia_real"),
            "recommendation": parsed_recommendation,
            "products": parsed_products,
            "latitude": None,
            "longitude": None,
            "generate_schedule": False,
            "source": parsed.get("source", "chatbot"),
        }

        state = ChatbotConversationState.query.filter_by(
            platform="telegram",
            chat_id=chat_message.chat_id
        ).first()

        if not state:
            state = ChatbotConversationState(
                platform="telegram",
                chat_id=chat_message.chat_id,
            )
            db.session.add(state)

        state.last_message = message_text
        state.pending_visit_suggestions_json = json.dumps(suggestions, ensure_ascii=False)
        state.visit_preview_json = json.dumps(visit_preview, ensure_ascii=False)
        state.confirmation_text = confirmation_text
        state.status = "awaiting_confirmation"

        db.session.commit()

        send_result = send_telegram_message(
            chat_id=chat_message.chat_id,
            text=confirmation_text
        )

        return jsonify({
            "ok": True,
            "telegram_summary": chatbot_service.build_internal_summary(chat_message),
            "parsed_message": parsed,
            "matched_client": matched_client.to_dict() if matched_client else None,
            "matched_property": matched_property.to_dict() if matched_property else None,
            "pending_visit_suggestions": suggestions,
            "same_culture_found": same_culture_found,
            "confirmation_text": confirmation_text,
            "send_result": send_result,
        }), 200

    except Exception as e:
        print(f"❌ Erro em /telegram/webhook: {e}")
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500



def compact_user_text_for_ai(text: str) -> str:
    if not text:
        return ""

    text = text.replace("\r", " ").replace("\n", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text









def get_week_date_range(reference_date=None):
    today = reference_date or get_local_today()
    start = today - _timedelta(days=today.weekday())
    end = start + _timedelta(days=6)
    return start, end






@bp.route('/telegram/bindings', methods=['POST'])
def create_telegram_binding():
    try:
        data = request.get_json(silent=True) or {}

        telegram_chat_id = str(data.get("telegram_chat_id") or "").strip()
        telegram_user_id = str(data.get("telegram_user_id") or "").strip()
        telegram_username = (data.get("telegram_username") or "").strip()
        display_name = (data.get("display_name") or "").strip()
        consultant_id = data.get("consultant_id")

        if not telegram_chat_id:
            return jsonify({
                "ok": False,
                "error": "telegram_chat_id is required"
            }), 400

        if not consultant_id:
            return jsonify({
                "ok": False,
                "error": "consultant_id is required"
            }), 400

        consultant = Consultant.query.get(consultant_id)
        if not consultant:
            return jsonify({
                "ok": False,
                "error": "consultant not found"
            }), 404

        existing = TelegramContactBinding.query.filter_by(
            telegram_chat_id=telegram_chat_id
        ).first()

        if existing:
            existing.telegram_user_id = telegram_user_id or existing.telegram_user_id
            existing.telegram_username = telegram_username or existing.telegram_username
            existing.display_name = display_name or existing.display_name
            existing.consultant_id = consultant_id
            existing.is_active = True
            db.session.commit()

            return jsonify({
                "ok": True,
                "message": "binding updated",
                "binding": existing.to_dict()
            }), 200

        binding = TelegramContactBinding(
            telegram_chat_id=telegram_chat_id,
            telegram_user_id=telegram_user_id or None,
            telegram_username=telegram_username or None,
            display_name=display_name or None,
            consultant_id=consultant_id,
            is_active=True
        )

        db.session.add(binding)
        db.session.commit()

        return jsonify({
            "ok": True,
            "message": "binding created",
            "binding": binding.to_dict()
        }), 201

    except Exception as e:
        print(f"❌ Erro em /telegram/bindings: {e}")
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500









@bp.route('/chatbot/preview-visit', methods=['POST'])
def chatbot_preview_visit():
    """
    Recebe uma mensagem simples e devolve uma prévia
    do payload que no futuro será enviado para /api/visits.
    Ainda não salva nada no banco.
    """
    try:
        data = request.get_json(silent=True) or {}
        message = (data.get("message") or "").strip()
        consultant_id = data.get("consultant_id", 1)

        if not message:
            return jsonify({
                "ok": False,
                "error": "message is required"
            }), 400

        parsed = parse_chatbot_message(message)

        matched_client, client_candidates, client_needs_confirmation = find_client_by_name(
            parsed.get("client_name")
        )

        matched_property, property_candidates, property_needs_confirmation = find_property_by_name(
            parsed.get("property_name"),
            matched_client.id if matched_client else None
        )

        visit_payload = {
            "client_id": matched_client.id if matched_client else None,
            "property_id": matched_property.id if matched_property else None,
            "plot_id": None,
            "consultant_id": consultant_id,
            "date": parsed.get("date"),
            "status": parsed.get("status", "planned"),
            "culture": parsed.get("culture") or "",
            "variety": "",
            "fenologia_real": parsed.get("fenologia_real"),
            "recommendation": parsed.get("recommendation") or "",
            "products": [],
            "latitude": None,
            "longitude": None,
            "generate_schedule": False,
            "source": parsed.get("source", "chatbot"),
        }

        return jsonify({
            "ok": True,
            "parsed_message": parsed,
            "matched_entities": {
                "client": matched_client.to_dict() if matched_client else None,
                "property": matched_property.to_dict() if matched_property else None,
            },
            "visit_preview": visit_payload
        }), 200

    except Exception as e:
        print(f"❌ Erro em /chatbot/preview-visit: {e}")
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500



@bp.route('/chatbot/suggest-pending-visits', methods=['POST'])
def chatbot_suggest_pending_visits():
    """
    Recebe uma mensagem do chatbot, resolve cliente/propriedade
    e sugere visitas pendentes compatíveis para confirmação.
    Ainda não salva nada.
    """
    try:
        data = request.get_json(silent=True) or {}
        message = (data.get("message") or "").strip()
        consultant_id = data.get("consultant_id", 1)

        if not message:
            return jsonify({
                "ok": False,
                "error": "message is required"
            }), 400

        parsed = parse_chatbot_message(message)

        matched_client, client_candidates, client_needs_confirmation = find_client_by_name(
            parsed.get("client_name")
        )

        matched_property, property_candidates, property_needs_confirmation = find_property_by_name(
            parsed.get("property_name"),
            matched_client.id if matched_client else None
        )

        pending_visits = []
        same_culture_found = False

        if matched_client:
            pending_visits, same_culture_found = find_pending_visits(
                client_id=matched_client.id,
                property_id=matched_property.id if matched_property else None,
                culture=parsed.get("culture"),
                limit=5
            )

        parsed_recommendation = (parsed.get("recommendation") or "").strip()
        if not parsed_recommendation:
            parsed_recommendation = extract_recommendation_fallback(message)

        visit_preview = {
            "client_id": matched_client.id if matched_client else None,
            "property_id": matched_property.id if matched_property else None,
            "plot_id": None,
            "consultant_id": consultant_id,
            "date": parsed.get("date"),
            "status": parsed.get("status", "planned"),
            "culture": parsed.get("culture") or "",
            "variety": "",
            "fenologia_real": parsed.get("fenologia_real"),
            "recommendation": parsed_recommendation,
            "products": normalize_products_from_parsed(parsed.get("products") or []),
            "latitude": None,
            "longitude": None,
            "generate_schedule": False,
            "source": parsed.get("source", "chatbot"),
        }

        suggestions = []
        for visit in pending_visits:
            suggestions.append({
                "id": visit.id,
                "date": visit.date.isoformat() if visit.date else None,
                "status": visit.status,
                "culture": visit.culture,
                "variety": visit.variety,
                "fenologia_real": visit.fenologia_real,
                "recommendation": (visit.recommendation or "").strip(),
                "client_id": visit.client_id,
                "property_id": visit.property_id,
                "plot_id": visit.plot_id,
                "property_name": visit.property.name if getattr(visit, "property", None) else "",
                "plot_name": visit.plot.name if getattr(visit, "plot", None) else "",
                "display_text": visit.to_dict().get("display_text"),
            })

        confirmation_text = build_pending_visits_confirmation_text(
            client_name=matched_client.name if matched_client else parsed.get("client_name"),
            requested_culture=parsed.get("culture"),
            suggestions=suggestions,
            same_culture_found=same_culture_found
        )

        return jsonify({
            "ok": True,
            "parsed_message": parsed,
            "matched_entities": {
                "client": matched_client.to_dict() if matched_client else None,
                "property": matched_property.to_dict() if matched_property else None,
            },
            "visit_preview": visit_preview,
            "pending_visit_suggestions": suggestions,
            "needs_confirmation": True if suggestions else False,
            "same_culture_found": same_culture_found,
            "requested_culture": parsed.get("culture"),
            "confirmation_text": confirmation_text,
        }), 200

    except Exception as e:
        print(f"❌ Erro em /chatbot/suggest-pending-visits: {e}")
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

@bp.route('/chatbot/resolve-confirmation', methods=['POST'])
def chatbot_resolve_confirmation():
    """
    Recebe a resposta do usuário após a sugestão de visitas pendentes.
    Ainda não salva nada no banco.
    Apenas decide se:
    - usa uma visita pendente existente
    - ou cria uma nova visita
    """
    try:
        data = request.get_json(silent=True) or {}

        user_reply = (data.get("user_reply") or "").strip().upper()
        pending_visit_suggestions = data.get("pending_visit_suggestions") or []
        visit_preview = data.get("visit_preview") or {}

        if not user_reply:
            return jsonify({
                "ok": False,
                "error": "user_reply is required"
            }), 400

        if user_reply == "NOVA":
            return jsonify({
                "ok": True,
                "action": "create_new_visit",
                "selected_pending_visit": None,
                "final_visit_payload": visit_preview,
                "message": "Usuário optou por criar uma nova visita."
            }), 200

        if user_reply.isdigit():
            idx = int(user_reply) - 1

            if idx < 0 or idx >= len(pending_visit_suggestions):
                return jsonify({
                    "ok": False,
                    "error": "opção inválida"
                }), 400

            selected = pending_visit_suggestions[idx]

            merged_payload = {
                **visit_preview,
                "client_id": selected.get("client_id") or visit_preview.get("client_id"),
                "property_id": selected.get("property_id") or visit_preview.get("property_id"),
                "plot_id": selected.get("plot_id") or visit_preview.get("plot_id"),
                "linked_pending_visit_id": selected.get("id"),
            }

            return jsonify({
                "ok": True,
                "action": "use_existing_pending_visit",
                "selected_pending_visit": selected,
                "final_visit_payload": merged_payload,
                "message": f"Usuário escolheu a visita pendente #{user_reply}."
            }), 200

        return jsonify({
            "ok": False,
            "error": "resposta inválida. Use um número ou NOVA"
        }), 400

    except Exception as e:
        print(f"❌ Erro em /chatbot/resolve-confirmation: {e}")
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500


@bp.route('/chatbot/commit-visit', methods=['POST'])
def chatbot_commit_visit():
    """
    Efetiva a visita no banco.
    - Se action == use_existing_pending_visit: atualiza a visita pendente escolhida e marca como done
    - Se action == create_new_visit: cria nova visita e marca como done
    """
    try:
        data = request.get_json(silent=True) or {}

        action = data.get("action")
        final_visit_payload = data.get("final_visit_payload") or {}
        selected_pending_visit = data.get("selected_pending_visit")

        if not action:
            return jsonify({
                "ok": False,
                "error": "action is required"
            }), 400

        if not final_visit_payload:
            return jsonify({
                "ok": False,
                "error": "final_visit_payload is required"
            }), 400

        # força status done
        final_visit_payload["status"] = "done"

        # =========================================================
        # 1) Atualizar visita pendente existente
        # =========================================================
        if action == "use_existing_pending_visit":
            pending_visit_id = final_visit_payload.get("linked_pending_visit_id")

            if not pending_visit_id:
                return jsonify({
                    "ok": False,
                    "error": "linked_pending_visit_id is required for use_existing_pending_visit"
                }), 400

            visit = Visit.query.get(pending_visit_id)
            if not visit:
                return jsonify({
                    "ok": False,
                    "error": "pending visit not found"
                }), 404

            # Atualizações principais
            if final_visit_payload.get("date"):
                visit.date = _date.fromisoformat(final_visit_payload["date"])

            visit.status = "done"
            visit.culture = final_visit_payload.get("culture") or visit.culture
            visit.variety = final_visit_payload.get("variety") or visit.variety
            visit.fenologia_real = final_visit_payload.get("fenologia_real")
            visit.recommendation = final_visit_payload.get("recommendation") or visit.recommendation
            visit.consultant_id = final_visit_payload.get("consultant_id") or visit.consultant_id
            visit.latitude = final_visit_payload.get("latitude")
            visit.longitude = final_visit_payload.get("longitude")

            # source só se a coluna existir no banco/model
            if hasattr(visit, "source"):
                visit.source = final_visit_payload.get("source", "chatbot")

            db.session.commit()

            return jsonify({
                "ok": True,
                "action": action,
                "message": "Visita pendente atualizada com sucesso.",
                "visit": visit.to_dict()
            }), 200



        # =========================================================
        # 2) Criar nova visita
        # =========================================================
        if action == "create_new_visit":
            if not final_visit_payload.get("client_id"):
                return jsonify({
                    "ok": False,
                    "error": "client_id is required to create a new visit"
                }), 400

            visit_date = None
            if final_visit_payload.get("date"):
                visit_date = _date.fromisoformat(final_visit_payload["date"])

            new_visit = Visit(
                client_id=final_visit_payload.get("client_id"),
                property_id=final_visit_payload.get("property_id"),
                plot_id=final_visit_payload.get("plot_id"),
                consultant_id=final_visit_payload.get("consultant_id"),
                date=visit_date,
                recommendation=final_visit_payload.get("recommendation") or "",
                status="done",
                culture=final_visit_payload.get("culture") or "",
                variety=final_visit_payload.get("variety") or "",
                fenologia_real=final_visit_payload.get("fenologia_real"),
                latitude=final_visit_payload.get("latitude"),
                longitude=final_visit_payload.get("longitude"),
            )

            if hasattr(new_visit, "source"):
                new_visit.source = final_visit_payload.get("source", "chatbot")

            db.session.add(new_visit)
            db.session.commit()

            return jsonify({
                "ok": True,
                "action": action,
                "message": "Nova visita criada com sucesso.",
                "visit": new_visit.to_dict()
            }), 201

        return jsonify({
            "ok": False,
            "error": "invalid action"
        }), 400

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro em /chatbot/commit-visit: {e}")
        return jsonify({
            "ok": False,
            "error": str(e)
        }), 500

# ============================================================
# 🌱 VISITS ENDPOINTS
# ============================================================

@bp.route('/visits', methods=['GET'])
def get_visits():
    """
    Rota unificada:
    - ?month=current → modo iOS (visitas só do mês)
    - ?scope=all → retorna todas as visitas (usado no Acompanhamentos/Calendar)
    - sem params → retorna visitas com filtros normais
    """
    from datetime import date

    try:
        month = request.args.get("month")
        scope = request.args.get("scope")

        # ============================================
        # 🍏 iOS → apenas visitas do mês atual
        # ============================================
        if month == "current":
            today = date.today()
            visits = (
                Visit.query
                .filter(db.extract('month', Visit.date) == today.month)
                .filter(db.extract('year', Visit.date) == today.year)
                .order_by(Visit.date.asc())
                .all()
            )

        # ============================================
        # 🔵 Acompanhamentos / Calendar Desktop → all
        # ============================================
        elif scope == "all":
            visits = Visit.query.order_by(Visit.date.asc().nullslast()).all()

        # ============================================
        # 🔧 Filtros normais (client_id, talhão, etc.)
        # ============================================
        else:
            client_id = request.args.get('client_id', type=int)
            property_id = request.args.get('property_id', type=int)
            plot_id = request.args.get('plot_id', type=int)
            consultant_id = request.args.get('consultant_id', type=int)
            status = request.args.get('status', type=str)

            q = Visit.query
            if client_id:
                q = q.filter_by(client_id=client_id)
            if property_id:
                q = q.filter_by(property_id=property_id)
            if plot_id:
                q = q.filter_by(plot_id=plot_id)
            if consultant_id:
                q = q.filter_by(consultant_id=consultant_id)

            visits = q.order_by(Visit.date.asc().nullslast()).all()

        # ============================================
        # 📸 Montagem final da resposta (unificada)
        # ============================================
        result = []
        backend_url = os.environ.get("RENDER_EXTERNAL_URL") or "https://agrocrm-backend.onrender.com"

        for v in visits:
            client = Client.query.get(v.client_id)
            consultant = Consultant.query.get(v.consultant_id) if v.consultant_id else None
            consultant_name = consultant.name if consultant else None

            photos = []
            for p in v.photos:
                file_name = os.path.basename(p.url)
                photos.append({
                    "id": p.id,
                    "url": resolve_photo_url(p.url),
                    "caption": p.caption or ""
                })

            culture = v.culture or (v.planting.culture if v.planting else "—")
            variety = v.variety or (v.planting.variety if v.planting else "—")

            d = v.to_dict()

            # 🔥 Sobrescreve SEMPRE com o valor real, mesmo se vazio
            d["fenologia_real"] = v.fenologia_real or ""

            # 🔥 ADICIONAR LISTA DE PRODUTOS DA VISITA
            d["products"] = [p.to_dict() for p in v.products]

            # Ajustes finais
            d["client_name"] = client.name if client else f"Cliente {v.client_id}"
            d["consultant_name"] = consultant_name or "—"
            d["culture"] = culture
            d["variety"] = variety
            d["photos"] = photos

            result.append(d)


        return jsonify(result), 200

    except Exception as e:
        print(f"⚠️ Erro ao listar visitas: {e}")
        return jsonify(error=str(e)), 500





@bp.route('/visits', methods=['POST'])
def create_visit():
    """
    Cria uma nova visita.
    Se 'generate_schedule' for True, gera automaticamente o cronograma fenológico
    com base na cultura e na data de plantio.
    """
    from datetime import date as _d, timedelta
    data = request.get_json(silent=True) or {}

    client_id = data.get('client_id')
    property_id = data.get('property_id')
    plot_id = data.get('plot_id')
    consultant_id = data.get('consultant_id')
    status = (data.get('status') or 'planned').strip().lower()
    gen_schedule = bool(data.get('generate_schedule'))
    culture = (data.get('culture') or '').strip() or None
    variety = (data.get('variety') or '').strip() or None
    date_str = data.get('date')
    recommendation = (data.get('recommendation') or '').strip()
    latitude = data.get('latitude')
    longitude = data.get('longitude')
    fenologia_real = data.get("fenologia_real")


    # ========================
    # ✅ VALIDAÇÕES BÁSICAS
    # ========================
    if not client_id:
        return jsonify(message="client_id é obrigatório"), 400
    if not Client.query.get(client_id):
        return jsonify(message="cliente não encontrado"), 404
    if property_id and not Property.query.get(property_id):
        return jsonify(message="propriedade não encontrada"), 404
    if plot_id and not Plot.query.get(plot_id):
        return jsonify(message="talhão não encontrado"), 404
    if consultant_id and not Consultant.query.get(int(consultant_id)):
        return jsonify(message="consultor não encontrado"), 404

    try:
        visit_date = _d.fromisoformat(date_str)
    except Exception:
        return jsonify(message="data inválida, esperado formato YYYY-MM-DD"), 400

    # ======================================================
    # 🌾 GERAÇÃO AUTOMÁTICA DO CRONOGRAMA FENOLÓGICO
    # ======================================================
    from models import PhenologyStage

    p = None
    if gen_schedule:
        if not culture or not variety:
            return jsonify(message="culture e variety são obrigatórios quando gerar cronograma"), 400

        # ✅ Cria o registro de plantio — mesmo que não haja talhão
        if not plot_id:
            p = Planting(
                plot_id=None,
                culture=culture,
                variety=variety,
                planting_date=visit_date
            )
            db.session.add(p)
            db.session.flush()
        else:
            p = Planting(
                plot_id=plot_id,
                culture=culture,
                variety=variety,
                planting_date=visit_date
            )
            db.session.add(p)
            db.session.flush()


        # Visita inicial (plantio)
        v0 = Visit(
            client_id=client_id,
            property_id=property_id or None,
            plot_id=plot_id or None,
            planting_id=p.id if p else None,
            consultant_id=consultant_id,
            date=visit_date,
            recommendation="Plantio",
            status=status,
            culture=culture,
            variety=variety,
            latitude=latitude,
            longitude=longitude
        )
        db.session.add(v0)

        # Gera visitas futuras conforme fenologia
        stages = PhenologyStage.query.filter_by(culture=culture).order_by(PhenologyStage.days.asc()).all()
        if culture.lower() == "soja":
            stages = [s for s in stages if "maturação fisiológica" not in s.name.lower()]

        for st in stages:
            if st.days == 0 or "plantio" in st.name.lower():
                continue

            fut_date = visit_date + timedelta(days=int(st.days))
            vv = Visit(
                client_id=client_id,  # garante o vínculo com o mesmo cliente
                property_id=property_id,  # idem para fazenda
                plot_id=plot_id,  # idem para talhão
                planting_id=p.id if p else None,
                consultant_id=consultant_id or v0.consultant_id,  # fallback seguro
                date=fut_date,
                recommendation=st.name.strip().capitalize(),
                status='planned',
                culture=culture,
                variety=variety,
                latitude=latitude,
                longitude=longitude
            )
            db.session.add(vv)


        db.session.commit()
        return jsonify(message="visita criada com cronograma", visit=v0.to_dict()), 201

    # ======================================================
    # 🌱 VISITA NORMAL (SEM CRONOGRAMA)
    # ======================================================
    v = Visit(
        client_id=client_id,
        property_id=property_id or None,
        plot_id=plot_id or None,
        consultant_id=consultant_id,
        date=visit_date,
        checklist=data.get('checklist'),
        diagnosis=data.get('diagnosis'),
        recommendation=recommendation,
        status=status,
        culture=culture,
        variety=variety,
        latitude=latitude,
        longitude=longitude,
        fenologia_real=fenologia_real
    )

    # 🌿 Preenche cultura e variedade a partir do plantio, se não vierem do frontend
    if not v.culture and v.plot_id:
        planting = Planting.query.filter_by(plot_id=v.plot_id).order_by(Planting.id.desc()).first()
        if planting:
            v.culture = planting.culture
            v.variety = planting.variety

    
    # ===========================================
    # 🌿 SALVAR PRODUTOS NA CRIAÇÃO DA VISITA
    # ===========================================
    db.session.add(v)
    db.session.commit()

    products = data.get("products", [])
    from models import VisitProduct

    for p in products:
        vp = VisitProduct(
            visit_id=v.id,
            product_name=p.get("product_name", ""),
            dose=p.get("dose", ""),
            unit=p.get("unit", ""),
            application_date=(
                datetime.strptime(p["application_date"], "%Y-%m-%d").date()
                if p.get("application_date")
                else None
            ),
        )
        db.session.add(vp)

    db.session.commit()
    return jsonify(message="visita criada", visit=v.to_dict()), 201






@bp.route('/visits/<int:visit_id>/pdf', methods=['GET'])
@cross_origin(origins=["https://agrocrm-frontend.onrender.com"])
def export_visit_pdf(visit_id):
    buffer, filename = build_visit_pdf_file(visit_id)
    return send_file(
        buffer,
        mimetype="application/pdf",
        as_attachment=True,
        download_name=filename
    )






@bp.route('/visits/bulk', methods=['POST'])
def create_visits_bulk():
    data = request.get_json(silent=True) or {}
    items = data.get('items', [])
    created = []

    for it in items:
        client_id = it.get('client_id')
        property_id = it.get('property_id')
        plot_id = it.get('plot_id')
        consultant_id = it.get('consultant_id')
        status = (it.get('status') or 'planned').strip().lower()

        if not (client_id and property_id and plot_id and it.get('date')):
            continue

        if not Client.query.get(client_id): continue
        if not Property.query.get(property_id): continue
        if not Plot.query.get(plot_id): continue
        if consultant_id and int(consultant_id) not in CONSULTANT_IDS: continue

        try:
            from datetime import date as _d
            visit_date = _d.fromisoformat(it['date'])
        except Exception:
            continue

        v = Visit(
            client_id=client_id,
            property_id=property_id,
            plot_id=plot_id,
            consultant_id=consultant_id,
            date=visit_date,
            checklist=None,
            diagnosis=None,
            recommendation=it.get('recommendation'),
            status=status
        )
        db.session.add(v)
        created.append(v)

    db.session.commit()
    return jsonify([v.to_dict() | {"status": v.status} for v in created]), 201



@bp.route('/visits/<int:visit_id>', methods=['PUT'])
def update_visit(visit_id: int):
    v = Visit.query.get(visit_id)
    if not v:
        return jsonify(message='visit not found'), 404

    data = request.get_json(silent=True) or {}
    print("📩 PAYLOAD RECEBIDO NO PUT:", data)

    # Campos simples
    for tf in ('checklist', 'diagnosis', 'fenologia_real'):
        if tf in data:
            setattr(v, tf, data[tf])

    if "recommendation" in data:
        rec = data.get("recommendation")
        if rec not in (None, "", " "):
            v.recommendation = rec.strip()

    # Só altera client_id se vier explicitamente e válido
    if 'client_id' in data:
        cid = data.get('client_id')
        if cid in (None, "", 0):
            v.client_id = v.client_id
        else:
            if not Client.query.get(cid):
                return jsonify(message='client not found'), 404
            v.client_id = cid

    # Só altera property_id se vier explicitamente
    if 'property_id' in data:
        pid = data.get('property_id')
        if pid in (None, "", 0):
            v.property_id = None
        else:
            if not Property.query.get(pid):
                return jsonify(message='property not found'), 404
            v.property_id = pid

    # Só altera plot_id se vier explicitamente
    if 'plot_id' in data:
        plid = data.get('plot_id')
        if plid in (None, "", 0):
            v.plot_id = None
        else:
            if not Plot.query.get(plid):
                return jsonify(message='plot not found'), 404
            v.plot_id = plid

    # Só altera consultant_id se vier explicitamente
    if 'consultant_id' in data:
        cid = data.get('consultant_id')
        if cid in (None, "", 0):
            v.consultant_id = None
        else:
            if int(cid) not in CONSULTANT_IDS:
                return jsonify(message='consultant not found'), 404
            v.consultant_id = int(cid)

    # Só altera culture se vier explicitamente
    if 'culture' in data:
        culture = (data.get('culture') or "").strip()
        v.culture = culture or None

    # Só altera variety se vier explicitamente
    if 'variety' in data:
        variety = (data.get('variety') or "").strip()
        v.variety = variety or None

    if "preserve_date" in data and data.get("preserve_date"):
        data["date"] = v.date.isoformat() if v.date else None

    if 'date' in data:
        if not data['date']:
            v.date = None
        else:
            try:
                from datetime import date as _d
                v.date = _d.fromisoformat(data['date'])
            except Exception:
                return jsonify(message='invalid date, expected YYYY-MM-DD'), 400

    if 'status' in data and data['status']:
        v.status = data['status'].strip().lower()

    if 'latitude' in data:
        v.latitude = parse_optional_float(data.get('latitude'))

    if 'longitude' in data:
        v.longitude = parse_optional_float(data.get('longitude'))

    print("🧠 Atualizando visita", visit_id, "com dados:", data)

    # ✅ PRODUTOS
    if "products" in data:
        from models import VisitProduct
        VisitProduct.query.filter_by(visit_id=visit_id).delete()
        for p in data["products"]:
            vp = VisitProduct(
                visit_id=visit_id,
                product_name=p.get("product_name", ""),
                dose=p.get("dose", ""),
                unit=p.get("unit", ""),
                application_date=(
                    datetime.strptime(p["application_date"], "%Y-%m-%d")
                    if p.get("application_date") else None
                ),
            )
            db.session.add(vp)

    db.session.commit()
    return jsonify(message='visit updated', visit=v.to_dict() | {"status": v.status}), 200




@bp.route('/visits/<int:visit_id>', methods=['DELETE'])
def delete_visit(visit_id):
    """
    Exclui uma visita. Se for a visita de plantio, remove também o plantio e TODAS
    as visitas geradas automaticamente (fenológicas) vinculadas ao mesmo plantio_id.
    """
    try:
        visit = Visit.query.get(visit_id)
        if not visit:
            print(f"⚠️ Visita {visit_id} não encontrada.")
            return jsonify({'error': 'Visita não encontrada'}), 404

        print(f"🗑 Solicitada exclusão da visita {visit_id}: {visit.recommendation}")

        # ✅ Detecção mais robusta de plantio
        rec = (visit.recommendation or "").lower()
        is_plantio = ("plantio" in rec) or (visit.planting_id is not None)

        if is_plantio:
            planting = Planting.query.get(visit.planting_id) if visit.planting_id else None

            # 1) Se tiver planting_id, apaga todas as visitas vinculadas (SEM filtro de data)
            if planting:
                linked_visits = Visit.query.filter(
                    Visit.planting_id == planting.id,
                    Visit.id != visit.id
                ).all()

                for lv in linked_visits:
                    print(f"   → Removendo visita vinculada {lv.id} ({lv.recommendation})")
                    db.session.delete(lv)

                print(f"🌾 Removendo plantio {planting.id} vinculado.")
                db.session.delete(planting)

            else:
                # 2) Fallback: sem planting_id, apaga "todas do mesmo contexto"
                #    (se você tiver plot_id, USE ELE! É o melhor vínculo do talhão)
                q = Visit.query.filter(
                    Visit.id != visit.id,
                    Visit.client_id == visit.client_id,
                    Visit.property_id == visit.property_id,
                    Visit.culture == visit.culture
                )

                # Se existir plot_id no seu model, habilite essa linha (recomendado):
                if getattr(visit, "plot_id", None):
                    q = q.filter(Visit.plot_id == visit.plot_id)

                linked_visits = q.all()

                for lv in linked_visits:
                    print(f"   → Removendo visita relacionada {lv.id} ({lv.recommendation})")
                    db.session.delete(lv)

            # 3) Por fim, remove a visita de plantio
            db.session.delete(visit)
            db.session.commit()
            print("✅ Plantio e visitas vinculadas excluídos com sucesso.")
            return jsonify({'message': 'Plantio e visitas vinculadas excluídos com sucesso'}), 200

        # Caso comum (visita isolada)
        print(f"🧾 Excluindo visita isolada {visit_id}")
        db.session.delete(visit)
        db.session.commit()
        print(f"✅ Visita {visit_id} excluída com sucesso.")
        return jsonify({'message': 'Visita excluída com sucesso'}), 200

    except Exception as e:
        print(f"❌ Erro interno ao excluir visita {visit_id}: {e}")
        db.session.rollback()
        return jsonify({'error': f'Erro interno ao excluir visita: {str(e)}'}), 500







# ==============================
# 📸 FOTOS — upload, legenda, exclusão (REVISADO)
# ==============================

@bp.route('/visits/<int:visit_id>/photos', methods=['POST', 'OPTIONS'])
@cross_origin(origins=["https://agrocrm-frontend.onrender.com"])
def upload_photos(visit_id):
    """Upload de múltiplas fotos com legendas (captions) — agora no Cloudflare R2."""
    visit = Visit.query.get_or_404(visit_id)

    # ✅ Modo leve por padrão (evita SIGKILL)
    # /pdf?mode=full -> tenta completo (ainda com limites)
    mode = (request.args.get("mode") or "lite").lower()

    # Limites seguros (você pode ajustar)
    MAX_VISITS   = int(request.args.get("max_visits") or (8 if mode == "lite" else 20))
    MAX_PHOTOS_V = int(request.args.get("max_photos") or (4 if mode == "lite" else 8))


    files = request.files.getlist('photos')
    captions = request.form.getlist('captions')

    if not files:
        return jsonify({"error": "Nenhum arquivo enviado"}), 400

    # ✅ R2 envs
    bucket = os.environ.get("R2_BUCKET")
    public_base = (os.environ.get("R2_PUBLIC_BASE_URL") or "").rstrip("/")

    if not bucket or not public_base:
        return jsonify({"error": "R2 não configurado: faltam variáveis de ambiente"}), 500

    r2 = get_r2_client()

    saved = []

    for i, file in enumerate(files):
        # 🔥 nome único
        unique = uuid.uuid4().hex
        original = secure_filename(file.filename or "foto.jpg")

        # (opcional) força extensão .jpg se vier sem
        if "." not in original:
            original = f"{original}.jpg"

        key = f"visits/{visit_id}/{unique}_{original}"

        # ✅ upload direto pro R2 (sem usar disco do Render)
        r2.upload_fileobj(
            Fileobj=file,
            Bucket=bucket,
            Key=key,
            ExtraArgs={
                "ContentType": file.mimetype or "image/jpeg",
                # Se quiser cache agressivo depois, dá pra adicionar CacheControl aqui
            },
        )

        caption = captions[i] if i < len(captions) else ""

        url = f"{public_base}/{key}"

        photo = Photo(
            visit_id=visit_id,
            url=url,
            caption=caption
        )
        db.session.add(photo)
        db.session.flush()

        saved.append({
            "id": photo.id,
            "url": url,
            "caption": caption or ""
        })

    db.session.commit()

    return jsonify({
        "message": f"{len(saved)} foto(s) salvas.",
        "photos": saved
    }), 201



@bp.route('/visits/<int:visit_id>/photos', methods=['GET'])
def list_photos(visit_id):
    visit = Visit.query.get_or_404(visit_id)

    photos = []
    for p in (visit.photos or []):
        photos.append({
            "id": p.id,
            "url": resolve_photo_url(p.url),  # ✅ devolve R2 se já for R2
            "caption": p.caption or ""
        })

    return jsonify(photos), 200




@bp.route('/photos/<int:photo_id>', methods=['PUT', 'OPTIONS'])
@cross_origin(origins=["https://agrocrm-frontend.onrender.com"])
def update_photo_caption(photo_id):
    """
    Atualiza a legenda de uma foto específica.
    """
    try:
        data = request.get_json() or {}
        caption = data.get("caption", "").strip()

        photo = Photo.query.get(photo_id)
        if not photo:
            return jsonify({"error": "Foto não encontrada"}), 404

        photo.caption = caption
        db.session.commit()  # ✅ garante persistência

        print(f"📝 Legenda atualizada -> Foto {photo_id}: {caption}")
        return jsonify({"success": True, "caption": caption}), 200

    except Exception as e:
        db.session.rollback()
        print(f"❌ Erro ao atualizar legenda da foto {photo_id}: {e}")
        return jsonify({"error": str(e)}), 500




@bp.route('/photos/<int:photo_id>', methods=['DELETE', 'OPTIONS'])
@cross_origin(origins=["https://agrocrm-frontend.onrender.com"])
def delete_single_photo(photo_id):
    """Exclui uma foto específica do banco e do disco."""
    photo = Photo.query.get_or_404(photo_id)
    try:
        from urllib.parse import urlparse
        parsed = urlparse(photo.url)
        filename = os.path.basename(parsed.path)
        file_path = os.path.join(UPLOAD_DIR, filename)

        if os.path.exists(file_path):
            os.remove(file_path)

        db.session.delete(photo)
        db.session.commit()
        return jsonify({"message": "Foto excluída com sucesso"}), 200
    except Exception as e:
        db.session.rollback()
        print(f"⚠️ Erro ao excluir foto: {e}")
        return jsonify({"error": f"Erro ao excluir foto: {e}"}), 500


@bp.route('/visits/<int:visit_id>/photos', methods=['DELETE'])
def delete_all_photos_of_visit(visit_id):
    """Exclui todas as fotos associadas a uma visita."""
    try:
        visit = Visit.query.get_or_404(visit_id)
        count = 0
        for photo in visit.photos:
            from urllib.parse import urlparse
            parsed = urlparse(photo.url)
            filename = os.path.basename(parsed.path)
            file_path = os.path.join(UPLOAD_DIR, filename)
            if os.path.exists(file_path):
                os.remove(file_path)
            db.session.delete(photo)
            count += 1

        db.session.commit()
        return jsonify({"message": f"{count} foto(s) excluídas com sucesso"}), 200
    except Exception as e:
        db.session.rollback()
        print(f"⚠️ Erro ao excluir fotos da visita: {e}")
        return jsonify({"error": str(e)}), 500

# ============================================================
# 🔗 Resolver URL pública da foto (R2 / legado)
# ============================================================
def resolve_photo_url(u: str) -> str:
    if not u:
        return ""

    # já é pública (R2/CDN)
    if u.startswith("http://") or u.startswith("https://"):
        return u

    # legado /uploads/... -> aponta pro backend
    if u.startswith("/uploads/"):
        backend_url = (os.environ.get("RENDER_EXTERNAL_URL") or "https://agrocrm-backend.onrender.com").rstrip("/")
        return f"{backend_url}{u}"

    return u




@bp.route("/visits/<int:visit_id>", methods=["GET"])
def get_visit(visit_id):
    v = Visit.query.get_or_404(visit_id)

    photos = []
    try:
        for p in (v.photos or []):
            photos.append({
                "id": p.id,
                "url": resolve_photo_url(p.url),
                "caption": p.caption or ""
            })
    except Exception:
        photos = []

    return jsonify({
        "id": v.id,
        "client_id": v.client_id,
        "property_id": v.property_id,
        "plot_id": v.plot_id,
        "consultant_id": v.consultant_id,
        "date": v.date.isoformat() if getattr(v, "date", None) else None,
        "recommendation": v.recommendation or "",
        "status": v.status or "planned",
        "fenologia_real": getattr(v, "fenologia_real", None),
        "latitude": getattr(v, "latitude", None),
        "longitude": getattr(v, "longitude", None),
        "products": [p.to_dict() for p in (getattr(v, "products", []) or [])],
        "photos": photos
    }), 200





@bp.route("/view/visit/<int:visit_id>", methods=["GET"])
def public_visit_view(visit_id):
    """🌿 Página pública de visualização de visita (NutriCRM Viewer)"""
    from models import Visit, Client, Property, Plot, Consultant

    # ================================
    # ✅ CORREÇÃO OBRIGATÓRIA AQUI
    # ================================
    backend_url = os.environ.get("RENDER_EXTERNAL_URL") or "https://agrocrm-backend.onrender.com"

    visit = Visit.query.get_or_404(visit_id)
    client = Client.query.get(visit.client_id)
    prop = Property.query.get(visit.property_id)
    plot = Plot.query.get(visit.plot_id)
    consultant = Consultant.query.get(visit.consultant_id) if hasattr(visit, "consultant_id") else None

    lat = getattr(plot, "latitude", None)
    lon = getattr(plot, "longitude", None)

    photos = []
    for p in (visit.photos or []):
        photos.append({
            "id": p.id,
            "url": resolve_photo_url(p.url),
            "caption": p.caption or ""
        })


    html_template = """
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
        <meta charset="UTF-8" />
        <meta name="viewport" content="width=device-width, initial-scale=1.0" />
        <title>Visita #{{ visit.id }} — NutriCRM</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css" rel="stylesheet" />
        <link rel="stylesheet" href="https://unpkg.com/leaflet@1.9.4/dist/leaflet.css" />
        <script src="https://unpkg.com/leaflet@1.9.4/dist/leaflet.js"></script>
        <style>
            body { background-color:#f9fafb; font-family:'Inter', sans-serif; padding-bottom:60px; }
            header { background:#1B5E20; color:white; padding:16px; display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; }
            header img { height:46px; object-fit:contain; }
            .info-card { background:white; border-radius:10px; box-shadow:0 2px 10px rgba(0,0,0,0.05); padding:20px; margin-top:20px; }
            .photos img { width:100%; border-radius:8px; cursor:pointer; transition:transform 0.2s; }
            .photos img:hover { transform:scale(1.02); }
            #map { height:300px; border-radius:10px; margin-top:15px; }
            footer { margin-top:40px; text-align:center; color:#888; }
            .download-btn { background:#2E7D32; color:white; padding:10px 18px; border-radius:6px; text-decoration:none; transition:opacity .2s; }
            .download-btn:hover { opacity:.85; }
            .lightbox { display:none; position:fixed; z-index:9999; top:0; left:0; width:100%; height:100%; background:rgba(0,0,0,0.9); justify-content:center; align-items:center; }
            .lightbox img { max-width:90%; max-height:90%; }
        </style>
    </head>
    <body>
        <header>
            <img src="/static/nutricrm_logo.png" alt="NutriCRM Logo" />
            <h4>Relatório Técnico — Visita #{{ visit.id }}</h4>
            <a class="download-btn" href="/api/visits/{{ visit.id }}/pdf" target="_blank">📄 Baixar PDF</a>
        </header>

        <main class="container">
            <div class="info-card">
                <h4>Informações Gerais</h4>
                <table class="table table-borderless mt-3">
                    <tr><th>Cliente:</th><td>{{ client.name if client else '-' }}</td></tr>
                    <tr><th>Fazenda:</th><td>{{ prop.name if prop else '-' }}</td></tr>
                    <tr><th>Talhão:</th><td>{{ plot.name if plot else '-' }}</td></tr>
                    <tr><th>Consultor:</th><td>{{ consultant.name if consultant else '-' }}</td></tr>
                    <tr><th>Data:</th><td>{{ visit.date.strftime('%d/%m/%Y') if visit.date else '-' }}</td></tr>
                    <tr><th>Status:</th><td>{{ visit.status }}</td></tr>
                </table>
                {% if lat and lon %}
                <div id="map"></div>
                {% endif %}
            </div>

            {% if visit.diagnosis %}
            <div class="info-card">
                <h4>Diagnóstico</h4>
                <p>{{ visit.diagnosis }}</p>
            </div>
            {% endif %}

            {% if visit.recommendation %}
            <div class="info-card">
                <h4>Recomendações Técnicas</h4>
                <p>{{ visit.recommendation }}</p>
            </div>
            {% endif %}

            {% if photos %}
            <div class="info-card photos">
                <h4>Fotos da Visita</h4>
                <div class="row mt-3">
                    {% for p in photos %}
                    <div class="col-md-6 mb-3">
                        <img src="{{ p.url }}" onclick="openLightbox('{{ p.url }}')" />
                    </div>
                    {% endfor %}
                </div>
            </div>
            {% endif %}
        </main>

        <div id="lightbox" class="lightbox" onclick="closeLightbox()">
            <img id="lightbox-img" src="">
        </div>

        <footer>
            <small>NutriCRM © 2025 — Relatório técnico automatizado</small>
        </footer>

        <script>
            function openLightbox(src) {
                document.getElementById('lightbox-img').src = src;
                document.getElementById('lightbox').style.display = 'flex';
            }
            function closeLightbox() {
                document.getElementById('lightbox').style.display = 'none';
            }

            {% if lat and lon %}
            const map = L.map('map').setView([{{ lat }}, {{ lon }}], 15);
            L.tileLayer('https://{s}.tile.openstreetmap.org/{z}/{x}/{y}.png', {
                attribution: '© OpenStreetMap contributors'
            }).addTo(map);
            L.marker([{{ lat }}, {{ lon }}]).addTo(map)
                .bindPopup("{{ plot.name if plot else 'Talhão' }}")
                .openPopup();
            {% endif %}
        </script>
    </body>
    </html>
    """

    return render_template_string(
        html_template,
        visit=visit,
        client=client,
        prop=prop,
        plot=plot,
        consultant=consultant,
        photos=photos,
        lat=lat,
        lon=lon
    )


@bp.route("/visits/<int:visit_id>/products", methods=["POST"])
def add_visit_product(visit_id):
    data = request.get_json()

    product = VisitProduct(
        visit_id=visit_id,
        product_name=data.get("product_name", "").strip(),
        dose=data.get("dose", "").strip(),
        unit=data.get("unit", "").strip(),
        application_date=datetime.strptime(data.get("application_date"), "%Y-%m-%d") if data.get("application_date") else None,
    )

    db.session.add(product)
    db.session.commit()

    return jsonify({"success": True, "product": product.to_dict()}), 201



@bp.route("/products/<int:product_id>", methods=["PUT"])
def update_visit_product(product_id):
    data = request.get_json()
    product = VisitProduct.query.get_or_404(product_id)

    product.product_name = data.get("product_name", product.product_name)
    product.dose = data.get("dose", product.dose)
    product.unit = data.get("unit", product.unit)
    product.application_date = (
        datetime.strptime(data["application_date"], "%Y-%m-%d")
        if data.get("application_date")
        else product.application_date
    )

    db.session.commit()
    return jsonify({"success": True, "product": product.to_dict()})



@bp.route("/products/<int:product_id>", methods=["DELETE"])
def delete_visit_product(product_id):
    product = VisitProduct.query.get_or_404(product_id)
    db.session.delete(product)
    db.session.commit()
    return jsonify({"success": True})





@bp.route('/phenology/schedule', methods=['GET'])
def get_phenology_schedule():
    """
    Retorna o cronograma fenológico real (do banco de dados)
    com base na tabela phenology_stage.
    """
    from datetime import datetime, timedelta

    culture = request.args.get("culture")
    planting_date = request.args.get("planting_date")

    if not culture or not planting_date:
        return jsonify({"error": "culture and planting_date required"}), 400

    try:
        planting_date = datetime.fromisoformat(planting_date).date()
    except ValueError:
        return jsonify({"error": "invalid planting_date format"}), 400

    # ✅ Agora busca diretamente da tabela correta
    stages = db.session.run(
        text("SELECT code, name, days FROM phenology_stage WHERE culture = :culture ORDER BY days"),
        {"culture": culture}
    ).fetchall()

    if not stages:
        print(f"⚠️ Nenhum estágio encontrado para {culture}.")
        return jsonify([]), 200

    events = []
    for s in stages:
        date = planting_date + timedelta(days=s.days)
        events.append({
            "stage": s.name,
            "code": s.code,
            "suggested_date": date.isoformat(),
        })

    print(f"✅ {len(events)} estágios retornados para {culture}.")
    return jsonify(events), 200



# ============================================================
# 🔧 TESTES E UTILITÁRIOS
# ============================================================


@bp.route('/status')
def status():
    return jsonify({
        "ok": True,
        "clients": Client.query.count(),
        "properties": Property.query.count(),
        "plots": Plot.query.count(),
        "visits": Visit.query.count(),
    }), 200



@bp.route('/hello', methods=['GET'])
def hello():
    return jsonify(message="Hello from Flask!")


@bp.route('/users', methods=['GET'])
def get_users():
    users = User.query.all()
    return jsonify([{'id': u.id, 'email': u.email} for u in users])


# ---- Clients CRUD -------------------------------------------------
from models import Client


@bp.route('/clients', methods=['GET'])
def get_clients():
    """Return list of clients."""
    clients = Client.query.order_by(Client.id.desc()).all()
    return jsonify([c.to_dict() for c in clients]), 200


@bp.route('/clients/<int:client_id>', methods=['GET'])
def get_client(client_id: int):
    client = Client.query.get(client_id)
    if not client:
        return jsonify(message='client not found'), 404
    return jsonify(client.to_dict()), 200


@bp.route('/clients', methods=['POST'])
def create_client():
    data = request.get_json() or {}
    name = data.get('name')
    if not name:
        return jsonify(message='name is required'), 400

    client = Client(
        name=name,
        document=data.get('document'),
        segment=data.get('segment'),
        vendor=data.get('vendor'),
    )
    db.session.add(client)
    db.session.commit()
    return jsonify(message='client created', client=client.to_dict()), 201


@bp.route('/clients/<int:client_id>', methods=['PUT'])
def update_client(client_id: int):
    client = Client.query.get(client_id)
    if not client:
        return jsonify(message='client not found'), 404
    data = request.get_json() or {}
    # Update allowed fields if present
    for field in ('name', 'document', 'segment', 'vendor'):
        if field in data:
            setattr(client, field, data.get(field))
    db.session.commit()
    return jsonify(message='client updated', client=client.to_dict()), 200


@bp.route('/clients/<int:client_id>', methods=['DELETE'])
def delete_client(client_id: int):
    client = Client.query.get(client_id)
    if not client:
        return jsonify(message='client not found'), 404
    db.session.delete(client)
    db.session.commit()
    return jsonify(message='client deleted'), 200


# ---- Properties CRUD -------------------------------------------------
from models import Property


@bp.route('/properties', methods=['GET'])
def get_properties():
    """List properties. Optional query param: client_id to filter by client."""
    client_id = request.args.get('client_id', type=int)
    q = Property.query
    if client_id:
        q = q.filter_by(client_id=client_id)
    props = q.order_by(Property.id.desc()).all()
    return jsonify([p.to_dict() for p in props]), 200


@bp.route('/properties/<int:prop_id>', methods=['GET'])
def get_property(prop_id: int):
    p = Property.query.get(prop_id)
    if not p:
        return jsonify(message='property not found'), 404
    return jsonify(p.to_dict()), 200


@bp.route('/properties', methods=['POST'])
def create_property():
    data = request.get_json() or {}

    client_id = data.get('client_id')
    name = (data.get('name') or '').strip()
    city_state = (data.get('city_state') or '').strip() or None
    area_ha = data.get('area_ha')
    latitude = parse_optional_float(data.get('latitude'))
    longitude = parse_optional_float(data.get('longitude'))

    if not client_id or not name:
        return jsonify(message='client_id and name are required'), 400

    client = Client.query.get(client_id)
    if not client:
        return jsonify(message='client not found'), 404

    try:
        prop = Property(
            client_id=int(client_id),
            name=name,
            city_state=city_state,
            area_ha=(float(area_ha) if area_ha not in (None, "") else None),
            latitude=latitude,
            longitude=longitude,
        )
        db.session.add(prop)
        db.session.commit()
        return jsonify(message='property created', property=prop.to_dict()), 201

    except Exception as e:
        db.session.rollback()
        return jsonify(message=str(e)), 500


@bp.route('/properties/<int:prop_id>', methods=['PUT'])
def update_property(prop_id: int):
    p = Property.query.get(prop_id)
    if not p:
        return jsonify(message='property not found'), 404

    data = request.get_json() or {}

    if 'client_id' in data and data.get('client_id') not in (None, ""):
        if not Client.query.get(data.get('client_id')):
            return jsonify(message='client not found'), 404
        p.client_id = int(data.get('client_id'))

    if 'name' in data:
        p.name = (data.get('name') or '').strip()

    if 'city_state' in data:
        p.city_state = (data.get('city_state') or '').strip() or None

    if 'area_ha' in data:
        p.area_ha = float(data.get('area_ha')) if data.get('area_ha') not in (None, "") else None

    if 'latitude' in data:
        p.latitude = parse_optional_float(data.get('latitude'))

    if 'longitude' in data:
        p.longitude = parse_optional_float(data.get('longitude'))

    try:
        db.session.commit()
        return jsonify(message='property updated', property=p.to_dict()), 200
    except Exception as e:
        db.session.rollback()
        return jsonify(message=str(e)), 500


@bp.route('/properties/<int:prop_id>', methods=['DELETE'])
def delete_property(prop_id: int):
    p = Property.query.get(prop_id)
    if not p:
        return jsonify(message='property not found'), 404
    db.session.delete(p)
    db.session.commit()
    return jsonify(message='property deleted'), 200


# ---- Plots (Talhões) CRUD -------------------------------------------
from models import Plot


@bp.route('/plots', methods=['GET'])
def get_plots():
    try:
        property_id = request.args.get('property_id', type=int)
        client_id = request.args.get('client_id', type=int)
        q = Plot.query
        if property_id:
            q = q.filter_by(property_id=property_id)
        if client_id:
            q = q.join(Property).filter(Property.client_id == client_id)
        plots = q.order_by(Plot.id.desc()).all()
        return jsonify([pl.to_dict() for pl in plots]), 200
    except Exception as e:
        print(f"❌ Erro em /plots: {e}")
        return jsonify(error=str(e)), 500



@bp.route('/plots/<int:plot_id>', methods=['GET'])
def get_plot(plot_id: int):
    pl = Plot.query.get(plot_id)
    if not pl:
        return jsonify(message='plot not found'), 404
    return jsonify(pl.to_dict()), 200


@bp.route('/plots', methods=['POST'])
def create_plot():
    data = request.get_json() or {}
    property_id = data.get('property_id')
    name = data.get('name')
    if not property_id or not name:
        return jsonify(message='property_id and name are required'), 400
    prop = Property.query.get(property_id)
    if not prop:
        return jsonify(message='property not found'), 404

    irrigated = data.get('irrigated')
    # normalize irrigated to boolean or None
    if irrigated is not None:
        if isinstance(irrigated, str):
            irrigated = irrigated.lower() in ('1', 'true', 'yes', 'sim')
        else:
            irrigated = bool(irrigated)

    pl = Plot(
        property_id=property_id,
        name=name,
        area_ha=(float(data.get('area_ha')) if data.get('area_ha') is not None else None),
        irrigated=irrigated,
    )
    db.session.add(pl)
    db.session.commit()
    return jsonify(message='plot created', plot=pl.to_dict()), 201


@bp.route('/plots/<int:plot_id>', methods=['PUT'])
def update_plot(plot_id: int):
    pl = Plot.query.get(plot_id)
    if not pl:
        return jsonify(message='plot not found'), 404
    data = request.get_json() or {}
    for field in ('name', 'area_ha', 'irrigated', 'property_id', 'latitude', 'longitude'):
        if field in data:
            if field == 'area_ha':
                setattr(pl, field, float(data.get(field)) if data.get(field) is not None else None)
            elif field == 'irrigated':
                v = data.get(field)
                if isinstance(v, str):
                    v = v.lower() in ('1', 'true', 'yes', 'sim')
                else:
                    v = bool(v)
                setattr(pl, field, v)
            else:
                setattr(pl, field, data.get(field))

    if 'property_id' in data and data.get('property_id') is not None:
        if not Property.query.get(data.get('property_id')):
            return jsonify(message='property not found'), 404

    db.session.commit()
    return jsonify(message='plot updated', plot=pl.to_dict()), 200


@bp.route('/plots/<int:plot_id>', methods=['DELETE'])
def delete_plot(plot_id: int):
    pl = Plot.query.get(plot_id)
    if not pl:
        return jsonify(message='plot not found'), 404
    db.session.delete(pl)
    db.session.commit()
    return jsonify(message='plot deleted'), 200


# ---- Plantings (Plantios) CRUD --------------------------------------
from models import Planting
from datetime import datetime


@bp.route('/plantings', methods=['GET'])
def get_plantings():
    """List plantings. Optional filters: plot_id, property_id, client_id"""
    plot_id = request.args.get('plot_id', type=int)
    property_id = request.args.get('property_id', type=int)
    client_id = request.args.get('client_id', type=int)

    q = Planting.query
    if plot_id:
        q = q.filter_by(plot_id=plot_id)
    if property_id:
        q = q.join(Plot).filter(Plot.property_id == property_id)
    if client_id:
        q = q.join(Plot).join(Property).filter(Property.client_id == client_id)

    items = q.order_by(Planting.id.desc()).all()
    return jsonify([it.to_dict() for it in items]), 200


@bp.route('/plantings/<int:pid>', methods=['GET'])
def get_planting(pid: int):
    p = Planting.query.get(pid)
    if not p:
        return jsonify(message='planting not found'), 404
    return jsonify(p.to_dict()), 200


@bp.route('/plantings', methods=['POST'])
def create_planting():
    data = request.get_json() or {}
    plot_id = data.get('plot_id')
    if not plot_id:
        return jsonify(message='plot_id is required'), 400

    plot = Plot.query.get(plot_id)
    if not plot:
        return jsonify(message='plot not found'), 404

    # cultura e variedade podem vir do select do frontend
    culture = data.get('culture')  # esperado: "Milho", "Soja", "Algodão" (case-insensitive ok)
    variety = data.get('variety')

    planting_date = None
    if data.get('planting_date'):
        try:
            planting_date = datetime.fromisoformat(data.get('planting_date')).date()
        except Exception:
            return jsonify(message='invalid planting_date, expected YYYY-MM-DD'), 400

    p = Planting(
        plot_id=plot_id,
        culture=culture,
        variety=variety,
        planting_date=planting_date,
    )
    db.session.add(p)
    db.session.flush()  # para ter p.id

    # === Gerar Visitas automáticas pela fenologia ===
    # Usa a tabela real phenology_stage do banco
    if planting_date and culture:
        from sqlalchemy import text

        # busca do banco conforme cultura
        stages = db.session.run(
            text("SELECT code, name, days FROM phenology_stage WHERE culture = :culture ORDER BY days"),
            {"culture": culture}
        ).fetchall()

        if stages:
            prop = Property.query.get(plot.property_id) if plot.property_id else None
            client = Client.query.get(prop.client_id) if (prop and prop.client_id) else None

            for st in stages:
                if st.days == 0:
                    continue  # ignora o plantio (já criado)

                visit_date = planting_date + datetime.timedelta(days=int(st.days))

                v = Visit(
                    client_id=(client.id if client else None),
                    property_id=(prop.id if prop else None),
                    plot_id=plot.id,
                    planting_id=p.id,
                    consultant_id=None,
                    date=visit_date,
                    checklist=None,
                    diagnosis=None,
                    recommendation=st.name,
                    culture=culture,
                    variety=variety,
                    status='planned'
                )
                db.session.add(v)

            print(f"✅ {len(stages)} visitas geradas para {culture}.")


    db.session.commit()
    return jsonify(message='planting created', planting=p.to_dict()), 201



@bp.route('/plantings/<int:pid>', methods=['PUT'])
def update_planting(pid: int):
    p = Planting.query.get(pid)
    if not p:
        return jsonify(message='planting not found'), 404
    data = request.get_json() or {}

    if 'plot_id' in data and data.get('plot_id') is not None:
        if not Plot.query.get(data.get('plot_id')):
            return jsonify(message='plot not found'), 404
        p.plot_id = data.get('plot_id')

    if 'culture' in data:
        p.culture = data.get('culture')
    if 'variety' in data:
        p.variety = data.get('variety')
    if 'planting_date' in data:
        if data.get('planting_date') in (None, ''):
            p.planting_date = None
        else:
            try:
                p.planting_date = datetime.fromisoformat(data.get('planting_date')).date()
            except Exception:
                return jsonify(message='invalid planting_date, expected YYYY-MM-DD'), 400

    db.session.commit()
    return jsonify(message='planting updated', planting=p.to_dict()), 200


@bp.route('/plantings/<int:pid>', methods=['DELETE'])
def delete_planting(pid: int):
    p = Planting.query.get(pid)
    if not p:
        return jsonify(message='planting not found'), 404
    db.session.delete(p)
    db.session.commit()
    return jsonify(message='planting deleted'), 200


# ---- Visits (Visitas) CRUD ------------------------------------------
from models import Visit
from datetime import date

# ---- Opportunities CRUD --------------------------------------------
from models import Opportunity


@bp.route('/opportunities', methods=['GET'])
def get_opportunities():
    """List opportunities. Optional filter: client_id"""
    client_id = request.args.get('client_id', type=int)
    q = Opportunity.query
    if client_id:
        q = q.filter_by(client_id=client_id)
    items = q.order_by(Opportunity.id.desc()).all()
    return jsonify([it.to_dict() for it in items]), 200


@bp.route('/opportunities/<int:oid>', methods=['GET'])
def get_opportunity(oid: int):
    o = Opportunity.query.get(oid)
    if not o:
        return jsonify(message='opportunity not found'), 404
    return jsonify(o.to_dict()), 200


@bp.route('/opportunities', methods=['POST'])
def create_opportunity():
    data = request.get_json() or {}
    client_id = data.get('client_id')
    if not client_id:
        return jsonify(message='client_id is required'), 400
    if not Client.query.get(client_id):
        return jsonify(message='client not found'), 404

    title = data.get('title')
    estimated_value = data.get('estimated_value')
    stage = data.get('stage') or 'prospecção'

    try:
        estimated_value = float(estimated_value) if estimated_value is not None else None
    except Exception:
        return jsonify(message='estimated_value must be a number'), 400

    o = Opportunity(
        client_id=client_id,
        title=title,
        estimated_value=estimated_value,
        stage=stage,
    )
    db.session.add(o)
    db.session.commit()
    return jsonify(message='opportunity created', opportunity=o.to_dict()), 201


@bp.route('/opportunities/<int:oid>', methods=['PUT'])
def update_opportunity(oid: int):
    o = Opportunity.query.get(oid)
    if not o:
        return jsonify(message='opportunity not found'), 404
    data = request.get_json() or {}
    if 'client_id' in data and data.get('client_id') is not None:
        if not Client.query.get(data.get('client_id')):
            return jsonify(message='client not found'), 404
        o.client_id = data.get('client_id')
    if 'title' in data:
        o.title = data.get('title')
    if 'estimated_value' in data:
        try:
            o.estimated_value = float(data.get('estimated_value')) if data.get('estimated_value') is not None else None
        except Exception:
            return jsonify(message='estimated_value must be a number'), 400
    if 'stage' in data:
        o.stage = data.get('stage') or 'prospecção'

    db.session.commit()
    return jsonify(message='opportunity updated', opportunity=o.to_dict()), 200


@bp.route('/opportunities/<int:oid>', methods=['DELETE'])
def delete_opportunity(oid: int):
    o = Opportunity.query.get(oid)
    if not o:
        return jsonify(message='opportunity not found'), 404
    db.session.delete(o)
    db.session.commit()
    return jsonify(message='opportunity deleted'), 200


@bp.route('/db-test', methods=['GET'])
def db_test():
    """Simple endpoint to test DB connectivity and return user count."""
    try:
        # run a trivial select to verify connection
        result = None
        with db.engine.connect() as conn:
            row = conn.run(text('SELECT 1')).fetchone()
            result = row[0] if row is not None else None

        # also give a user count
        user_count = User.query.count()
        return jsonify(status='ok', ping=result, users=user_count), 200
    except Exception as e:
        return jsonify(status='error', error=str(e)), 500


@bp.route('/users', methods=['POST'])
def create_user():
    """Create a new user. Expects JSON: {email, password}"""
    data = request.get_json() or {}
    email = data.get('email')
    password = data.get('password')
    if not email or not password:
        return jsonify(message='email and password are required'), 400

    if User.query.filter_by(email=email).first():
        return jsonify(message='user with that email already exists'), 409

    user = User(email=email)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    return jsonify(message='user created', id=user.id, email=user.email), 201


@bp.route('/login', methods=['POST'])
def login():
    data = request.get_json() or {}
    email = data.get('email')
    password = data.get('password')
    if not email or not password:
        return jsonify(message='email and password are required'), 400

    user = User.query.filter_by(email=email).first()
    if not user or not user.check_password(password):
        return jsonify(message='invalid credentials'), 401

    # create JWT
    secret = os.environ.get('JWT_SECRET', 'dev-secret')
    payload = {
        'sub': user.id,
        'email': user.email,
        'exp': datetime.datetime.utcnow() + datetime.timedelta(hours=24)
    }
    token = jwt.encode(payload, secret, algorithm='HS256')
    return jsonify(message='ok', token=token), 200


def _get_current_user_from_token(req):
    auth = req.headers.get('Authorization', '')
    if not auth.startswith('Bearer '):
        return None, 'missing token'
    token = auth.split(' ', 1)[1]
    secret = os.environ.get('JWT_SECRET', 'dev-secret')
    try:
        payload = jwt.decode(token, secret, algorithms=['HS256'])
    except jwt.ExpiredSignatureError:
        return None, 'token expired'
    except jwt.InvalidTokenError:
        return None, 'invalid token'

    user = User.query.get(payload.get('sub'))
    if not user:
        return None, 'user not found'
    return user, None


@bp.route('/me', methods=['GET'])
def me():
    user, err = _get_current_user_from_token(request)
    if err:
        return jsonify(message=err), 401
    return jsonify(id=user.id, email=user.email), 200

